import streamlit as st
import pandas as pd
import io
from math import ceil

from streamlit_app.utils.cmi_helpers import linea_color
from streamlit_app.utils.formatting import meta_his_signo, ejecucion_his_signo
from streamlit_app.components.ui.pagination import PaginationManager

# ── Columnas relevantes para el export a Excel ───────────────────────────────
_EXPORT_COLS = {
    "Id": "Código",
    "Indicador": "Indicador",
    "Linea": "Línea Estratégica",
    "Objetivo": "Objetivo",
    "Meta": "Meta",
    "Ejecucion": "Ejecución",
    "cumplimiento_pct": "Cumplimiento (%)",
    "Nivel de cumplimiento": "Estado",
    "Proceso": "Proceso",
    "Periodicidad": "Periodicidad",
    "Anio": "Año",
}

_BADGE_COLORS = {
    "Peligro": ("#B71C1C", "#FEE2E2"),
    "Alerta": ("#B45309", "#FEF3C7"),
    "Cumplimiento": ("#166534", "#DCFCE7"),
    "Sobrecumplimiento": ("#1D4ED8", "#DBEAFE"),
    "Pendiente de reporte": ("#475569", "#F1F5F9"),
}

_ROWS_PER_PAGE_OPTIONS = [25, 50, 100]


def _nivel_badge_html(nivel: str) -> str:
    tc, bg = _BADGE_COLORS.get(nivel, ("#334155", "#F8FAFC"))
    return (
        f"<span style='background:{bg};color:{tc};padding:3px 10px;"
        f"border-radius:999px;font-size:0.72rem;font-weight:700;"
        f"white-space:nowrap;display:inline-flex;align-items:center;gap:6px'>"
        f"<span style='width:8px;height:8px;border-radius:50%;background:{tc};display:inline-block'></span>"
        f"{nivel}</span>"
    )


def _linea_pill_html(linea: str) -> str:
    color = linea_color(linea)
    return (
        f"<span style='background:{color}22;color:{color};border:1.5px solid {color};"
        f"padding:3px 9px;border-radius:999px;font-size:0.72rem;font-weight:700;"
        f"white-space:nowrap;display:inline-block;max-width:180px;"
        f"overflow:hidden;text-overflow:ellipsis' title='{linea}'>{linea}</span>"
    )


def _build_excel(df: pd.DataFrame) -> bytes:
    """Genera Excel con columnas relevantes, encabezados amigables y formato visual."""
    present_cols = {k: v for k, v in _EXPORT_COLS.items() if k in df.columns}
    df_exp = df[list(present_cols.keys())].copy().rename(columns=present_cols)

    if "Cumplimiento (%)" in df_exp.columns:
        df_exp["Cumplimiento (%)"] = (
            pd.to_numeric(df_exp["Cumplimiento (%)"], errors="coerce").round(1)
        )

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_exp.to_excel(writer, sheet_name="Indicadores CMI", index=False)
        ws = writer.sheets["Indicadores CMI"]

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        thin = Side(style="thin", color="D1D5DB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Encabezado
        hdr_fill = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
        hdr_font = Font(color="FFFFFF", bold=True, size=11)
        for col_idx in range(1, len(df_exp.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 35

        # Anchos de columna
        col_widths = {"Indicador": 48, "Objetivo": 36, "Línea Estratégica": 26}
        for col_idx, col_name in enumerate(df_exp.columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 18)

        # Filas de datos con color por Estado
        status_colors = {
            "Sobrecumplimiento": "DBEAFE",
            "Cumplimiento": "DCFCE7",
            "Alerta": "FEF3C7",
            "Peligro": "FEE2E2",
        }
        estado_col_idx = next(
            (i + 1 for i, c in enumerate(df_exp.columns) if c == "Estado"), None
        )
        for row_idx in range(2, len(df_exp) + 2):
            for col_idx in range(1, len(df_exp.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                cell.border = border
            if estado_col_idx:
                estado_val = ws.cell(row=row_idx, column=estado_col_idx).value
                fill_color = status_colors.get(str(estado_val), None)
                if fill_color:
                    ws.cell(row=row_idx, column=estado_col_idx).fill = PatternFill(
                        start_color=fill_color, end_color=fill_color, fill_type="solid"
                    )

        ws.freeze_panes = "A2"

    return buf.getvalue()




def _render_row(row: pd.Series, idx: int, global_idx: int) -> None:
    """Renderiza una fila de la tabla paginada."""
    st.markdown("<div class='tl-sep'>", unsafe_allow_html=True)
    cols = st.columns([0.55, 3.8, 2.1, 2.8, 1.0, 1.0, 1.3, 2.1, 1.4])

    indicador = str(row.get("Indicador", "") or "—")
    linea = str(row.get("Linea", "") or "—")
    objetivo = str(row.get("Objetivo", "") or "—")
    meta_raw = row.get("Meta")
    ejec_raw = row.get("Ejecucion")
    cump = pd.to_numeric(
        row.get("cumplimiento_pct")
        if "cumplimiento_pct" in row.index else row.get("Cumplimiento_pct")
        if "Cumplimiento_pct" in row.index else row.get("cumplimiento")
        if "cumplimiento" in row.index else row.get("Cumplimiento"),
        errors="coerce",
    )
    nivel = str(row.get("Nivel de cumplimiento", "—") or "—")

    # Formateo canónico: respeta signo (%, $, ENT, DEC…) y decimales de cada indicador.
    # Los nombres de columna Meta y Ejecucion NO se modifican.
    def _safe_fmt_meta(row_d: dict) -> str:
        v = row_d.get("Meta")
        if v is None or str(v).strip() in ("", "nan", "None"):
            return "—"
        try:
            if pd.isna(float(str(v).strip())):
                return "—"
        except (ValueError, TypeError):
            pass
        return meta_his_signo(row_d) or "—"

    def _safe_fmt_ejec(row_d: dict) -> str:
        v = row_d.get("Ejecucion")
        if v is None or str(v).strip() in ("", "nan", "None"):
            return "—"
        try:
            if pd.isna(float(str(v).strip())):
                return "—"
        except (ValueError, TypeError):
            pass
        return ejecucion_his_signo(row_d) or "—"

    row_dict = row.to_dict()
    meta_fmt = _safe_fmt_meta(row_dict)
    ejec_fmt = _safe_fmt_ejec(row_dict)

    cols[0].markdown(
        f"<div class='tl-num' style='padding-top:5px'>{row.get('Id', '')}</div>",
        unsafe_allow_html=True,
    )
    cols[1].markdown(
        f"<div class='tl-ind' title='{indicador}'>{indicador}</div>",
        unsafe_allow_html=True,
    )
    cols[2].markdown(_linea_pill_html(linea), unsafe_allow_html=True)
    cols[3].markdown(
        f"<div class='tl-obj' title='{objetivo}'>{objetivo}</div>",
        unsafe_allow_html=True,
    )
    cols[4].markdown(f"<div class='tl-num'>{meta_fmt}</div>", unsafe_allow_html=True)
    cols[5].markdown(f"<div class='tl-num'>{ejec_fmt}</div>", unsafe_allow_html=True)
    cols[6].markdown(
        f"<div class='tl-num'>{f'{cump:.1f}%' if pd.notna(cump) else '—'}</div>",
        unsafe_allow_html=True,
    )
    cols[7].markdown(_nivel_badge_html(nivel), unsafe_allow_html=True)

    bkey = f"vf_{global_idx}_{row.get('Id', '')}_{abs(hash(indicador)) % 99999}"
    if cols[8].button("Ver ficha", key=bkey, use_container_width=True):
        from streamlit_app.components.cmi_tabs.modal_ficha import render_modal_ficha
        render_modal_ficha(row)

    st.markdown("</div>", unsafe_allow_html=True)


def render_tab_listado(df: pd.DataFrame):
    """Renderiza el tab Listado de Indicadores con paginación, filtros locales y export."""
    if df.empty:
        st.info("No hay datos para mostrar.")
        return

    # ── CSS de tabla responsive ───────────────────────────────────────────────
    st.markdown(
        """
        <style>
        .tl-ind { font-size:0.84rem;font-weight:600;color:#0F172A;line-height:1.4;
                  overflow:hidden;display:-webkit-box;-webkit-line-clamp:2;
                  -webkit-box-orient:vertical; }
        .tl-obj { font-size:0.77rem;color:#475569;overflow:hidden;
                  display:-webkit-box;-webkit-line-clamp:2;-webkit-box-orient:vertical; }
        .tl-num { font-size:0.83rem;color:#0F172A;font-variant-numeric:tabular-nums; }
        .tl-hdr { font-size:0.71rem;font-weight:700;color:#64748B;
                  text-transform:uppercase;letter-spacing:.06em;
                  border-bottom:2px solid #1A3A5C;padding-bottom:4px; }
        .tl-sep { border-top:1px solid #E2E8F0;padding-top:5px;padding-bottom:5px; }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.markdown("#### Listado Completo de Indicadores")

    # ── Filtros locales ───────────────────────────────────────────────────────
    fc1, fc2, fc3, fc4, fc5 = st.columns([2, 2, 3, 2, 1.5])
    with fc1:
        lineas_disp = (
            sorted(df["Linea"].dropna().astype(str).unique().tolist())
            if "Linea" in df.columns else []
        )
        linea_sel = st.selectbox("Línea estratégica", ["Todas"] + lineas_disp, key="tl_linea")
    with fc2:
        df_for_obj = df if linea_sel == "Todas" else df[df["Linea"] == linea_sel]
        objetivos_disp = (
            sorted(df_for_obj["Objetivo"].dropna().astype(str).unique().tolist())
            if "Objetivo" in df_for_obj.columns else []
        )
        objetivo_sel = st.selectbox("Objetivo estratégico", ["Todos"] + objetivos_disp, key="tl_obj")
    with fc3:
        search = st.text_input(
            "Buscar indicador", placeholder="Texto en nombre…",
            key="tl_search", label_visibility="visible",
        )
    with fc4:
        estados_disp = (
            ["Todos"] + sorted(df["Nivel de cumplimiento"].dropna().unique().tolist())
            if "Nivel de cumplimiento" in df.columns else ["Todos"]
        )
        sel_estado = st.selectbox("Estado", estados_disp, key="tl_estado")
    with fc5:
        rpp = st.selectbox("Filas/pág.", _ROWS_PER_PAGE_OPTIONS, index=0, key="tl_rpp")

    # ── Aplicar filtros ───────────────────────────────────────────────────────
    df_f = df.copy()
    if linea_sel != "Todas" and "Linea" in df_f.columns:
        df_f = df_f[df_f["Linea"] == linea_sel]
    if objetivo_sel != "Todos" and "Objetivo" in df_f.columns:
        df_f = df_f[df_f["Objetivo"] == objetivo_sel]
    if search:
        df_f = df_f[df_f["Indicador"].astype(str).str.contains(search, case=False, na=False)]
    if sel_estado != "Todos" and "Nivel de cumplimiento" in df_f.columns:
        df_f = df_f[df_f["Nivel de cumplimiento"] == sel_estado]

    if "cumplimiento_pct" in df_f.columns:
        df_f = df_f.assign(
            _cump_sort=pd.to_numeric(df_f["cumplimiento_pct"], errors="coerce")
        ).sort_values("_cump_sort", ascending=False).drop(columns="_cump_sort")

    if df_f.empty:
        st.warning("No hay indicadores que coincidan con los filtros seleccionados.")
        return

    # ── Inicializar PaginationManager ─────────────────────────────────────────
    pm = PaginationManager(
        df=df_f,
        key_prefix="tl_table",
        rows_per_page=rpp,
    )
    total, page, df_page = pm.paginate()
    total_pages = max(1, ceil(total / rpp))
    start_idx = (page - 1) * rpp
    end_idx = min(start_idx + rpp, total)

    # ── Barra superior: conteo + export ──────────────────────────────────────
    bar1, _, bar3 = st.columns([3, 2, 1.4])
    with bar1:
        st.markdown(
            f"<span style='font-size:0.82rem;color:#64748B'>"
            f"<b>{total}</b> indicadores · Selecciona <b>Ver ficha</b> para abrir la ficha técnica</span>",
            unsafe_allow_html=True,
        )
    with bar3:
        st.download_button(
            label="📥 Exportar Excel",
            data=_build_excel(df_f),
            file_name="indicadores_cmi.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    # ── Encabezado de columnas ────────────────────────────────────────────────
    hdr_cols = st.columns([0.55, 3.8, 2.1, 2.8, 1.0, 1.0, 1.3, 2.1, 1.4])
    for col, label in zip(
        hdr_cols,
        ["ID", "Indicador", "Línea", "Objetivo", "Meta", "Ejec.", "Cump. %", "Estado", "Acción"],
    ):
        col.markdown(f"<div class='tl-hdr'>{label}</div>", unsafe_allow_html=True)

    st.markdown("<div style='height:2px'></div>", unsafe_allow_html=True)

    # ── Renderizar filas de la página ─────────────────────────────────────────
    pm.render_rows(df_page, _render_row)

    # ── Controles de paginación ───────────────────────────────────────────────
    pm.render_controls(total, page, total_pages, start_idx, end_idx, position="bottom")


