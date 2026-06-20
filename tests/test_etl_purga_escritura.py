"""
tests/test_etl_purga_escritura.py
Cobertura de scripts/etl/escritura.py y scripts/etl/purga.py.

Cubre funciones puras e in-memory openpyxl:
  escritura.py:
    - _ejec_score
    - llaves_de_df
    - get_last_data_row
    - deduplicar_sheet (con _build_col_map mocked)
  purga.py:
    - limpiar_cierres_existentes (usa row[5] hardcoded)
    - purgar_filas_invalidas (con _build_col_map mocked)
"""
from unittest.mock import patch

import openpyxl
import pandas as pd
import pytest


# ─────────────────────────────────────────────────────────────────────────────
# _ejec_score — función pura
# ─────────────────────────────────────────────────────────────────────────────

from scripts.etl.escritura import _ejec_score


class TestEjecScore:
    def test_none_retorna_0(self):
        assert _ejec_score(None) == 0

    def test_cero_retorna_1(self):
        assert _ejec_score(0.0) == 1

    def test_cero_entero_retorna_1(self):
        assert _ejec_score(0) == 1

    def test_valor_positivo_retorna_2(self):
        assert _ejec_score(85.0) == 2

    def test_valor_negativo_retorna_2(self):
        assert _ejec_score(-5.0) == 2

    def test_string_numerico_no_cero_retorna_2(self):
        assert _ejec_score("85.5") == 2

    def test_string_cero_retorna_1(self):
        assert _ejec_score("0.0") == 1

    def test_string_vacio_retorna_0(self):
        assert _ejec_score("") == 0

    def test_string_nan_float_convierte_a_2(self):
        # float("nan") != 0.0 → score 2 (función usa float() antes de revisar "nan")
        assert _ejec_score("nan") == 2

    def test_string_no_numerico_no_vacio_retorna_1(self):
        assert _ejec_score("NA") == 1


# ─────────────────────────────────────────────────────────────────────────────
# llaves_de_df — función pura sobre DataFrame
# ─────────────────────────────────────────────────────────────────────────────

from scripts.etl.escritura import llaves_de_df


class TestLlavesDeDf:
    def test_genera_llave_correcta(self):
        df = pd.DataFrame({
            "Id": ["10"],
            "Fecha": ["2025-06-30"],
        })
        llaves = llaves_de_df(df)
        assert "10-2025-06-30" in llaves

    def test_df_vacio_retorna_set_vacio(self):
        df = pd.DataFrame(columns=["Id", "Fecha"])
        assert llaves_de_df(df) == set()

    def test_fecha_invalida_se_descarta(self):
        df = pd.DataFrame({
            "Id": ["10", "11"],
            "Fecha": [None, "2025-06-30"],
        })
        llaves = llaves_de_df(df)
        assert len(llaves) == 1
        assert "11-2025-06-30" in llaves

    def test_columna_id_faltante_retorna_vacio(self):
        df = pd.DataFrame({"Fecha": ["2025-06-30"]})
        assert llaves_de_df(df) == set()

    def test_id_float_se_normaliza(self):
        df = pd.DataFrame({
            "Id": [68.0],
            "Fecha": ["2025-06-30"],
        })
        llaves = llaves_de_df(df)
        assert "68-2025-06-30" in llaves

    def test_multiples_registros(self):
        df = pd.DataFrame({
            "Id": ["10", "10"],
            "Fecha": ["2025-06-30", "2025-12-31"],
        })
        llaves = llaves_de_df(df)
        assert len(llaves) == 2


# ─────────────────────────────────────────────────────────────────────────────
# get_last_data_row — openpyxl in-memory
# ─────────────────────────────────────────────────────────────────────────────

from scripts.etl.escritura import get_last_data_row


def _ws_con_filas(n_filas: int):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Id"])  # header row
    for i in range(1, n_filas + 1):
        ws.append([str(i)])
    return ws


class TestGetLastDataRow:
    def test_hoja_con_datos_retorna_ultima_fila(self):
        ws = _ws_con_filas(3)
        assert get_last_data_row(ws) == 4  # header + 3 rows

    def test_hoja_solo_header_retorna_1(self):
        ws = _ws_con_filas(0)
        assert get_last_data_row(ws) == 1

    def test_fila_vacia_en_medio_no_cuenta(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Id"])         # row 1: header
        ws.append(["10"])         # row 2: data
        ws.append([None])         # row 3: empty
        ws.append(["20"])         # row 4: data
        assert get_last_data_row(ws) == 4


# ─────────────────────────────────────────────────────────────────────────────
# deduplicar_sheet — openpyxl + mock _build_col_map
# ─────────────────────────────────────────────────────────────────────────────

from scripts.etl.escritura import deduplicar_sheet


def _ws_dedup():
    """Worksheet con 2 filas duplicadas para el mismo Id+Fecha."""
    wb = openpyxl.Workbook()
    ws = wb.active
    # Headers: Id(1), B(2), C(3), D(4), E(5), Fecha(6), ..., Ejecucion(11)
    ws.append(["Id", "B", "C", "D", "E", "Fecha", "G", "H", "I", "J", "Ejecucion"])
    # Fila 2: ejec=0 (peor)
    ws.append(["10", None, None, None, None, "2025-06-30",
               None, None, None, None, 0])
    # Fila 3: ejec=85 (mejor)
    ws.append(["10", None, None, None, None, "2025-06-30",
               None, None, None, None, 85])
    return ws


class TestDeduplicarSheet:
    @patch("scripts.etl.escritura._build_col_map")
    def test_elimina_duplicado_peor(self, mock_map):
        mock_map.return_value = {"Fecha": 6, "Ejecucion": 11}
        ws = _ws_dedup()
        n = deduplicar_sheet(ws)
        assert n == 1  # one row removed
        # Remaining rows: header + 1 data row
        data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
        assert len(data_rows) == 1

    @patch("scripts.etl.escritura._build_col_map")
    def test_sin_duplicados_no_elimina(self, mock_map):
        mock_map.return_value = {"Fecha": 6, "Ejecucion": 11}
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Id", "B", "C", "D", "E", "Fecha",
                   "G", "H", "I", "J", "Ejecucion"])
        ws.append(["10", None, None, None, None, "2025-06-30",
                   None, None, None, None, 85])
        ws.append(["11", None, None, None, None, "2025-06-30",
                   None, None, None, None, 90])
        n = deduplicar_sheet(ws)
        assert n == 0

    @patch("scripts.etl.escritura._build_col_map")
    def test_hoja_vacia_retorna_0(self, mock_map):
        mock_map.return_value = {"Fecha": 6, "Ejecucion": 11}
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Id", "Fecha", "Ejecucion"])
        n = deduplicar_sheet(ws)
        assert n == 0


# ─────────────────────────────────────────────────────────────────────────────
# limpiar_cierres_existentes — openpyxl in-memory (usa row[5] hardcoded)
# ─────────────────────────────────────────────────────────────────────────────

from scripts.etl.purga import limpiar_cierres_existentes
from scripts.etl.config import AÑO_CIERRE_ACTUAL


def _ws_cierres(rows):
    """
    Crea worksheet donde:
      - col A (idx 0) = Id
      - col F (idx 5) = Fecha (posición hardcoded en limpiar_cierres_existentes)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consolidado Cierres"
    # Headers: A=Id, B-E = relleno, F=Fecha
    ws.append(["Id", "B", "C", "D", "E", "Fecha"])
    for row in rows:
        ws.append(row)
    return ws


class TestLimpiarCierresExistentes:
    def test_conserva_diciembre_y_borra_junio(self):
        año = AÑO_CIERRE_ACTUAL - 1  # historical year
        ws = _ws_cierres([
            ["10", None, None, None, None, f"{año}-06-30"],
            ["10", None, None, None, None, f"{año}-12-31"],
        ])
        n = limpiar_cierres_existentes(ws)
        assert n == 1
        rows_left = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
        assert len(rows_left) == 1
        # The remaining row should be December
        fecha_restante = pd.to_datetime(rows_left[0][5])
        assert fecha_restante.month == 12

    def test_sin_diciembre_conserva_unico_registro(self):
        año = AÑO_CIERRE_ACTUAL - 1
        ws = _ws_cierres([
            ["10", None, None, None, None, f"{año}-06-30"],
        ])
        n = limpiar_cierres_existentes(ws)
        assert n == 0
        rows_left = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
        assert len(rows_left) == 1

    def test_año_futuro_conserva_todos(self):
        año_futuro = AÑO_CIERRE_ACTUAL + 1
        ws = _ws_cierres([
            ["10", None, None, None, None, f"{año_futuro}-03-31"],
            ["10", None, None, None, None, f"{año_futuro}-06-30"],
        ])
        n = limpiar_cierres_existentes(ws)
        # Future years are all kept
        rows_left = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
        assert len(rows_left) == 2

    def test_hoja_vacia_retorna_0(self):
        ws = _ws_cierres([])
        n = limpiar_cierres_existentes(ws)
        assert n == 0

    def test_ids_distintos_conservan_su_diciembre(self):
        año = AÑO_CIERRE_ACTUAL - 1
        ws = _ws_cierres([
            ["10", None, None, None, None, f"{año}-06-30"],
            ["10", None, None, None, None, f"{año}-12-31"],
            ["11", None, None, None, None, f"{año}-06-30"],
            ["11", None, None, None, None, f"{año}-12-31"],
        ])
        n = limpiar_cierres_existentes(ws)
        assert n == 2  # 2 Junes removed (one per Id)
        rows_left = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
        assert len(rows_left) == 2


# ─────────────────────────────────────────────────────────────────────────────
# purgar_filas_invalidas — openpyxl + mock _build_col_map
# ─────────────────────────────────────────────────────────────────────────────

from scripts.etl.purga import purgar_filas_invalidas


def _ws_purga(rows):
    """
    Worksheet para purgar_filas_invalidas.
    Con _build_col_map mocked → {"Id": 1, "Fecha": 2, "Anio": 3}
    col A=Id (idx 0), col B=Fecha (idx 1), col C=Anio (idx 2)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Id", "Fecha", "Anio"])  # headers
    for row in rows:
        ws.append(row)
    return ws


class TestPurgarFilasInvalidas:
    @patch("scripts.etl.purga._build_col_map")
    def test_borra_fila_con_año_futuro(self, mock_map):
        mock_map.return_value = {"Id": 1, "Fecha": 2, "Anio": 3}
        año_futuro = AÑO_CIERRE_ACTUAL + 1
        ws = _ws_purga([
            ["10", f"{año_futuro}-06-30", año_futuro],
        ])
        n = purgar_filas_invalidas(ws)
        assert n >= 1
        rows_left = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
        assert len(rows_left) == 0

    @patch("scripts.etl.purga._build_col_map")
    def test_conserva_fila_con_año_valido(self, mock_map):
        mock_map.return_value = {"Id": 1, "Fecha": 2, "Anio": 3}
        año = AÑO_CIERRE_ACTUAL - 1
        ws = _ws_purga([
            ["10", f"{año}-12-31", año],
        ])
        n = purgar_filas_invalidas(ws)
        assert n == 0
        rows_left = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
        assert len(rows_left) == 1

    @patch("scripts.etl.purga._build_col_map")
    def test_kawak_validos_filtra_id_ausente(self, mock_map):
        mock_map.return_value = {"Id": 1, "Fecha": 2, "Anio": 3}
        año = AÑO_CIERRE_ACTUAL - 1
        ws = _ws_purga([
            ["10", f"{año}-12-31", año],
        ])
        n = purgar_filas_invalidas(ws, kawak_validos={("99", año)})
        assert n >= 1
        rows_left = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
        assert len(rows_left) == 0

    @patch("scripts.etl.purga._build_col_map")
    def test_kawak_validos_permite_id_presente(self, mock_map):
        mock_map.return_value = {"Id": 1, "Fecha": 2, "Anio": 3}
        año = AÑO_CIERRE_ACTUAL - 1
        ws = _ws_purga([
            ["10", f"{año}-12-31", año],
        ])
        n = purgar_filas_invalidas(ws, kawak_validos={("10", año)})
        assert n == 0

    @patch("scripts.etl.purga._build_col_map")
    def test_hoja_vacia_retorna_0(self, mock_map):
        mock_map.return_value = {"Id": 1, "Fecha": 2, "Anio": 3}
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Id", "Fecha", "Anio"])
        n = purgar_filas_invalidas(ws)
        assert n == 0
