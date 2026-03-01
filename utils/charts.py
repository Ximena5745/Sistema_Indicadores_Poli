"""
utils/charts.py — Componentes de gráficos reutilizables.
"""
import io
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Font

from config import COLORES, COLOR_CATEGORIA

COLOR_CAT = {
    "Peligro":           "#D32F2F",
    "Alerta":            "#FBAF17",
    "Cumplimiento":      "#43A047",
    "Sobrecumplimiento": "#1A3A5C",
    "Sin dato":          "#BDBDBD",
}


def grafico_historico_indicador(df_ind: pd.DataFrame, titulo: str = "") -> go.Figure:
    """
    Gráfico de línea con zonas de color para el histórico de un indicador.
    df_ind debe tener: Cumplimiento_norm, Categoria, Periodo, Fecha.
    """
    if df_ind.empty:
        fig = go.Figure()
        fig.update_layout(title="Sin datos disponibles")
        return fig

    df_ind = df_ind.sort_values("Fecha").copy()
    cum_pct = (df_ind["Cumplimiento_norm"] * 100).round(1)
    y_max = max(130, float(cum_pct.max()) + 15) if not cum_pct.empty else 130

    colores_pts = df_ind["Categoria"].map(COLOR_CAT).fillna("#9E9E9E")

    fig = go.Figure()

    # Zonas de fondo
    fig.add_hrect(
        y0=0, y1=80, fillcolor="#FFCDD2", opacity=0.45, line_width=0,
        annotation_text="Peligro < 80%", annotation_position="top left",
        annotation_font_size=10,
    )
    fig.add_hrect(
        y0=80, y1=100, fillcolor="#FFF9C4", opacity=0.45, line_width=0,
        annotation_text="Alerta 80–100%", annotation_position="top left",
        annotation_font_size=10,
    )
    fig.add_hrect(
        y0=100, y1=105, fillcolor="#E8F5E9", opacity=0.50, line_width=0,
        annotation_text="Cumplimiento 100–105%", annotation_position="top left",
        annotation_font_size=10,
    )
    fig.add_hrect(
        y0=105, y1=y_max, fillcolor="#D0E4FF", opacity=0.45, line_width=0,
        annotation_text="Sobrecumplimiento > 105%", annotation_position="top left",
        annotation_font_size=10,
    )

    # Línea de datos
    x_vals = df_ind["Fecha"] if "Fecha" in df_ind.columns else df_ind["Periodo"]
    custom = df_ind[["Periodo", "Categoria"]].values if "Periodo" in df_ind.columns else None

    fig.add_trace(go.Scatter(
        x=x_vals,
        y=cum_pct,
        mode="lines+markers+text",
        line=dict(color="#455A64", width=2),
        marker=dict(
            size=12,
            color=colores_pts,
            line=dict(width=2, color="white"),
        ),
        text=cum_pct.astype(str) + "%",
        textposition="top center",
        textfont=dict(size=9),
        hovertemplate=(
            "<b>Periodo:</b> %{customdata[0]}<br>"
            "<b>Cumplimiento:</b> %{y:.1f}%<br>"
            "<b>Estado:</b> %{customdata[1]}<extra></extra>"
        ) if custom is not None else "%{y:.1f}%<extra></extra>",
        customdata=custom,
        showlegend=False,
    ))

    # Líneas de referencia
    fig.add_hline(y=100, line_dash="dash", line_color="#2E7D32", line_width=1.5)
    fig.add_hline(y=80,  line_dash="dot",  line_color="#C62828", line_width=1.5)

    # Leyenda de colores
    for cat, color in COLOR_CAT.items():
        if cat == "Sin dato":
            continue
        fig.add_trace(go.Scatter(
            x=[None], y=[None], mode="markers",
            marker=dict(size=10, color=color),
            name=cat, showlegend=True,
        ))

    fig.update_layout(
        title=titulo,
        yaxis=dict(title="Cumplimiento (%)", ticksuffix="%", range=[0, y_max]),
        xaxis_title="",
        plot_bgcolor="white",
        paper_bgcolor="white",
        legend=dict(orientation="h", y=-0.2),
        margin=dict(t=50, b=60),
    )
    return fig


def tabla_historica_indicador(df_ind: pd.DataFrame) -> pd.DataFrame:
    """Prepara DataFrame histórico con columnas formateadas para mostrar."""
    cols = ["Periodo", "Anio", "Meta", "Ejecucion", "Cumplimiento_norm", "Categoria"]
    cols_disp = [c for c in cols if c in df_ind.columns]
    df_t = df_ind[cols_disp].copy().sort_values("Anio") if "Anio" in df_ind.columns else df_ind[cols_disp].copy()
    if "Cumplimiento_norm" in df_t.columns:
        df_t["Cumplimiento_norm"] = (df_t["Cumplimiento_norm"] * 100).round(1).astype(str) + "%"
        df_t.rename(columns={"Cumplimiento_norm": "Cumplimiento%"}, inplace=True)
    return df_t


def exportar_excel(df: pd.DataFrame, nombre_hoja: str = "Datos") -> bytes:
    """Retorna bytes xlsx con formato institucional para st.download_button."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=nombre_hoja)
        ws = writer.sheets[nombre_hoja]
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1A3A5C")
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
    return output.getvalue()


def colorear_tabla_categoria(df: pd.DataFrame, col_categoria: str = "Categoria") -> pd.DataFrame.style:
    """Aplica estilos de color a filas según categoría."""
    COLOR_BG = {
        "Peligro":           "#FDE8F3",
        "Alerta":            "#FEF3D0",
        "Cumplimiento":      "#E0F7FA",
        "Sobrecumplimiento": "#D0E4FF",
        "Sin dato":          "#F5F5F5",
    }

    def estilo_fila(row):
        cat = row.get(col_categoria, "")
        bg = COLOR_BG.get(cat, "")
        return [f"background-color: {bg}" if bg else "" for _ in row]

    return df.style.apply(estilo_fila, axis=1)


def panel_detalle_indicador(df_ind: pd.DataFrame, id_ind: str, df_full: pd.DataFrame):
    """
    Renderiza el panel completo de detalle de un indicador.
    Llamar desde un st.dialog o st.expander.
    """
    from utils.calculos import generar_recomendaciones, calcular_tendencia

    if df_ind.empty:
        st.warning("Sin datos para este indicador.")
        return

    df_ind_sorted = df_ind.sort_values("Fecha")
    ultimo = df_ind_sorted.iloc[-1]

    nombre = ultimo.get("Indicador", "—")
    proceso = ultimo.get("Proceso", "—")
    subproceso = ultimo.get("Subproceso", "—")
    periodicidad = ultimo.get("Periodicidad", "—")
    clasificacion = ultimo.get("Clasificacion", "—")
    cum_norm = ultimo.get("Cumplimiento_norm", None)
    categoria = ultimo.get("Categoria", "Sin dato")

    cum_pct_str = f"{round(float(cum_norm)*100, 1)}%" if pd.notna(cum_norm) else "—"

    BADGE_COLOR = {
        "Peligro":          "#C62828",
        "Alerta":           "#F57F17",
        "Cumplimiento":     "#2E7D32",
        "Sobrecumplimiento":"#0277BD",
        "Sin dato":         "#9E9E9E",
    }
    badge_col = BADGE_COLOR.get(categoria, "#9E9E9E")

    st.markdown(f"### {id_ind} — {nombre}")

    c1, c2, c3, c4 = st.columns([2, 2, 2, 1])
    with c1:
        st.markdown(f"**Proceso:** {proceso}")
    with c2:
        st.markdown(f"**Subproceso:** {subproceso}")
    with c3:
        st.markdown(f"**Periodicidad:** {periodicidad} · {clasificacion}")
    with c4:
        st.markdown(
            f"<div style='font-size:1.5rem; font-weight:bold;'>{cum_pct_str}</div>"
            f"<span style='background:{badge_col};color:white;padding:2px 10px;"
            f"border-radius:12px;font-size:0.85rem'>{categoria}</span>",
            unsafe_allow_html=True,
        )

    st.divider()

    # Tabla histórica
    df_tabla = tabla_historica_indicador(df_ind_sorted)
    st.markdown("**Histórico**")
    st.dataframe(df_tabla, use_container_width=True, hide_index=True)

    # Gráfico histórico
    fig = grafico_historico_indicador(df_ind_sorted)
    st.plotly_chart(fig, use_container_width=True)

    # Recomendaciones
    st.divider()
    cum_series = df_ind_sorted["Cumplimiento_norm"].dropna() * 100
    tendencia, recs = generar_recomendaciones(categoria, cum_series)

    tendencia_icons = {
        "Mejorando":            "✅ Mejorando",
        "Empeorando":           "⚠️ Empeorando",
        "Estable":              "➡️ Estable",
        "Sin datos suficientes":"ℹ️ Sin datos suficientes",
    }
    st.markdown(
        f"**Recomendaciones**  |  Tendencia: **{tendencia_icons.get(tendencia, tendencia)}**"
    )
    for rec in recs:
        st.markdown(f"- {rec}")
