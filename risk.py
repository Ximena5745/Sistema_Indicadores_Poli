import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# Crear workbook
wb = Workbook()

# Eliminar hoja por defecto
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# ===========================
# HOJA 1: MATRIZ DE INDICADORES
# ===========================
ws1 = wb.create_sheet("Matriz de Indicadores", 0)

# Encabezados
headers = [
    "#",
    "Nombre del Indicador",
    "Tipo",
    "Clasificación",
    "Fórmula",
    "Descripción Técnica",
    "Justificación",
    "Fuente de Datos",
    "Frecuencia de Medición",
    "Relación Normativa"
]

# Estilo de encabezados
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Escribir encabezados
for col_num, header in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

# Datos de indicadores
indicadores = [
    {
        "num": 1,
        "nombre": "Índice de Madurez de Gestión de Riesgos por Proceso",
        "tipo": "Resultado",
        "clasificacion": "Efectividad",
        "formula": "(Σ Procesos con controles nivel fuerte / Total procesos institucionales) × 100",
        "descripcion": "Mide el porcentaje de procesos institucionales que han alcanzado nivel de diseño fuerte en sus controles de riesgo, según escala de madurez definida en metodología institucional",
        "justificacion": "Indicador estratégico que revela la penetración real del sistema de gestión de riesgos en la organización, permite priorizar recursos de intervención y demuestra evolución institucional",
        "fuente": "Matriz de riesgos institucional, Módulo de gestión de riesgos",
        "frecuencia": "Semestral",
        "normativa": "ISO 9001: 6.1 (Acciones para abordar riesgos)\nISO 21001: 6.1.2 (Riesgos en organizaciones educativas)\nCNA: Factor 6 (Organización, gestión y administración)"
    },
    {
        "num": 2,
        "nombre": "Tasa de Materialización Efectiva de Riesgos Críticos",
        "tipo": "Impacto",
        "clasificacion": "Eficacia",
        "formula": "(N° riesgos críticos materializados / Total riesgos críticos identificados) × 100",
        "descripcion": "Cuantifica únicamente riesgos de nivel crítico (impacto alto + probabilidad alta) que se materializaron generando consecuencias reales versus el universo de riesgos críticos documentados",
        "justificacion": "Diferencia entre incidentes menores y crisis organizacionales; permite evaluar si la gestión está protegiendo objetivos estratégicos o solo gestionando burocráticamente matrices",
        "fuente": "Sistema de incidentes, Matriz de riesgos (filtro: riesgos críticos), Registro de eventos adversos",
        "frecuencia": "Trimestral",
        "normativa": "ISO 9001: 10.2 (No conformidad y acción correctiva)\nISO 45001: 6.1.2 (Identificación de peligros)\nISO 14001: 6.1.1 (Aspectos ambientales)\nReporte: Consejo Superior, Rectoría"
    },
    {
        "num": 3,
        "nombre": "Velocidad de Respuesta en Tratamiento de Riesgos Emergentes",
        "tipo": "Producto",
        "clasificacion": "Eficiencia",
        "formula": "Promedio de días entre (Fecha identificación riesgo nuevo - Fecha implementación primer control)",
        "descripcion": "Mide tiempo transcurrido desde detección de riesgo emergente (no contemplado en matrices) hasta implementación efectiva del primer control documentado",
        "justificacion": "Evalúa agilidad organizacional ante cambios normativos, tecnológicos o del entorno; separa gestión proactiva de gestión reactiva tardía",
        "fuente": "Matriz de riesgos (campo fecha identificación y fecha control), Sistema de gestión documental",
        "frecuencia": "Trimestral",
        "normativa": "ISO 9001: 6.1 (Acciones riesgos y oportunidades)\nISO 21001: 8.2.4 (Evaluación continua)\nAplicable: Informes de gestión a Ministerios ante cambios normativos"
    },
    {
        "num": 4,
        "nombre": "Índice de Robustez de Controles (IRC)",
        "tipo": "Resultado",
        "clasificacion": "Efectividad",
        "formula": "[(Σ Puntaje controles verificados / Puntaje máximo posible) × 100] donde puntaje = f(Actividad, Documentación, Frecuencia, Naturaleza, Evidencia, Tipo)",
        "descripcion": "Evalúa solidez real de controles mediante verificación multidimensional de 6 variables críticas, generando un score ponderado por control",
        "justificacion": "Trasciende verificación binaria (existe/no existe control); permite identificar controles débiles aunque documentados, priorizando rediseño",
        "fuente": "Herramienta tecnológica de riesgos (módulo verificación controles), Evidencias de ejecución",
        "frecuencia": "Cuatrimestral",
        "normativa": "ISO 9001: 9.1 (Seguimiento, medición, análisis)\nISO 45001: 9.1.1 (Seguimiento y medición)\nISO 14001: 9.1.1 (Seguimiento, medición, análisis)"
    },
    {
        "num": 5,
        "nombre": "Tasa de Planes de Continuidad Activados con Éxito",
        "tipo": "Impacto",
        "clasificacion": "Eficacia",
        "formula": "(N° planes de continuidad activados exitosamente / Total activaciones requeridas) × 100",
        "descripcion": "Mide efectividad de planes de continuidad cuando son activados por materialización de riesgos, evaluando si permitieron mantener operaciones críticas dentro de parámetros aceptables",
        "justificacion": "Indicador de resiliencia organizacional real; distingue entre tener planes documentados (cumplimiento formal) vs. tener capacidad de respuesta efectiva",
        "fuente": "Registro de activaciones, Informes post-incidente, Evaluaciones de recuperación",
        "frecuencia": "Anual (o por evento)",
        "normativa": "ISO 22301 (referencia): Continuidad de negocio\nISO 21001: 8.7 (Aprendizaje organizacional)\nCNA: Factor 7 (Bienestar institucional) en continuidad servicios"
    },
    {
        "num": 6,
        "nombre": "Cobertura de Riesgos Regulatorios Críticos",
        "tipo": "Producto",
        "clasificacion": "Eficacia",
        "formula": "(N° requisitos normativos críticos con control asociado / Total requisitos normativos críticos aplicables) × 100",
        "descripcion": "Verifica que cada requisito legal/normativo crítico (MEN, Min Ambiente, Min Trabajo, CNA, SNIES) tenga control específico vinculado en matriz de riesgos",
        "justificacion": "Protege a la institución de sanciones, pérdida de registros calificados o acreditación; asegura trazabilidad entre compliance y gestión de riesgos",
        "fuente": "Matriz legal, Matriz de riesgos (filtro: riesgos de cumplimiento), Registros SNIES",
        "frecuencia": "Semestral",
        "normativa": "ISO 9001: 4.2 (Comprensión necesidades partes interesadas)\nISO 21001: 4.2 (Necesidades partes interesadas educativas)\nReporte: SNIES, CNA, Ministerios sectorial"
    },
    {
        "num": 7,
        "nombre": "Índice de Transformación Positiva de Riesgos",
        "tipo": "Resultado",
        "clasificacion": "Efectividad",
        "formula": "[(N° riesgos que disminuyeron nivel + N° riesgos eliminados) / Total riesgos gestionados en periodo anterior] × 100",
        "descripcion": "Cuantifica riesgos que efectivamente redujeron su nivel de exposición (impacto×probabilidad) o fueron eliminados, versus actualización pasiva de matrices",
        "justificacion": "Mide gestión transformadora vs. gestión burocrática; permite identificar si controles generan valor real o solo documentación",
        "fuente": "Matriz de riesgos (comparativo periodo t vs t-1), Informes de cambio de nivel de riesgo",
        "frecuencia": "Semestral",
        "normativa": "ISO 9001: 10.3 (Mejora continua)\nISO 45001: 10.3 (Mejora continua)\nISO 14001: 10.3 (Mejora continua)"
    },
    {
        "num": 8,
        "nombre": "Efectividad de Planes de Mejoramiento Derivados de Riesgos",
        "tipo": "Resultado",
        "clasificacion": "Eficacia",
        "formula": "(N° hallazgos de auditoría/revisión cerrados exitosamente / Total hallazgos generados por gestión de riesgos) × 100",
        "descripcion": "Evalúa cierre efectivo de no conformidades u oportunidades de mejora identificadas durante supervisión/monitoreo de controles",
        "justificacion": "Conecta gestión de riesgos con mejora continua; evita que hallazgos queden documentados sin resolución efectiva",
        "fuente": "Sistema de acciones correctivas/preventivas, Informes de auditoría, Matriz de seguimiento a hallazgos",
        "frecuencia": "Trimestral",
        "normativa": "ISO 9001: 10.2 (No conformidad y acción correctiva)\nISO 21001: 10.2 (No conformidad y mejora)\nCNA: Sistemas de aseguramiento de calidad"
    },
    {
        "num": 9,
        "nombre": "Índice de Alineación Estratégica de Riesgos",
        "tipo": "Resultado",
        "clasificacion": "Efectividad",
        "formula": "(N° proyectos estratégicos con riesgos identificados y controlados / Total proyectos estratégicos vigentes) × 100",
        "descripcion": "Verifica que todos los proyectos del Plan Estratégico Institucional tengan análisis de riesgos y controles asociados en matriz institucional",
        "justificacion": "Asegura que gestión de riesgos está alineada con objetivos estratégicos, no solo con operaciones rutinarias; protege inversión y transformación institucional",
        "fuente": "Plan Estratégico Institucional, Matriz de riesgos (módulo proyectos estratégicos), Plan de Retos",
        "frecuencia": "Cuatrimestral",
        "normativa": "ISO 9001: 5.1.1 (Liderazgo y compromiso)\nISO 21001: 5.1.1 (Enfoque al estudiante y partes interesadas)\nReporte: Consejo Superior, Comité Rectoral"
    },
    {
        "num": 10,
        "nombre": "Tasa de Actualización Metodológica ante Cambios Normativos",
        "tipo": "Producto",
        "clasificacion": "Eficiencia",
        "formula": "(N° actualizaciones metodología GR ejecutadas / N° cambios normativos aplicables identificados) × 100",
        "descripcion": "Mide si la metodología de gestión de riesgos se actualiza oportunamente ante cambios en ISO, decretos MEN, resoluciones ambientales, normas SST, etc.",
        "justificacion": "Garantiza vigencia y pertinencia del sistema de gestión de riesgos; evita obsolescencia metodológica que convertiría el proceso en inútil",
        "fuente": "Registro de cambios metodología GR, Matriz de seguimiento normativo, Actas Comité de Riesgos",
        "frecuencia": "Anual",
        "normativa": "ISO 9001: 7.1.6 (Conocimientos organizacionales)\nISO 45001: 6.1.1 (Generalidades en SST)\nISO 14001: 6.1.1 (Generalidades ambiental)\nAplicable: Auditorías de certificación"
    }
]

# Estilos para datos
data_alignment = Alignment(vertical="top", wrap_text=True)
border_style = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# Colores alternados para filas
row_fill_1 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
row_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Escribir datos
for idx, ind in enumerate(indicadores, 2):
    row_fill = row_fill_1 if idx % 2 == 0 else row_fill_2
    
    ws1.cell(row=idx, column=1, value=ind["num"])
    ws1.cell(row=idx, column=2, value=ind["nombre"])
    ws1.cell(row=idx, column=3, value=ind["tipo"])
    ws1.cell(row=idx, column=4, value=ind["clasificacion"])
    ws1.cell(row=idx, column=5, value=ind["formula"])
    ws1.cell(row=idx, column=6, value=ind["descripcion"])
    ws1.cell(row=idx, column=7, value=ind["justificacion"])
    ws1.cell(row=idx, column=8, value=ind["fuente"])
    ws1.cell(row=idx, column=9, value=ind["frecuencia"])
    ws1.cell(row=idx, column=10, value=ind["normativa"])
    
    for col in range(1, 11):
        cell = ws1.cell(row=idx, column=col)
        cell.alignment = data_alignment
        cell.border = border_style
        cell.fill = row_fill

# Ajustar anchos de columna
ws1.column_dimensions['A'].width = 5
ws1.column_dimensions['B'].width = 45
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 55
ws1.column_dimensions['F'].width = 60
ws1.column_dimensions['G'].width = 60
ws1.column_dimensions['H'].width = 50
ws1.column_dimensions['I'].width = 18
ws1.column_dimensions['J'].width = 55

# Altura de fila de encabezado
ws1.row_dimensions[1].height = 40

# ===========================
# HOJA 2: ANÁLISIS DE MEJORAS
# ===========================
ws2 = wb.create_sheet("Análisis de Mejoras", 1)

# Título
ws2.merge_cells('A1:F1')
title_cell = ws2['A1']
title_cell.value = "ANÁLISIS DE MEJORAS IDENTIFICADAS - PROCESO GESTIÓN DE RIESGOS"
title_cell.font = Font(bold=True, size=14, color="FFFFFF")
title_cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

# Sección 1: Fortalezas
ws2.merge_cells('A3:F3')
fortaleza_title = ws2['A3']
fortaleza_title.value = "✅ FORTALEZAS DEL PROCESO ACTUAL"
fortaleza_title.font = Font(bold=True, size=12, color="FFFFFF")
fortaleza_title.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
fortaleza_title.alignment = Alignment(horizontal="left", vertical="center")

fortalezas = [
    "Alcance integral que abarca todos los roles institucionales (académicos, administrativos) y grupos de interés",
    "Vinculación directa con proyectos estratégicos institucionales a través del Plan de Retos",
    "Metodología de gestión de riesgos documentada y actualizable",
    "Existencia de herramienta tecnológica específica para gestión de riesgos",
    "Visión holística de riesgos (estratégicos, operacionales, financieros, tecnológicos, reputacionales, legales)",
    "Definición clara de productos documentados (matrices actualizadas, informes de comportamiento)"
]

row = 4
for fortaleza in fortalezas:
    ws2.merge_cells(f'A{row}:F{row}')
    cell = ws2[f'A{row}']
    cell.value = f"• {fortaleza}"
    cell.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
    ws2.row_dimensions[row].height = 30
    row += 1

# Sección 2: Vacíos Críticos
row += 1
ws2.merge_cells(f'A{row}:F{row}')
vacios_title = ws2[f'A{row}']
vacios_title.value = "⚠️ VACÍOS CRÍTICOS IDENTIFICADOS"
vacios_title.font = Font(bold=True, size=12, color="FFFFFF")
vacios_title.fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
vacios_title.alignment = Alignment(horizontal="left", vertical="center")

row += 1
# Encabezados de tabla de vacíos
vacios_headers = ["Dimensión", "Vacío Identificado", "Impacto", "Recomendación"]
for col_num, header in enumerate(vacios_headers, 1):
    cell = ws2.cell(row=row, column=col_num)
    cell.value = header
    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_style

vacios_data = [
    {
        "dimension": "ENTRADAS DEL PROCESO",
        "vacio": "No se especifican las entradas del proceso. No hay claridad sobre quién, cómo y cuándo se alimenta la identificación inicial de riesgos",
        "impacto": "Riesgo de identificación reactiva o incompleta de riesgos. Dependencia de criterio subjetivo sin fuentes estructuradas",
        "recomendacion": "Definir entradas explícitas: matriz de contexto organizacional (ISO 9001:4.1), cambios normativos identificados, resultados de auditorías internas/externas, incidentes reportados, tendencias de PQRS, análisis de partes interesadas"
    },
    {
        "dimension": "INTEGRACIÓN NORMATIVA",
        "vacio": "No se evidencia mecanismo sistemático de captura de cambios normativos externos (MEN, Min Ambiente, Min Trabajo, nuevas ISO)",
        "impacto": "Obsolescencia de la metodología. Riesgos regulatorios no identificados oportunamente. Potenciales sanciones o pérdida de acreditaciones",
        "recomendacion": "Implementar entrada formal: 'Monitoreo normativo' con responsable definido. Incluir alertas tempranas de cambios legales como trigger de actualización metodológica"
    },
    {
        "dimension": "CAPTURA DE RIESGOS EMERGENTES",
        "vacio": "No hay mecanismo explícito para capturar riesgos emergentes de partes interesadas externas (estudiantes, empleadores, sector productivo, comunidad)",
        "impacto": "Visión endogámica de riesgos. Posible desconexión con expectativas de stakeholders clave. Riesgos reputacionales no anticipados",
        "recomendacion": "Establecer entrada periódica: 'Consulta a partes interesadas' sobre riesgos percibidos. Integrar resultados de encuestas de satisfacción, empleabilidad y pertinencia como fuentes de identificación"
    },
    {
        "dimension": "MEDICIÓN DE EFECTIVIDAD",
        "vacio": "Los indicadores actuales son mayormente operativos. Falta medición de impacto real y transformación en el perfil de riesgo institucional",
        "impacto": "Imposibilidad de demostrar valor agregado de la gestión de riesgos a alta dirección. Percepción de burocracia sin resultados tangibles",
        "recomendacion": "Implementar los indicadores estratégicos propuestos (Índice de Madurez, Transformación Positiva, Alineación Estratégica) que miden cambio real vs. actividad documental"
    },
    {
        "dimension": "INTEGRACIÓN CON MEJORA CONTINUA",
        "vacio": "No se evidencia vínculo estructurado entre hallazgos de gestión de riesgos y sistema de mejora continua (acciones correctivas/preventivas)",
        "impacto": "Hallazgos de supervisión pueden quedar documentados sin seguimiento efectivo. Desconexión entre procesos de evaluación y mejora",
        "recomendacion": "Formalizar salida del proceso: 'Solicitudes de acción correctiva/preventiva derivadas de GR' con trazabilidad en sistema integrado de gestión. Implementar indicador #8 propuesto"
    },
    {
        "dimension": "PLANES DE CONTINUIDAD",
        "vacio": "Se menciona 'determinar necesidad' de planes de continuidad, pero no hay criterios claros ni indicadores de efectividad de estos planes",
        "impacto": "Planes pueden existir documentalmente sin capacidad real de activación. Resiliencia organizacional no verificada",
        "recomendacion": "Establecer criterios obligatorios para PCN (procesos críticos según análisis de impacto). Implementar indicador #5 (Tasa de Planes Activados con Éxito) y realizar simulacros periódicos"
    }
]

row += 1
for vacio in vacios_data:
    ws2.cell(row=row, column=1, value=vacio["dimension"])
    ws2.cell(row=row, column=2, value=vacio["vacio"])
    ws2.cell(row=row, column=3, value=vacio["impacto"])
    ws2.cell(row=row, column=4, value=vacio["recomendacion"])
    
    for col in range(1, 5):
        cell = ws2.cell(row=row, column=col)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border_style
    
    ws2.row_dimensions[row].height = 80
    row += 1

# Ajustar anchos
ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 45
ws2.column_dimensions['C'].width = 40
ws2.column_dimensions['D'].width = 50
ws2.column_dimensions['E'].width = 15
ws2.column_dimensions['F'].width = 15

# Sección 3: Recomendaciones Estratégicas
row += 2
ws2.merge_cells(f'A{row}:F{row}')
rec_title = ws2[f'A{row}']
rec_title.value = "🎯 RECOMENDACIONES ESTRATÉGICAS PRIORIZADAS"
rec_title.font = Font(bold=True, size=12, color="FFFFFF")
rec_title.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
rec_title.alignment = Alignment(horizontal="left", vertical="center")

recomendaciones = [
    {
        "prioridad": "CRÍTICA",
        "accion": "Definir explícitamente las ENTRADAS del proceso",
        "plazo": "Inmediato (30 días)",
        "responsable": "Líder Gestión de Riesgos + Líder Sistema Integrado de Gestión"
    },
    {
        "prioridad": "ALTA",
        "accion": "Implementar sistema de monitoreo normativo automatizado",
        "plazo": "Corto plazo (90 días)",
        "responsable": "Oficina Jurídica + Gestión de Riesgos"
    },
    {
        "prioridad": "ALTA",
        "accion": "Rediseñar batería de indicadores incorporando los 10 indicadores estratégicos propuestos",
        "plazo": "Corto plazo (60 días)",
        "responsable": "Gestión de Riesgos + Planeación Institucional"
    },
    {
        "prioridad": "MEDIA",
        "accion": "Establecer mecanismo de consulta semestral a partes interesadas sobre riesgos percibidos",
        "plazo": "Mediano plazo (120 días)",
        "responsable": "Gestión de Riesgos + Relacionamiento con Stakeholders"
    },
    {
        "prioridad": "MEDIA",
        "accion": "Integrar formalmente GR con sistema de acciones correctivas/preventivas",
        "plazo": "Mediano plazo (90 días)",
        "responsable": "Gestión de Riesgos + Auditoría Interna + Mejora Continua"
    },
    {
        "prioridad": "BAJA",
        "accion": "Definir criterios para PCN obligatorios y programa de simulacros",
        "plazo": "Largo plazo (180 días)",
        "responsable": "Gestión de Riesgos + Líderes de procesos críticos"
    }
]

row += 1
rec_headers = ["Prioridad", "Acción Recomendada", "Plazo Sugerido", "Responsable Sugerido"]
for col_num, header in enumerate(rec_headers, 1):
    cell = ws2.cell(row=row, column=col_num)
    cell.value = header
    cell.fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_style

row += 1
for rec in recomendaciones:
    # Color según prioridad
    if rec["prioridad"] == "CRÍTICA":
        priority_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        priority_font = Font(bold=True, color="FFFFFF")
    elif rec["prioridad"] == "ALTA":
        priority_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        priority_font = Font(bold=True)
    else:
        priority_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        priority_font = Font(bold=True)
    
    cell_prior = ws2.cell(row=row, column=1, value=rec["prioridad"])
    cell_prior.fill = priority_fill
    cell_prior.font = priority_font
    cell_prior.alignment = Alignment(horizontal="center", vertical="center")
    cell_prior.border = border_style
    
    ws2.cell(row=row, column=2, value=rec["accion"])
    ws2.cell(row=row, column=3, value=rec["plazo"])
    ws2.cell(row=row, column=4, value=rec["responsable"])
    
    for col in range(2, 5):
        cell = ws2.cell(row=row, column=col)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border_style
    
    ws2.row_dimensions[row].height = 40
    row += 1

# ===========================
# HOJA 3: COMPARATIVO INDICADORES
# ===========================
ws3 = wb.create_sheet("Comparativo Indicadores", 2)

# Título
ws3.merge_cells('A1:E1')
title_cell = ws3['A1']
title_cell.value = "COMPARATIVO: INDICADORES ACTUALES vs. INDICADORES PROPUESTOS"
title_cell.font = Font(bold=True, size=14, color="FFFFFF")
title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 30

# Indicadores actuales
ws3.merge_cells('A3:E3')
actual_title = ws3['A3']
actual_title.value = "INDICADORES ACTUALES"
actual_title.font = Font(bold=True, size=11, color="FFFFFF")
actual_title.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
actual_title.alignment = Alignment(horizontal="left", vertical="center")

actuales_headers = ["Indicador Actual", "Tipo", "Limitación Identificada", "Nivel de Valor", "Observación"]
row = 4
for col_num, header in enumerate(actuales_headers, 1):
    cell = ws3.cell(row=row, column=col_num)
    cell.value = header
    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_style

actuales_data = [
    {
        "nombre": "Nivel de Satisfacción Servicios Prestados - Riesgos",
        "tipo": "Percepción",
        "limitacion": "Mide satisfacción del usuario interno, no efectividad real del proceso en reducir riesgos",
        "valor": "Bajo",
        "observacion": "Útil para mejora de servicio, pero no para evaluar desempeño estratégico de GR"
    },
    {
        "nombre": "Frecuencia de incidentes de riesgos",
        "tipo": "Lagging",
        "limitacion": "No discrimina entre incidentes menores y crisis críticas. Trata todos los riesgos por igual",
        "valor": "Medio",
        "observacion": "Reemplazar por Indicador #2 propuesto (Tasa Materialización Riesgos CRÍTICOS)"
    },
    {
        "nombre": "Eficacia de controles",
        "tipo": "Proceso",
        "limitacion": "Verifica existencia de variables, pero no solidez real ni efectividad probada del control",
        "valor": "Medio",
        "observacion": "Mejorar con Indicador #4 propuesto (IRC - Índice Robustez Controles con scoring ponderado)"
    },
    {
        "nombre": "Nivel de cumplimiento del plan de seguimiento a controles",
        "tipo": "Actividad",
        "limitacion": "Mide cumplimiento de agenda, no resultados del seguimiento ni mejoras generadas",
        "valor": "Bajo",
        "observacion": "Indicador puramente operativo sin valor estratégico. Eliminar o complementar con Indicador #8"
    },
    {
        "nombre": "Respuesta ante nuevas necesidades de control",
        "tipo": "Producto",
        "limitacion": "No mide velocidad ni calidad de respuesta, solo cantidad de controles nuevos",
        "valor": "Bajo",
        "observacion": "Reemplazar por Indicador #3 propuesto (Velocidad de Respuesta con medición en días)"
    }
]

row += 1
for actual in actuales_data:
    # Color según valor
    if actual["valor"] == "Bajo":
        valor_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    elif actual["valor"] == "Medio":
        valor_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    else:
        valor_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    ws3.cell(row=row, column=1, value=actual["nombre"])
    ws3.cell(row=row, column=2, value=actual["tipo"])
    ws3.cell(row=row, column=3, value=actual["limitacion"])
    
    cell_valor = ws3.cell(row=row, column=4, value=actual["valor"])
    cell_valor.fill = valor_fill
    cell_valor.alignment = Alignment(horizontal="center", vertical="center")
    
    ws3.cell(row=row, column=5, value=actual["observacion"])
    
    for col in range(1, 6):
        cell = ws3.cell(row=row, column=col)
        if col != 4:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border_style
    
    ws3.row_dimensions[row].height = 60
    row += 1

# Indicadores propuestos
row += 2
ws3.merge_cells(f'A{row}:E{row}')
propuesto_title = ws3[f'A{row}']
propuesto_title.value = "INDICADORES PROPUESTOS - VENTAJAS DIFERENCIALES"
propuesto_title.font = Font(bold=True, size=11, color="FFFFFF")
propuesto_title.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
propuesto_title.alignment = Alignment(horizontal="left", vertical="center")

row += 1
propuestos_headers = ["Indicador Propuesto", "Tipo", "Ventaja Diferencial", "Valor Estratégico", "Aplicación"]
for col_num, header in enumerate(propuestos_headers, 1):
    cell = ws3.cell(row=row, column=col_num)
    cell.value = header
    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_style

propuestos_data = [
    {
        "nombre": "Índice de Madurez de GR por Proceso",
        "tipo": "Resultado",
        "ventaja": "Permite benchmarking interno y priorización de recursos en procesos con menor madurez",
        "valor": "MUY ALTO",
        "aplicacion": "Decisiones de inversión en rediseño de controles"
    },
    {
        "nombre": "Tasa Materialización Riesgos Críticos",
        "tipo": "Impacto",
        "ventaja": "Discrimina crisis reales de incidentes menores. Enfoque en protección de objetivos estratégicos",
        "valor": "MUY ALTO",
        "aplicacion": "Reporte a Consejo Superior y Rectoría"
    },
    {
        "nombre": "Índice de Robustez de Controles (IRC)",
        "tipo": "Resultado",
        "ventaja": "Scoring multidimensional (6 variables) vs. verificación binaria. Identifica controles débiles",
        "valor": "ALTO",
        "aplicacion": "Priorización de rediseño de controles"
    },
    {
        "nombre": "Índice de Transformación Positiva",
        "tipo": "Resultado",
        "ventaja": "Mide gestión transformadora vs. burocrática. Demuestra valor agregado real de GR",
        "valor": "MUY ALTO",
        "aplicacion": "Evidencia de mejora continua para certificaciones ISO"
    },
    {
        "nombre": "Cobertura Riesgos Regulatorios",
        "tipo": "Producto",
        "ventaja": "Protección directa contra sanciones, pérdida de registros o acreditación. Trazabilidad compliance",
        "valor": "CRÍTICO",
        "aplicacion": "Reporte a entes externos (SNIES, CNA, Ministerios)"
    }
]

row += 1
for prop in propuestos_data:
    # Color según valor
    if prop["valor"] == "CRÍTICO":
        valor_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        valor_font = Font(bold=True, color="FFFFFF")
    elif prop["valor"] == "MUY ALTO":
        valor_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        valor_font = Font(bold=True)
    else:
        valor_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        valor_font = Font(bold=False)
    
    ws3.cell(row=row, column=1, value=prop["nombre"])
    ws3.cell(row=row, column=2, value=prop["tipo"])
    ws3.cell(row=row, column=3, value=prop["ventaja"])
    
    cell_valor = ws3.cell(row=row, column=4, value=prop["valor"])
    cell_valor.fill = valor_fill
    cell_valor.font = valor_font
    cell_valor.alignment = Alignment(horizontal="center", vertical="center")
    
    ws3.cell(row=row, column=5, value=prop["aplicacion"])
    
    for col in range(1, 6):
        cell = ws3.cell(row=row, column=col)
        if col != 4:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border_style
    
    ws3.row_dimensions[row].height = 50
    row += 1

# Ajustar anchos
ws3.column_dimensions['A'].width = 40
ws3.column_dimensions['B'].width = 15
ws3.column_dimensions['C'].width = 50
ws3.column_dimensions['D'].width = 15
ws3.column_dimensions['E'].width = 45

# ===========================
# HOJA 4: ROADMAP DE IMPLEMENTACIÓN
# ===========================
ws4 = wb.create_sheet("Roadmap Implementación", 3)

# Título
ws4.merge_cells('A1:G1')
title_cell = ws4['A1']
title_cell.value = "ROADMAP DE IMPLEMENTACIÓN - INDICADORES ESTRATÉGICOS GR"
title_cell.font = Font(bold=True, size=14, color="FFFFFF")
title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 30

# Encabezados
roadmap_headers = ["Fase", "Indicador", "Prioridad", "Requisitos Previos", "Fuente de Datos a Habilitar", "Plazo", "Responsable"]
row = 3
for col_num, header in enumerate(roadmap_headers, 1):
    cell = ws4.cell(row=row, column=col_num)
    cell.value = header
    cell.fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_style

roadmap_data = [
    {
        "fase": "FASE 1\nRápida Implementación",
        "indicador": "#6 - Cobertura Riesgos Regulatorios",
        "prioridad": "CRÍTICA",
        "requisitos": "Matriz legal actualizada, Vinculación controles-requisitos en herramienta",
        "fuente": "Matriz legal, Matriz de riesgos (módulo compliance)",
        "plazo": "30 días",
        "responsable": "Jurídica + GR"
    },
    {
        "fase": "FASE 1\nRápida Implementación",
        "indicador": "#9 - Alineación Estratégica",
        "prioridad": "ALTA",
        "requisitos": "Plan Estratégico vigente, Módulo proyectos en herramienta GR",
        "fuente": "Plan Estratégico, Matriz de riesgos (proyectos estratégicos)",
        "plazo": "45 días",
        "responsable": "Planeación + GR"
    },
    {
        "fase": "FASE 2\nConsolidación",
        "indicador": "#1 - Índice de Madurez GR",
        "prioridad": "ALTA",
        "requisitos": "Escala de madurez definida, Clasificación de controles por nivel",
        "fuente": "Matriz de riesgos, Auditorías de procesos",
        "plazo": "60 días",
        "responsable": "GR + SIG"
    },
    {
        "fase": "FASE 2\nConsolidación",
        "indicador": "#4 - IRC (Robustez Controles)",
        "prioridad": "ALTA",
        "requisitos": "Modelo de scoring implementado en herramienta, Evidencias digitalizadas",
        "fuente": "Herramienta GR (módulo verificación), Repositorio evidencias",
        "plazo": "90 días",
        "responsable": "GR + TI"
    },
    {
        "fase": "FASE 2\nConsolidación",
        "indicador": "#7 - Transformación Positiva",
        "prioridad": "MEDIA",
        "requisitos": "Histórico de matrices (comparativo temporal), Trazabilidad cambios nivel riesgo",
        "fuente": "Matriz de riesgos (histórico), Informes de cambio",
        "plazo": "90 días",
        "responsable": "GR"
    },
    {
        "fase": "FASE 3\nMaduración",
        "indicador": "#2 - Tasa Materialización Críticos",
        "prioridad": "ALTA",
        "requisitos": "Sistema de registro de incidentes, Clasificación criticidad de riesgos",
        "fuente": "Sistema de incidentes, Matriz de riesgos (filtro críticos)",
        "plazo": "120 días",
        "responsable": "GR + Todos los procesos"
    },
    {
        "fase": "FASE 3\nMaduración",
        "indicador": "#3 - Velocidad de Respuesta",
        "prioridad": "MEDIA",
        "requisitos": "Campos de fecha en herramienta GR (identificación y control), Procedimiento riesgos emergentes",
        "fuente": "Herramienta GR (timestamps), Sistema documental",
        "plazo": "120 días",
        "responsable": "GR + TI"
    },
    {
        "fase": "FASE 3\nMaduración",
        "indicador": "#8 - Efectividad Planes Mejoramiento",
        "prioridad": "MEDIA",
        "requisitos": "Integración GR con sistema acciones correctivas, Trazabilidad hallazgos-cierre",
        "fuente": "Sistema CAPA, Informes auditoría, Seguimiento hallazgos",
        "plazo": "150 días",
        "responsable": "GR + Auditoría + Mejora"
    },
    {
        "fase": "FASE 4\nExcelencia",
        "indicador": "#5 - Planes Continuidad Exitosos",
        "prioridad": "BAJA",
        "requisitos": "PCN diseñados y documentados, Programa de simulacros, Criterios de éxito definidos",
        "fuente": "Registro activaciones, Informes post-incidente, Evaluaciones",
        "plazo": "180 días",
        "responsable": "GR + Líderes procesos"
    },
    {
        "fase": "FASE 4\nExcelencia",
        "indicador": "#10 - Actualización Metodológica",
        "prioridad": "BAJA",
        "requisitos": "Registro de cambios metodología, Monitoreo normativo sistemático",
        "fuente": "Control de cambios metodología, Matriz seguimiento normativo",
        "plazo": "180 días",
        "responsable": "GR + Jurídica"
    }
]

row += 1
for item in roadmap_data:
    # Color según prioridad
    if item["prioridad"] == "CRÍTICA":
        priority_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        priority_font = Font(bold=True, color="FFFFFF")
    elif item["prioridad"] == "ALTA":
        priority_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        priority_font = Font(bold=True)
    elif item["prioridad"] == "MEDIA":
        priority_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        priority_font = Font(bold=False)
    else:
        priority_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        priority_font = Font(bold=False)
    
    ws4.cell(row=row, column=1, value=item["fase"])
    ws4.cell(row=row, column=2, value=item["indicador"])
    
    cell_prior = ws4.cell(row=row, column=3, value=item["prioridad"])
    cell_prior.fill = priority_fill
    cell_prior.font = priority_font
    cell_prior.alignment = Alignment(horizontal="center", vertical="center")
    
    ws4.cell(row=row, column=4, value=item["requisitos"])
    ws4.cell(row=row, column=5, value=item["fuente"])
    ws4.cell(row=row, column=6, value=item["plazo"])
    ws4.cell(row=row, column=7, value=item["responsable"])
    
    for col in range(1, 8):
        cell = ws4.cell(row=row, column=col)
        if col != 3:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border_style
    
    ws4.row_dimensions[row].height = 50
    row += 1

# Ajustar anchos
ws4.column_dimensions['A'].width = 18
ws4.column_dimensions['B'].width = 35
ws4.column_dimensions['C'].width = 12
ws4.column_dimensions['D'].width = 45
ws4.column_dimensions['E'].width = 40
ws4.column_dimensions['F'].width = 12
ws4.column_dimensions['G'].width = 25

# Guardar archivo
output_dir = os.path.join("data", "output")
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

output_path = os.path.join(output_dir, "Matriz_Indicadores_Gestion_Riesgos.xlsx")
wb.save(output_path)

print(f"OK: Archivo Excel generado exitosamente")
print(f"Ubicacion: {output_path}")
print(f"\nContenido del archivo:")
print(f"  - Hoja 1: Matriz completa de 10 indicadores estrategicos")
print(f"  - Hoja 2: Analisis detallado de mejoras y vacios identificados")
print(f"  - Hoja 3: Comparativo indicadores actuales vs. propuestos")
print(f"  - Hoja 4: Roadmap de implementacion por fases")
