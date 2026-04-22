#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
"""
generar_reporte.py  (versión dinámica 2026)
==========================================
Procesa fuentes LMI + Kawak y genera 'Seguimiento_Reporte.xlsx' con:

  1. Hoja "Seguimiento"       → copia completa del LMI + columna "Revisar"
  2. Hojas por periodicidad   → columnas Periodo X renombradas a fechas reales
                                + columnas "Reportado" / "Estado del indicador"
  3. Hoja "Resumen"           → tabla resumen por periodicidad
  4. Hoja "Tracking Mensual"  → matriz Id × mes/año  (Reportado/Pendiente/No aplica)
                                Calendario dinámico desde FECHA_INICIO hasta hoy

FUENTES DE DATOS
────────────────
  Fuente operativa (LMI):
    data/raw/lmi_reporte.xlsx
    → Columnas: Id, Periodicidad, Periodo 1 … Periodo N

  Fuente histórica de RESULTADOS (API Kawak):
    data/raw/Fuentes Consolidadas/Consolidado_API_Kawak.xlsx
    → Generado por scripts/consolidar_api.py (PARTE 2: data/raw/API/*.xlsx)
    → Columnas clave: ID, fecha, resultado
    → Fuente primaria para determinar si un indicador fue reportado

  Catálogo de indicadores (Kawak):
    data/raw/Fuentes Consolidadas/Indicadores Kawak.xlsx
    → Generado por scripts/consolidar_api.py (PARTE 1: data/raw/Kawak/*.xlsx)
    → Columnas: Año, Id, Periodicidad, Indicador, ...
    → Se usa para enriquecer la periodicidad de indicadores no presentes en LMI

CONFIGURACIÓN
─────────────
  RUTA_ORIGEN            : archivo fuente LMI (.xlsx)
  RUTA_SALIDA            : archivo de salida (.xlsx)
  RUTA_KAWAK_API         : Consolidado_API_Kawak.xlsx  (resultados históricos)
  RUTA_KAWAK_CATALOGO    : Indicadores Kawak.xlsx      (catálogo de periodicidad)
  FECHA_INICIO           : primer mes del calendario histórico (Tracking Mensual)
  FECHA_REFERENCIA_MANUAL: None = auto (último día del mes anterior al actual)
                           O fijar manualmente, ej: date(2025, 12, 31)
  COLUMNA_REVISAR        : columna usada para detectar inicio de nuevo indicador
"""

import os
from datetime import date, timedelta
from typing import Dict, List, Optional, Tuple

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Importar lógica centralizada de categorización
try:
    from core.semantica import categorizar_cumplimiento
except ImportError:
    # Fallback si no se puede importar
    def categorizar_cumplimiento(c, id_indicador=None):
        """Fallback local si semantica no está disponible."""
        try:
            v = float(c)
        except (TypeError, ValueError):
            return "Sin dato"
        if v >= 1.05: return "Sobrecumplimiento"
        if v >= 1.00: return "Cumplimiento"
        if v >= 0.80: return "Alerta"
        return "Peligro"

# ── Configuración ──────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

RUTA_ORIGEN = os.path.join(BASE_DIR, "data", "raw", "lmi_reporte.xlsx")
RUTA_SALIDA = os.path.join(BASE_DIR, "data", "output", "Seguimiento_Reporte.xlsx")

# Fuente de resultados históricos (generada por consolidar_api.py PARTE 2)
# Columnas: ID | fecha | resultado | ... (minúsculas, de la API)
RUTA_KAWAK_API = os.path.join(
    BASE_DIR, "data", "raw", "Fuentes Consolidadas", "Consolidado_API_Kawak.xlsx"
)

# Catálogo de indicadores Kawak (generado por consolidar_api.py PARTE 1)
# Columnas: Año | Id | Indicador | Clasificacion | Periodicidad | Sentido | ...
# NO tiene Fecha ni Resultado — solo metadata
RUTA_KAWAK_CATALOGO = os.path.join(
    BASE_DIR, "data", "raw", "Fuentes Consolidadas", "Indicadores Kawak.xlsx"
)

# Directorio con los archivos anuales de Kawak (catálogos activos por año)
# Contiene: 2025.xlsx, 2026.xlsx, ...
# Fuente autoritativa de IDs activos para el Tracking Mensual
RUTA_KAWAK_DIR = os.path.join(BASE_DIR, "data", "raw", "Kawak")

# Columna cuyo cambio entre filas marca un nuevo indicador
COLUMNA_REVISAR = "Id"

# ── Configuración dinámica de fechas ──────────────────────────────────────────
# None = automático: último día del mes anterior al mes actual.
# Fijarlo manualmente si se necesita un corte específico.
FECHA_REFERENCIA_MANUAL: Optional[date] = None

# Inicio del calendario histórico para el "Tracking Mensual"
FECHA_INICIO = date(2025, 1, 1)

# ── Reglas de negocio especiales ──────────────────────────────────────────────

# Indicadores que reportan en "año vencido":
# Su dato del año N (en Kawak) cuenta como "Reportado" para el año N+1.
# Ej: dato 2024 en Kawak → Reportado en dic-2025 del tracking.
IDS_AÑO_VENCIDO: set = {"226", "227", "228"}

# Indicadores con vigencia parcial: solo aplican desde la fecha indicada.
# Meses anteriores a esa fecha se omiten del tracking (no generan fila).
IDS_VIGENCIA_DESDE: Dict[str, date] = {
    "515": date(2025, 7, 1),
    "516": date(2025, 7, 1),
    "526": date(2025, 7, 1),
    "530": date(2025, 7, 1),
    "531": date(2025, 7, 1),
    "538": date(2025, 7, 1),
    "539": date(2025, 7, 1),
    "554": date(2025, 7, 1),
    "555": date(2025, 7, 1),
}

# Procesos cuyos indicadores muestran "No aplica" cuando no tienen reporte
# (en lugar de "Pendiente").  Comparación case-insensitive, coincidencia parcial.
PROCESOS_NO_APLICA_SIN_REPORTE: List[str] = [
    "Gestión de Unidades Académicas",
]

# ── Fuente de cierres históricos ──────────────────────────────────────────────
RUTA_CONSOLIDADOS = os.path.join(BASE_DIR, "data", "output", "Resultados Consolidados.xlsx")

# ── Umbrales de cumplimiento (alineados con core/config.py) ───────────────────
UMBRAL_PELIGRO_D           = 0.80   # < 80%  → Peligro
UMBRAL_ALERTA_D            = 1.00   # 80–99% → Alerta  / ≥100% → Cumplimiento
UMBRAL_SOBRECUMPLIMIENTO_D = 1.05   # ≥105%  → Sobrecumplimiento

# ── Paleta de colores ──────────────────────────────────────────────────────────
C_HEADER    = "1F4E79"
C_SI        = "C6EFCE"
C_NO        = "FFCCCC"
C_PENDIENTE = "FFEB9C"
C_REVISAR1  = "DDEEFF"
C_NO_APLICA = "EEEEEE"

# Fondos para niveles de cumplimiento (matriz de calor)
C_NIVEL_BG = {
    "Sobrecumplimiento": "D0E4FF",
    "Cumplimiento":      "E8F5E9",
    "Alerta":            "FEF3D0",
    "Peligro":           "FFCDD2",
    "Sin dato":          "F0F0F0",
}


# ══════════════════════════════════════════════════════════════════════════════
#  FECHA DE REFERENCIA DINÁMICA
# ══════════════════════════════════════════════════════════════════════════════

def _calcular_fecha_ref() -> date:
    """Retorna el último día del mes anterior al mes actual."""
    hoy = date.today()
    return hoy.replace(day=1) - timedelta(days=1)


def get_fecha_referencia() -> date:
    """Punto de acceso único para la fecha de referencia."""
    return FECHA_REFERENCIA_MANUAL or _calcular_fecha_ref()


# ══════════════════════════════════════════════════════════════════════════════
#  UTILIDADES BÁSICAS
# ══════════════════════════════════════════════════════════════════════════════

def _ultimo_dia(year: int, month: int) -> date:
    if month == 12:
        return date(year, 12, 31)
    return date(year, month + 1, 1) - timedelta(days=1)


def _retroceder(year: int, mes_actual: int, meses_ciclo: list) -> tuple:
    idx = meses_ciclo.index(mes_actual)
    if idx == 0:
        return year - 1, meses_ciclo[-1]
    return year, meses_ciclo[idx - 1]


def _tiene_dato(v) -> bool:
    if v is None:
        return False
    s = str(v).strip()
    return s not in ("", "-", "nan", "NaN", "None")


def _tiene_dato_kawak(v) -> bool:
    try:
        if pd.isna(v):
            return False
    except (TypeError, ValueError):
        pass
    s = str(v).strip()
    return s not in ("", "-", "nan", "NaN", "None")


def _id_normalizar(x) -> str:
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except (TypeError, ValueError):
        pass
    try:
        f = float(x)
        return str(int(f)) if f == int(f) else str(f)
    except (ValueError, TypeError):
        return str(x).strip()


def _detectar_col(df: pd.DataFrame, candidatos: list) -> Optional[str]:
    """Detección dinámica de columna: busca por nombre exacto, luego case-insensitive."""
    # Búsqueda exacta primero
    for cand in candidatos:
        if cand in df.columns:
            return cand
    # Fallback case-insensitive
    col_lower = {c.lower(): c for c in df.columns}
    for cand in candidatos:
        found = col_lower.get(cand.lower())
        if found:
            return found
    return None


# ══════════════════════════════════════════════════════════════════════════════
#  FECHAS DE PERÍODO (hojas por periodicidad — retroactivo desde fecha_ref)
# ══════════════════════════════════════════════════════════════════════════════

def get_period_dates(periodicidad: str, n: int = 13,
                     fecha_ref: Optional[date] = None) -> list:
    """
    Retorna n fechas (último día de cada período) hacia atrás desde fecha_ref.
    Periodo 1 = período más reciente <= fecha_ref.
    """
    ref = fecha_ref or get_fecha_referencia()
    dates = []

    if periodicidad == "Mensual":
        y, m = ref.year, ref.month
        for _ in range(n):
            dates.append(_ultimo_dia(y, m))
            m -= 1
            if m == 0:
                m, y = 12, y - 1

    elif periodicidad == "Bimestral":
        ciclo = [2, 4, 6, 8, 10, 12]
        y, m = ref.year, ref.month
        previos = [x for x in ciclo if x <= m]
        cur = max(previos) if previos else 12
        if not previos:
            y -= 1
        for _ in range(n):
            dates.append(_ultimo_dia(y, cur))
            y, cur = _retroceder(y, cur, ciclo)

    elif periodicidad == "Trimestral":
        ciclo = [3, 6, 9, 12]
        y, m = ref.year, ref.month
        previos = [x for x in ciclo if x <= m]
        cur = max(previos) if previos else 12
        if not previos:
            y -= 1
        for _ in range(n):
            dates.append(_ultimo_dia(y, cur))
            y, cur = _retroceder(y, cur, ciclo)

    elif periodicidad == "Semestral":
        ciclo = [6, 12]
        y, m = ref.year, ref.month
        previos = [x for x in ciclo if x <= m]
        cur = max(previos) if previos else 12
        if not previos:
            y -= 1
        for _ in range(n):
            dates.append(_ultimo_dia(y, cur))
            y, cur = _retroceder(y, cur, ciclo)

    elif periodicidad == "Anual":
        # Período anual = diciembre de cada año.
        # Si ref.month < 12, el diciembre de ref.year aún no ha ocurrido
        # → el período más reciente es diciembre del año anterior.
        y = ref.year if ref.month == 12 else ref.year - 1
        for i in range(n):
            dates.append(date(y - i, 12, 31))

    else:
        dates = [None] * n

    return dates


# ══════════════════════════════════════════════════════════════════════════════
#  LECTURA DE ARCHIVOS
# ══════════════════════════════════════════════════════════════════════════════

def leer_xlsx(ruta: str) -> pd.DataFrame:
    """Lee LMI .xlsx con openpyxl; celdas vacías → None; todo como texto."""
    df = pd.read_excel(
        ruta,
        engine="openpyxl",
        keep_default_na=False,
        na_values=[""],
        dtype=str,
    )
    df.columns = [str(c).strip() for c in df.columns]
    df = df.where(df != "nan", other=None)
    df = df.where(df != "NaN", other=None)
    return df


def leer_kawak_api(ruta: str) -> Dict:
    """
    Lee Consolidado_API_Kawak.xlsx y retorna {(id_str, year, month): resultado}.

    Estructura esperada (generada por consolidar_api.py PARTE 2):
      Columnas: ID | fecha | resultado | ... (minúsculas, origen API)
    Si el mismo (id, year, month) aparece varias veces conserva el último
    registro no vacío (el archivo ya está ordenado por ID + fecha).
    """
    if not os.path.exists(ruta):
        print(f"    INFO: {ruta} no encontrado — se omite lookup Kawak.")
        return {}

    try:
        df = pd.read_excel(ruta, engine="openpyxl",
                           keep_default_na=False, na_values=[""])
    except Exception as exc:
        print(f"    ADVERTENCIA Kawak API: no se pudo leer: {exc}")
        return {}

    df.columns = [str(c).strip() for c in df.columns]

    # Columnas del Consolidado_API_Kawak: "ID" (mayús.), "fecha", "resultado"
    col_id  = _detectar_col(df, ["ID", "Id", "id", "Codigo", "Código"])
    col_fec = _detectar_col(df, ["fecha", "Fecha", "FechaPeriodo", "Periodo", "Mes"])
    col_res = _detectar_col(df, ["resultado", "Resultado", "Valor", "valor", "Result"])

    if not col_id or not col_fec or not col_res:
        print(f"    ADVERTENCIA Kawak API: columnas no encontradas.\n"
              f"    Disponibles: {df.columns.tolist()}")
        return {}

    lookup: Dict = {}
    omitidas = 0
    for _, row in df.iterrows():
        kid = _id_normalizar(row[col_id])
        if not kid:
            continue

        fecha_raw = row[col_fec]
        resultado = row[col_res]

        try:
            if isinstance(fecha_raw, (pd.Timestamp, date)):
                ts = pd.Timestamp(fecha_raw)
            elif isinstance(fecha_raw, (int, float)) and not pd.isna(fecha_raw):
                ts = pd.Timestamp("1899-12-30") + pd.Timedelta(days=int(fecha_raw))
            else:
                ts = pd.Timestamp(str(fecha_raw).strip())

            if pd.isna(resultado):
                continue
            # El archivo está ordenado por fecha → el último registro gana
            lookup[(kid, ts.year, ts.month)] = resultado
        except Exception:
            omitidas += 1

    if omitidas:
        print(f"    INFO Kawak API: {omitidas} filas con fecha no parseable omitidas.")

    print(f"    Kawak API cargado: {len(lookup)} registros (Id × período).")
    return lookup


def leer_catalogo_kawak(ruta: str) -> Dict[str, str]:
    """
    Lee Indicadores Kawak.xlsx (catálogo de metadata) y retorna
    {id_str: periodicidad} usando el año más reciente disponible.

    Estructura (generada por consolidar_api.py PARTE 1):
      Columnas: Año | Id | Indicador | Clasificacion | Proceso | Tipo | Periodicidad | Sentido
    NO contiene Fecha ni Resultado — es solo metadata de los indicadores.
    """
    if not os.path.exists(ruta):
        print(f"    INFO: {ruta} no encontrado — catálogo Kawak omitido.")
        return {}

    try:
        df = pd.read_excel(ruta, engine="openpyxl",
                           keep_default_na=False, na_values=[""])
    except Exception as exc:
        print(f"    ADVERTENCIA catálogo Kawak: no se pudo leer: {exc}")
        return {}

    df.columns = [str(c).strip() for c in df.columns]

    col_id   = _detectar_col(df, ["Id", "ID", "id"])
    col_per  = _detectar_col(df, ["Periodicidad", "periodicidad", "frecuencia"])
    col_anio = _detectar_col(df, ["Año", "Anio", "anio", "año", "Year"])

    if not col_id or not col_per:
        print(f"    ADVERTENCIA catálogo Kawak: columnas Id/Periodicidad no encontradas.\n"
              f"    Disponibles: {df.columns.tolist()}")
        return {}

    # Si hay columna de año, ordenar para que el más reciente quede último
    if col_anio:
        try:
            df[col_anio] = pd.to_numeric(df[col_anio], errors="coerce")
            df = df.sort_values(col_anio, na_position="first")
        except Exception:
            pass

    catalogo: Dict[str, str] = {}
    for _, row in df.iterrows():
        kid = _id_normalizar(row[col_id])
        if not kid:
            continue
        perio = str(row.get(col_per, "")).strip()
        if perio and perio not in ("nan", "NaN", "None", ""):
            catalogo[kid] = perio  # El más reciente sobrescribe

    print(f"    Catálogo Kawak cargado: {len(catalogo)} indicadores con periodicidad.")
    return catalogo


def leer_catalogo_por_anio(kawak_dir: str, year: int) -> Tuple[Dict[str, str], Dict[str, dict]]:
    """
    Lee data/raw/Kawak/{year}.xlsx — fuente autoritativa de IDs activos para ese año.
    Retorna:
      catalogo : {id_str: periodicidad}
      meta     : {id_str: {col: valor, ...}}  (metadata descriptiva sin Id/Periodicidad)

    Si el archivo no existe, retorna ({}, {}) sin error fatal.
    """
    ruta = os.path.join(kawak_dir, f"{year}.xlsx")
    if not os.path.exists(ruta):
        print(f"    INFO: {ruta} no encontrado — año {year} omitido del tracking.")
        return {}, {}

    try:
        df = pd.read_excel(ruta, engine="openpyxl", keep_default_na=False, na_values=[""])
    except Exception as exc:
        print(f"    ADVERTENCIA catálogo {year}: {exc}")
        return {}, {}

    df.columns = [str(c).strip() for c in df.columns]
    col_id  = _detectar_col(df, ["Id", "ID", "id"])
    col_per = _detectar_col(df, ["Periodicidad", "periodicidad", "frecuencia"])

    if not col_id or not col_per:
        print(f"    ADVERTENCIA: columnas Id/Periodicidad no encontradas en {ruta}\n"
              f"    Disponibles: {df.columns.tolist()}")
        return {}, {}

    # Columnas de metadata (todo excepto Id y Periodicidad)
    excluir_meta = {col_id, col_per}
    meta_cols = [c for c in df.columns if c not in excluir_meta]

    catalogo: Dict[str, str] = {}
    meta: Dict[str, dict] = {}

    for _, row in df.iterrows():
        kid = _id_normalizar(row[col_id])
        if not kid:
            continue
        perio = str(row.get(col_per, "")).strip()
        if perio and perio not in ("nan", "NaN", "None", ""):
            catalogo[kid] = perio
            if kid not in meta:
                meta[kid] = {c: row.get(c) for c in meta_cols}

    print(f"    Kawak {year}: {len(catalogo)} indicadores activos.")
    return catalogo, meta


# ══════════════════════════════════════════════════════════════════════════════
#  LÓGICA DE SEGUIMIENTO TRADICIONAL (hojas por periodicidad)
# ══════════════════════════════════════════════════════════════════════════════

def agregar_revisar(df: pd.DataFrame, col: str) -> pd.DataFrame:
    df = df.copy()
    valores = df[col].tolist()
    revisar = [1] + [
        0 if valores[i] == valores[i - 1] else 1
        for i in range(1, len(valores))
    ]
    df["Revisar"] = revisar
    return df


def agregar_columnas_seguimiento(df: pd.DataFrame, col_p1: str,
                                  col_p2: Optional[str] = None) -> pd.DataFrame:
    df = df.copy()
    df["Reportado"] = df[col_p1].apply(
        lambda v: "Sí" if _tiene_dato(v) else "No"
    )
    df["Estado del indicador"] = df[col_p1].apply(
        lambda v: "Reportado" if _tiene_dato(v) else "Pendiente de reporte"
    )
    return df


def enriquecer_desde_kawak(df_p: pd.DataFrame, kawak: Dict,
                            col_p1: str, col_p2: Optional[str],
                            fecha_p1, fecha_p2) -> tuple:
    if not kawak:
        return df_p, 0

    df_p = df_p.copy()
    n_act = 0

    for i, row in df_p.iterrows():
        if row.get("Estado del indicador") != "Pendiente de reporte":
            continue

        kid      = _id_normalizar(row.get("Id", ""))
        cambiado = False

        if fecha_p1 and not _tiene_dato(row.get(col_p1, "")):
            res1 = kawak.get((kid, fecha_p1.year, fecha_p1.month))
            if res1 is not None and _tiene_dato_kawak(res1):
                df_p.at[i, col_p1] = res1
                cambiado = True

        if col_p2 and fecha_p2 and not _tiene_dato(row.get(col_p2, "")):
            res2 = kawak.get((kid, fecha_p2.year, fecha_p2.month))
            if res2 is not None and _tiene_dato_kawak(res2):
                df_p.at[i, col_p2] = res2
                cambiado = True

        if cambiado:
            p1_ok = _tiene_dato(df_p.at[i, col_p1])
            df_p.at[i, "Reportado"]            = "Sí" if p1_ok else "No"
            df_p.at[i, "Estado del indicador"] = (
                "Reportado" if p1_ok else "Pendiente de reporte"
            )
            if p1_ok:
                n_act += 1

    return df_p, n_act


# ══════════════════════════════════════════════════════════════════════════════
#  TRACKING MENSUAL — lógica principal
# ══════════════════════════════════════════════════════════════════════════════

def _mes_cierre_periodo(month: int, periodicidad: str) -> int:
    """
    Dado un mes calendario cualquiera, retorna el mes de CIERRE del período
    al que pertenece según la periodicidad del indicador.

    Ejemplos:
      Trimestral: Feb (2) → 3  |  May (5) → 6  |  Oct (10) → 12
      Bimestral:  Mar (3) → 4  |  Nov (11) → 12
      Semestral:  Abr (4) → 6  |  Sep (9)  → 12
      Anual:      cualquier mes → 12
      Mensual:    mes → mismo mes
    """
    perio = str(periodicidad).strip()
    if perio == "Mensual":
        return month
    if perio == "Bimestral":
        return month if month % 2 == 0 else month + 1
    if perio == "Trimestral":
        if month <= 3:  return 3
        if month <= 6:  return 6
        if month <= 9:  return 9
        return 12
    if perio == "Semestral":
        return 6 if month <= 6 else 12
    if perio == "Anual":
        return 12
    return month  # periodicidad desconocida → sin normalizar


def normalizar_kawak_lookup(kawak_raw: Dict,
                             catalogo: Dict[str, str]) -> Dict:
    """
    Re-mapea las claves del lookup Kawak al mes de CIERRE del período
    correspondiente según la periodicidad de cada indicador.

    Problema que resuelve:
      La API Kawak almacena el dato con la fecha real del reporte
      (ej: 2025-02-10 para un Trimestral de Q1).
      El tracking busca por el mes de cierre del período (2025-03 para Q1).
      Sin normalizar, la búsqueda falla y el indicador aparece como Pendiente.

    Regla de precedencia:
      Si dos fechas distintas del mismo indicador cierran al mismo período,
      se conserva el último valor encontrado (el más reciente en el archivo).
    """
    normalizado: Dict = {}
    for (kid, year, month), valor in kawak_raw.items():
        perio = catalogo.get(kid, "Mensual")
        mes_norm = _mes_cierre_periodo(month, perio)
        normalizado[(kid, year, mes_norm)] = valor  # último gana
    return normalizado


def aplica_periodicidad(periodicidad: str, mes: int) -> bool:
    """
    Determina si un mes calendario aplica según la periodicidad del indicador.
      Mensual    → todos los meses
      Bimestral  → meses pares (2, 4, 6, 8, 10, 12)
      Trimestral → 3, 6, 9, 12
      Semestral  → 6, 12
      Anual      → 12
    Periodicidad desconocida → False (conservador: no marcar como Pendiente).
    """
    perio = str(periodicidad).strip()
    if perio == "Mensual":
        return True
    if perio == "Bimestral":
        return mes % 2 == 0
    if perio == "Trimestral":
        return mes in (3, 6, 9, 12)
    if perio == "Semestral":
        return mes in (6, 12)
    if perio == "Anual":
        return mes == 12
    return False


def generar_calendario(fecha_inicio: date, fecha_fin: date) -> List[Tuple[int, int]]:
    """Genera lista ordenada de (año, mes) desde fecha_inicio hasta fecha_fin inclusive."""
    calendario: List[Tuple[int, int]] = []
    y, m = fecha_inicio.year, fecha_inicio.month
    fin_y, fin_m = fecha_fin.year, fecha_fin.month
    while (y < fin_y) or (y == fin_y and m <= fin_m):
        calendario.append((y, m))
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return calendario


def construir_lmi_lookup(df: pd.DataFrame, fecha_ref: date) -> Dict:
    """
    Construye {(id_str, year, month): True} desde el DataFrame LMI.
    Marca presencia cuando la celda del período tiene un dato válido.
    Usa la periodicidad de cada fila para mapear Periodo X → (year, month).
    """
    lookup: Dict = {}
    periodo_cols = [c for c in df.columns if str(c).startswith("Periodo ")]
    if not periodo_cols:
        return lookup

    n = len(periodo_cols)
    for _, row in df.iterrows():
        kid = _id_normalizar(row.get("Id", ""))
        if not kid:
            continue
        perio = str(row.get("Periodicidad", "")).strip()
        fechas = get_period_dates(perio, n, fecha_ref)

        for col, fecha in zip(periodo_cols, fechas):
            if fecha is None:
                continue
            if _tiene_dato(row.get(col)):
                lookup[(kid, fecha.year, fecha.month)] = True

    return lookup


def construir_catalogo(df_lmi: pd.DataFrame,
                        catalogo_kawak: Dict[str, str]) -> Dict[str, str]:
    """
    Construye {id_str: periodicidad} usando Kawak del año vigente como fuente
    autoritativa de IDs activos, con LMI como enriquecimiento.

    Regla:
      1. LMI carga todos sus IDs con su periodicidad (si la tiene).
      2. Kawak agrega IDs que no estén en LMI y completa periodicidades vacías.
         Esto incluye indicadores activos en Kawak que aún no figuren en el LMI.
      3. Se descartan IDs sin periodicidad resoluble.

    Nota: el catálogo Kawak que se pasa aquí es SIEMPRE del año vigente
    (leer_catalogo_kawak lee Indicadores Kawak.xlsx filtrado por año reciente),
    por lo que no incorpora indicadores históricos descontinuados.
    """
    catalogo: Dict[str, str] = {}

    # 1) LMI — carga sus IDs con periodicidad
    for _, row in df_lmi.iterrows():
        kid = _id_normalizar(row.get("Id", ""))
        if not kid:
            continue
        perio = str(row.get("Periodicidad", "")).strip()
        if perio and perio not in ("nan", "NaN", "None", ""):
            catalogo[kid] = perio
        elif kid not in catalogo:
            catalogo[kid] = ""  # ID presente pero sin periodicidad aún

    # 2) Kawak completa periodicidad faltante Y agrega IDs activos no en LMI
    for kid, perio in catalogo_kawak.items():
        if kid not in catalogo:
            catalogo[kid] = perio  # ID nuevo desde Kawak vigente
        elif not catalogo[kid] and perio:
            catalogo[kid] = perio  # completa periodicidad vacía

    # 3) Remover los que quedaron sin periodicidad resoluble
    catalogo = {k: v for k, v in catalogo.items() if v}

    return catalogo


def _extraer_meta_indicadores(df: pd.DataFrame) -> Dict[str, dict]:
    """
    Extrae metadata descriptiva de cada indicador desde el LMI.
    Toma la primera aparición de cada Id (fila donde Revisar == 1).
    Excluye columnas de control y columnas de período.
    """
    excluir = {"Revisar", "Id", "Periodicidad"}
    periodo_cols = {c for c in df.columns if str(c).startswith("Periodo ")}
    meta_cols = [c for c in df.columns if c not in excluir and c not in periodo_cols]

    meta: Dict[str, dict] = {}
    for _, row in df.iterrows():
        kid = _id_normalizar(row.get("Id", ""))
        if not kid or kid in meta:
            continue
        meta[kid] = {col: row.get(col) for col in meta_cols}
    return meta


# Nombres de meses en español para la columna Mes_Nombre
_MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


def _proceso_es_no_aplica(proceso: str) -> bool:
    """Retorna True si el proceso del indicador usa 'No aplica' en lugar de 'Pendiente'."""
    proceso_norm = str(proceso or "").strip().lower()
    for p in PROCESOS_NO_APLICA_SIN_REPORTE:
        if p.lower() in proceso_norm:
            return True
    return False


def construir_tracking_largo(
    catalogo_por_anio: Dict[int, Dict[str, str]],
    kawak_lookup: Dict,
    lmi_lookup: Dict,
    calendario: List[Tuple[int, int]],
    meta_por_anio: Optional[Dict[int, Dict[str, dict]]] = None,
) -> pd.DataFrame:
    """
    Genera DataFrame en FORMATO LARGO (long) filtrable mes a mes.
    Una fila por (Id × período que APLICA según periodicidad).

    El catálogo de IDs activos es AÑO-ESPECÍFICO:
      catalogo_por_anio[2025] = IDs activos en 2025 (data/raw/Kawak/2025.xlsx)
      catalogo_por_anio[2026] = IDs activos en 2026 (data/raw/Kawak/2026.xlsx)
    Esto permite que indicadores se activen o inactiven entre años.

    Columnas: Id | <meta Kawak> | Periodicidad | Año | Mes | Mes_Nombre | Periodo | Estado | Fuente

    Reglas de negocio:
      1. IDS_AÑO_VENCIDO    → busca también en año anterior de Kawak.
      2. IDS_VIGENCIA_DESDE → omite meses anteriores a la fecha de inicio.
      3. PROCESOS_NO_APLICA → sin dato = "No aplica" en vez de "Pendiente".
    """
    if not catalogo_por_anio:
        return pd.DataFrame()

    meta_por_anio = meta_por_anio or {}

    def _sort_key(kid: str) -> tuple:
        try:
            return (0, int(kid), kid)
        except (ValueError, TypeError):
            return (1, 0, kid)

    # Orden de columnas de metadata (unión de todos los años)
    prioridad = ["Indicador", "Nombre", "Proceso", "Subproceso", "Linea",
                 "Clasificacion", "Tipo", "Sentido", "Área", "Area"]
    todas_meta_cols: set = set()
    for meta_year in meta_por_anio.values():
        for m in meta_year.values():
            todas_meta_cols.update(m.keys())
    meta_cols_order = [c for c in prioridad if c in todas_meta_cols]
    meta_cols_order += [c for c in sorted(todas_meta_cols) if c not in meta_cols_order]

    rows = []
    for (year, month) in calendario:
        catalogo_year = catalogo_por_anio.get(year, {})
        meta_year     = meta_por_anio.get(year, {})

        for kid in sorted(catalogo_year.keys(), key=_sort_key):
            perio   = catalogo_year[kid]
            meta    = meta_year.get(kid, {})
            proceso = str(meta.get("Proceso", "") or "").strip()

            # ── Regla base: mes debe aplicar según periodicidad ──────────
            if not aplica_periodicidad(perio, month):
                continue

            # ── Regla 2: vigencia parcial ────────────────────────────────
            if kid in IDS_VIGENCIA_DESDE:
                if date(year, month, 1) < IDS_VIGENCIA_DESDE[kid]:
                    continue

            # ── Regla 1: año vencido ─────────────────────────────────────
            if kid in IDS_AÑO_VENCIDO:
                en_kawak = (
                    (kid, year,     month) in kawak_lookup or
                    (kid, year - 1, month) in kawak_lookup
                )
            else:
                en_kawak = (kid, year, month) in kawak_lookup

            en_lmi = (kid, year, month) in lmi_lookup

            # ── Estado y fuente ──────────────────────────────────────────
            if en_kawak:
                estado, fuente = "Reportado", "Kawak"
            elif en_lmi:
                estado, fuente = "Reportado", "LMI"
            else:
                estado = "No aplica" if _proceso_es_no_aplica(proceso) else "Pendiente"
                fuente = "—"

            row_data: Dict = {"Id": kid}
            for col in meta_cols_order:
                row_data[col] = meta.get(col)
            row_data["Periodicidad"] = perio
            row_data["Año"]          = year
            row_data["Mes"]          = month
            row_data["Mes_Nombre"]   = _MESES_ES.get(month, str(month))
            row_data["Periodo"]      = f"{month}/{year}"
            row_data["Estado"]       = estado
            row_data["Fuente"]       = fuente
            rows.append(row_data)

    return pd.DataFrame(rows)


# ══════════════════════════════════════════════════════════════════════════════
#  ESCRITURA CON OPENPYXL
# ══════════════════════════════════════════════════════════════════════════════

def _estilo_header(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor=C_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borde = Side(style="thin", color="FFFFFF")
    cell.border = Border(left=borde, right=borde, bottom=borde)


def escribir_hoja(ws, df: pd.DataFrame, mapa_fechas: dict = None):
    """Escribe el DataFrame en ws con formato institucional."""
    mapa_fechas = mapa_fechas or {}

    for ci, col in enumerate(df.columns, 1):
        val = col
        if col in mapa_fechas and mapa_fechas[col] is not None:
            val = mapa_fechas[col].strftime("%d/%m/%Y")
        cell = ws.cell(row=1, column=ci, value=val)
        _estilo_header(cell)
        ws.column_dimensions[get_column_letter(ci)].width = max(len(str(val)) + 2, 10)

    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, raw_val in enumerate(row, 1):
            col_name = df.columns[ci - 1]

            if col_name == "Id":
                val = _id_normalizar(raw_val) if raw_val is not None else None
            elif raw_val is not None and str(raw_val).strip() == "-":
                val = None
            else:
                val = raw_val

            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if col_name == "Reportado":
                if val == "Sí":
                    cell.fill = PatternFill("solid", fgColor=C_SI)
                elif val == "No":
                    cell.fill = PatternFill("solid", fgColor=C_NO)

            elif col_name == "Estado del indicador":
                if val == "Reportado":
                    cell.fill = PatternFill("solid", fgColor=C_SI)
                elif val == "Pendiente de reporte":
                    cell.fill = PatternFill("solid", fgColor=C_PENDIENTE)

            elif col_name == "Revisar":
                try:
                    if int(val) == 1:
                        cell.fill = PatternFill("solid", fgColor=C_REVISAR1)
                except (TypeError, ValueError):
                    pass

            ancho_actual = ws.column_dimensions[get_column_letter(ci)].width
            contenido_ancho = len(str(val)) + 2 if val is not None else 0
            ws.column_dimensions[get_column_letter(ci)].width = min(
                40, max(ancho_actual, contenido_ancho)
            )

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions


def escribir_tracking_mensual(ws, df: pd.DataFrame):
    """
    Escribe la hoja "Tracking Mensual" en FORMATO LARGO con código de colores:
      Verde    → Reportado
      Amarillo → Pendiente
    Todas las columnas tienen filtro automático.
    Congela hasta la columna "Periodicidad" inclusive.
    """
    if df.empty:
        ws.cell(row=1, column=1, value="Sin datos — verifica las fuentes.")
        return

    fill_reportado = PatternFill("solid", fgColor=C_SI)
    fill_pendiente = PatternFill("solid", fgColor=C_PENDIENTE)

    # Anchos sugeridos por tipo de columna
    ANCHOS = {
        "Id": 8, "Periodicidad": 14, "Año": 7, "Mes": 6,
        "Mes_Nombre": 13, "Periodo": 10, "Estado": 13, "Fuente": 10,
    }

    # Encabezados
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        _estilo_header(cell)
        ws.column_dimensions[get_column_letter(ci)].width = ANCHOS.get(
            col, min(max(len(str(col)) + 4, 14), 40)
        )

    # Identificar índice de columna "Periodicidad" para congelar hasta ahí
    cols_list = list(df.columns)
    try:
        col_freeze = cols_list.index("Periodicidad") + 2  # columna siguiente a Periodicidad
    except ValueError:
        col_freeze = 2
    ws.freeze_panes = f"{get_column_letter(col_freeze)}2"

    # Datos
    for ri, row in enumerate(df.itertuples(index=False), 2):
        estado_val = None
        for ci, val in enumerate(row, 1):
            col_name = cols_list[ci - 1]
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if col_name == "Estado":
                estado_val = val

        # Colorear fila completa según Estado
        if estado_val in ("Reportado", "Pendiente"):
            fill = fill_reportado if estado_val == "Reportado" else fill_pendiente
            col_estado = cols_list.index("Estado") + 1
            ws.cell(row=ri, column=col_estado).fill = fill

    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}1"


# ══════════════════════════════════════════════════════════════════════════════
#  CIERRES HISTÓRICOS — leer + escribir
# ══════════════════════════════════════════════════════════════════════════════

def _nivel_cumplimiento(c, id_indicador=None) -> str:
    """Clasifica un valor de cumplimiento decimal en nivel de semáforo.
    
    Usa core.semantica para consistencia con el resto del sistema.
    """
    categoria = categorizar_cumplimiento(c, id_indicador=id_indicador)
    return categoria


def leer_consolidado_cierres(ruta: str, ids_activos: set) -> pd.DataFrame:
    """
    Lee la hoja 'Consolidado Cierres' de Resultados Consolidados.xlsx.
    Filtra a los IDs activos indicados, normaliza columnas clave y agrega
    la columna 'Nivel' calculada desde 'Cumplimiento'.
    """
    if not os.path.exists(ruta):
        print(f"    INFO: {ruta} no encontrado — hojas de cierres omitidas.")
        return pd.DataFrame()
    try:
        df = pd.read_excel(ruta, sheet_name="Consolidado Cierres", engine="openpyxl")
    except Exception as exc:
        print(f"    ADVERTENCIA Consolidado Cierres: {exc}")
        return pd.DataFrame()

    df.columns = [str(c).strip() for c in df.columns]

    # Normalizar Id
    col_id = _detectar_col(df, ["Id", "ID", "id"])
    if col_id:
        df[col_id] = df[col_id].apply(_id_normalizar)
        if col_id != "Id":
            df = df.rename(columns={col_id: "Id"})

    # Normalizar Fecha
    col_fec = _detectar_col(df, ["Fecha", "fecha"])
    if col_fec:
        df[col_fec] = pd.to_datetime(df[col_fec], errors="coerce")
        if col_fec != "Fecha":
            df = df.rename(columns={col_fec: "Fecha"})

    # Normalizar columnas de cumplimiento
    for col in ["Cumplimiento", "Cumplimiento Real"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Filtrar a IDs activos
    if ids_activos and "Id" in df.columns:
        df = df[df["Id"].isin(ids_activos)].copy()

    # Agregar columna Nivel
    if "Cumplimiento" in df.columns:
        df["Nivel"] = df["Cumplimiento"].apply(_nivel_cumplimiento)

    # Ordenar Id numérico → Fecha ascendente
    if "Fecha" in df.columns:
        df = df.sort_values(
            ["Id", "Fecha"],
            key=lambda s: pd.to_numeric(s, errors="coerce") if s.name == "Id" else s
        ).reset_index(drop=True)

    return df


def escribir_consolidado_cierres(ws, df: pd.DataFrame):
    """
    Escribe la hoja 'Consolidado Cierres' con color en columna Nivel
    y formato porcentual en Cumplimiento.
    """
    if df.empty:
        ws.cell(row=1, column=1, value="Sin datos de cierres históricos.")
        return

    # Columnas a mostrar (priorizadas + resto)
    prio = ["Id", "Indicador", "Proceso", "Periodicidad", "Sentido",
            "Fecha", "Periodo", "Meta", "Ejecucion",
            "Cumplimiento", "Cumplimiento Real", "Nivel"]
    cols = [c for c in prio if c in df.columns]
    cols += [c for c in df.columns if c not in cols and c not in
             {"Llave", "Meta_Signo", "Ejecucion_Signo",
              "Decimales_Meta", "Decimales_Ejecucion", "Tipo_Registro"}]
    df_out = df[cols].copy()

    # Encabezados
    for ci, col in enumerate(df_out.columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        _estilo_header(cell)
        ws.column_dimensions[get_column_letter(ci)].width = max(len(str(col)) + 4, 12)

    col_nivel_idx = df_out.columns.tolist().index("Nivel") + 1 if "Nivel" in df_out.columns else None
    col_cum_idx   = df_out.columns.tolist().index("Cumplimiento") + 1 if "Cumplimiento" in df_out.columns else None

    for ri, row in enumerate(df_out.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            col_name = df_out.columns[ci - 1]
            # Fecha → objeto date nativo para que Excel la reconozca
            if col_name == "Fecha" and pd.notna(val):
                try:
                    val = pd.Timestamp(val).date()
                except Exception:
                    pass
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Color Nivel
        if col_nivel_idx:
            nivel_val = str(row[col_nivel_idx - 1])
            bg = C_NIVEL_BG.get(nivel_val)
            if bg:
                ws.cell(row=ri, column=col_nivel_idx).fill = PatternFill("solid", fgColor=bg)

        # Color Cumplimiento (degradado por nivel)
        if col_cum_idx:
            cum_val = row[col_cum_idx - 1]
            nivel_c = _nivel_cumplimiento(cum_val)
            bg_c = C_NIVEL_BG.get(nivel_c)
            if bg_c:
                ws.cell(row=ri, column=col_cum_idx).fill = PatternFill("solid", fgColor=bg_c)

    ws.freeze_panes = "G2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df_out.columns))}1"


def escribir_matriz_calor(ws, df_cierres: pd.DataFrame):
    """
    Escribe la hoja 'Matriz Calor': indicadores (filas) × períodos de cierre (columnas).
    Cada celda muestra el % de cumplimiento con color de fondo según nivel.
    Columnas fijas izquierda: Id | Indicador | Proceso | Periodicidad
    Columnas dinámicas: una por Fecha de cierre (dd/mm/AAAA), ordenadas ascendente.
    """
    if df_cierres.empty or "Fecha" not in df_cierres.columns or "Id" not in df_cierres.columns:
        ws.cell(row=1, column=1, value="Sin datos para la matriz de calor.")
        return

    # ── Metadata fija por indicador (primera aparición) ──────────────────────
    META_COLS = [c for c in ["Indicador", "Proceso", "Periodicidad"] if c in df_cierres.columns]
    meta = (df_cierres.drop_duplicates(subset=["Id"], keep="first")
            .set_index("Id")[META_COLS])

    # ── Pivot Id × Fecha → Cumplimiento (último valor si hay duplicados) ──────
    pivot = (df_cierres.dropna(subset=["Fecha", "Cumplimiento"])
             .pivot_table(index="Id", columns="Fecha",
                          values="Cumplimiento", aggfunc="last"))
    pivot.columns = pd.to_datetime(pivot.columns)
    pivot = pivot.sort_index(axis=1)  # fechas ascendente

    # ── Unir metadata + pivot ─────────────────────────────────────────────────
    result = meta.join(pivot, how="outer").reset_index()

    # Ordenar filas por Id numérico
    result["_sort"] = pd.to_numeric(result["Id"], errors="coerce")
    result = result.sort_values("_sort").drop(columns=["_sort"]).reset_index(drop=True)

    # ── Columnas del Excel ────────────────────────────────────────────────────
    fixed_cols = ["Id"] + META_COLS
    date_cols  = [c for c in result.columns if isinstance(c, pd.Timestamp)]
    all_cols   = fixed_cols + date_cols

    # ── Encabezados ───────────────────────────────────────────────────────────
    ANCHOS_FIJOS = {"Id": 8, "Indicador": 40, "Proceso": 28, "Periodicidad": 14}
    for ci, col in enumerate(all_cols, 1):
        if isinstance(col, pd.Timestamp):
            label = col.strftime("%d/%m/%Y")
            ancho = 13
        else:
            label = col
            ancho = ANCHOS_FIJOS.get(col, max(len(str(col)) + 3, 12))
        cell = ws.cell(row=1, column=ci, value=label)
        _estilo_header(cell)
        ws.column_dimensions[get_column_letter(ci)].width = ancho

    # ── Datos ─────────────────────────────────────────────────────────────────
    n_fixed = len(fixed_cols)
    for ri, row in enumerate(result.itertuples(index=False), 2):
        row_dict = dict(zip(all_cols, row))
        for ci, col in enumerate(all_cols, 1):
            val = row_dict.get(col)
            if ci <= n_fixed:
                # Columnas fijas: texto plano
                cell = ws.cell(row=ri, column=ci, value=val if pd.notna(val) else "")
                cell.alignment = Alignment(horizontal="left" if ci > 1 else "center",
                                           vertical="center")
            else:
                # Columna de cumplimiento: mostrar como % y colorear
                if pd.notna(val):
                    pct_str = f"{float(val) * 100:.1f}%"
                    nivel   = _nivel_cumplimiento(val)
                    bg      = C_NIVEL_BG.get(nivel, "FFFFFF")
                    cell = ws.cell(row=ri, column=ci, value=pct_str)
                    cell.fill = PatternFill("solid", fgColor=bg)
                else:
                    cell = ws.cell(row=ri, column=ci, value="")
                    cell.fill = PatternFill("solid", fgColor="F5F5F5")
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Formato final ─────────────────────────────────────────────────────────
    freeze_col = get_column_letter(n_fixed + 1)
    ws.freeze_panes = f"{freeze_col}2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_cols))}1"


# ══════════════════════════════════════════════════════════════════════════════
#  HOJA RESUMEN
# ══════════════════════════════════════════════════════════════════════════════

def crear_hoja_resumen(wb, resumen_data: list):
    ws = wb.create_sheet(title="Resumen", index=0)
    headers = ["Periodicidad", "Total indicadores", "Reportados (período actual)",
               "Pendientes de reporte", "% Reporte"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        _estilo_header(cell)
        ws.column_dimensions[get_column_letter(ci)].width = max(len(h) + 4, 18)

    for ri, row in enumerate(resumen_data, 2):
        ws.cell(row=ri, column=1, value=row["Periodicidad"])
        ws.cell(row=ri, column=2, value=row["Total"])
        ws.cell(row=ri, column=3, value=row["Reportados"])
        ws.cell(row=ri, column=4, value=row["Pendientes"])
        pct = round(row["Reportados"] / row["Total"] * 100, 1) if row["Total"] else 0
        pct_cell = ws.cell(row=ri, column=5, value=f"{pct}%")
        if pct >= 80:
            pct_cell.fill = PatternFill("solid", fgColor=C_SI)
        elif pct >= 50:
            pct_cell.fill = PatternFill("solid", fgColor=C_PENDIENTE)
        else:
            pct_cell.fill = PatternFill("solid", fgColor=C_NO)


# ══════════════════════════════════════════════════════════════════════════════
#  FLUJO PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

def main():
    sep = "=" * 66
    fecha_ref = get_fecha_referencia()
    hoy = date.today()

    print(sep)
    print("  generar_reporte.py (dinámico 2026) → Seguimiento_Reporte.xlsx")
    print(sep)
    print(f"  Fecha de ejecución : {hoy.strftime('%d/%m/%Y')}")
    print(f"  Fecha de referencia: {fecha_ref.strftime('%d/%m/%Y')}  "
          f"({'manual' if FECHA_REFERENCIA_MANUAL else 'automática — mes anterior'})")
    print(f"  Inicio calendario  : {FECHA_INICIO.strftime('%m/%Y')}")

    # ── 1. Leer fuente LMI ───────────────────────────────────────────────────
    print(f"\n[1] Leyendo fuente LMI: {RUTA_ORIGEN}")
    if not os.path.exists(RUTA_ORIGEN):
        sys.exit(f"    ERROR: No se encontró {RUTA_ORIGEN}")

    df = leer_xlsx(RUTA_ORIGEN)
    print(f"    OK → {len(df)} filas  |  {len(df.columns)} columnas")

    # ── 2. Columna Revisar ───────────────────────────────────────────────────
    if COLUMNA_REVISAR not in df.columns:
        sys.exit(f"    ERROR: Columna '{COLUMNA_REVISAR}' no encontrada en la fuente.")

    print(f"\n[2] Columna 'Revisar' (basada en '{COLUMNA_REVISAR}')...")
    df = agregar_revisar(df, COLUMNA_REVISAR)
    indicadores_unicos = df["Revisar"].sum()
    print(f"    {indicadores_unicos} indicadores únicos detectados.")

    # ── 3. Identificar columnas de período ───────────────────────────────────
    periodo_cols = [c for c in df.columns if str(c).startswith("Periodo ")]
    if not periodo_cols:
        sys.exit("    ERROR: No se encontraron columnas 'Periodo X' en la fuente.")
    n_periodos = len(periodo_cols)
    print(f"\n[3] Columnas de período: {n_periodos}  ({periodo_cols[0]} … {periodo_cols[-1]})")

    # ── 4. Crear workbook ────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── 5. Hoja "Seguimiento" ────────────────────────────────────────────────
    print(f"\n[4] Hoja 'Seguimiento' → {len(df)} filas...")
    ws_seg = wb.create_sheet(title="Seguimiento")
    escribir_hoja(ws_seg, df)
    print("    OK")

    # ── 6. Cargar Kawak — resultados históricos (API) ────────────────────────
    print(f"\n[5] Cargando Kawak API (resultados históricos)...")
    print(f"    {RUTA_KAWAK_API}")
    kawak_lookup = leer_kawak_api(RUTA_KAWAK_API)

    # ── 7. Hojas por periodicidad ────────────────────────────────────────────
    periodicidades = [
        p for p in df["Periodicidad"].dropna().unique()
        if p and str(p).strip() not in ("", "nan", "NaN", "None")
    ]
    print(f"\n[6] Periodicidades detectadas: {periodicidades}")
    resumen_data = []

    for perio in periodicidades:
        df_p = df[df["Periodicidad"] == perio].copy().reset_index(drop=True)
        total = len(df_p)
        print(f"\n    ── {perio} ({total} filas)")

        fechas = get_period_dates(perio, n_periodos, fecha_ref)
        mapa = {col: f for col, f in zip(periodo_cols, fechas)}
        if fechas[0]:
            ultimo = fechas[-1].strftime('%d/%m/%Y') if fechas[-1] else "N/A"
            print(f"       Periodo 1 → {fechas[0].strftime('%d/%m/%Y')}  "
                  f"| Periodo {n_periodos} → {ultimo}")

        col_p1 = periodo_cols[0]
        col_p2 = periodo_cols[1] if n_periodos >= 2 else None
        df_p = agregar_columnas_seguimiento(df_p, col_p1, col_p2)

        fecha_p1 = fechas[0] if fechas else None
        fecha_p2 = fechas[1] if len(fechas) > 1 else None
        df_p, n_kawak = enriquecer_desde_kawak(
            df_p, kawak_lookup, col_p1, col_p2, fecha_p1, fecha_p2
        )

        reportados = (df_p["Reportado"] == "Sí").sum()
        pendientes = (df_p["Estado del indicador"] == "Pendiente de reporte").sum()
        print(f"       Reportados (LMI)          : {reportados - n_kawak}/{total}")
        if kawak_lookup:
            print(f"       Actualizados desde Kawak : {n_kawak}")
        print(f"       Reportados total          : {reportados}/{total}")
        print(f"       Pendientes de reporte     : {pendientes}/{total}")

        ws = wb.create_sheet(title=perio[:31])
        escribir_hoja(ws, df_p, mapa)

        resumen_data.append({
            "Periodicidad": perio,
            "Total": total,
            "Reportados": int(reportados),
            "Pendientes": int(pendientes),
        })

    # ── 8. Hoja "Resumen" ────────────────────────────────────────────────────
    print("\n[7] Hoja 'Resumen'...")
    crear_hoja_resumen(wb, resumen_data)
    print("    OK")

    # ── 9. Tracking Mensual ──────────────────────────────────────────────────
    print(f"\n[8] Hoja 'Tracking Mensual' (catálogo año-específico, formato largo)...")

    # Calendario dinámico: FECHA_INICIO → fecha_ref
    calendario = generar_calendario(FECHA_INICIO, fecha_ref)
    print(f"    Calendario: {calendario[0][1]}/{calendario[0][0]} → "
          f"{calendario[-1][1]}/{calendario[-1][0]}  ({len(calendario)} meses)")

    # Catálogos año-específicos desde data/raw/Kawak/{year}.xlsx
    # Un indicador inactivo en 2026 no genera filas para 2026
    years_en_calendario = sorted({y for y, _ in calendario})
    catalogo_por_anio: Dict[int, Dict[str, str]] = {}
    meta_por_anio:     Dict[int, Dict[str, dict]] = {}
    for year in years_en_calendario:
        cat, met = leer_catalogo_por_anio(RUTA_KAWAK_DIR, year)
        if cat:
            catalogo_por_anio[year] = cat
            meta_por_anio[year]     = met
        else:
            print(f"    ADVERTENCIA: sin catálogo para {year} — meses de ese año omitidos.")

    total_ids = sum(len(c) for c in catalogo_por_anio.values())
    print(f"    Catálogos: {list(catalogo_por_anio.keys())}  "
          f"→ {total_ids} entradas ({', '.join(f'{y}:{len(c)}' for y,c in catalogo_por_anio.items())})")

    # Catálogo unificado para normalizar el lookup Kawak (periodicidad por ID)
    catalogo_unificado: Dict[str, str] = {}
    for cat in catalogo_por_anio.values():
        for kid, perio in cat.items():
            catalogo_unificado.setdefault(kid, perio)

    # Normalizar Kawak lookup al mes de CIERRE del período según periodicidad
    kawak_lookup_norm = normalizar_kawak_lookup(kawak_lookup, catalogo_unificado)
    print(f"    Kawak normalizado: {len(kawak_lookup_norm)} registros "
          f"(raw: {len(kawak_lookup)}  →  por período de cierre)")

    # Lookup LMI: (id, year, month) → True  (fuente secundaria)
    lmi_lookup = construir_lmi_lookup(df, fecha_ref)
    print(f"    Registros LMI con dato: {len(lmi_lookup)}")

    # Construir tabla larga con catálogo año-específico
    tracking_df = construir_tracking_largo(
        catalogo_por_anio, kawak_lookup_norm, lmi_lookup, calendario, meta_por_anio
    )

    if not tracking_df.empty:
        reportados_t = (tracking_df["Estado"] == "Reportado").sum()
        pendientes_t = (tracking_df["Estado"] == "Pendiente").sum()
        kawak_t      = (tracking_df["Fuente"]  == "Kawak").sum()
        lmi_t        = (tracking_df["Fuente"]  == "LMI").sum()
        print(f"    Filas generadas : {len(tracking_df):,}  "
              f"({len(catalogo_unificado)} indicadores únicos × períodos que aplican)")
        print(f"    Reportado: {reportados_t:,}  "
              f"(Kawak: {kawak_t:,} | LMI: {lmi_t:,})  |  "
              f"Pendiente: {pendientes_t:,}")

    ws_track = wb.create_sheet(title="Tracking Mensual")
    escribir_tracking_mensual(ws_track, tracking_df)
    print("    OK")

    # ── 10. Consolidado Cierres + Matriz Calor ───────────────────────────────
    print(f"\n[9] Hojas de cierres históricos...")
    print(f"    Fuente: {RUTA_CONSOLIDADOS}")
    df_cierres = leer_consolidado_cierres(RUTA_CONSOLIDADOS, set(catalogo_unificado.keys()))

    if not df_cierres.empty:
        print(f"    Consolidado Cierres: {len(df_cierres)} registros "
              f"| {df_cierres['Id'].nunique()} IDs "
              f"| {df_cierres['Fecha'].nunique()} fechas de cierre")

        ws_cierres = wb.create_sheet(title="Consolidado Cierres")
        escribir_consolidado_cierres(ws_cierres, df_cierres)
        print("    Hoja 'Consolidado Cierres' → OK")

        ws_calor = wb.create_sheet(title="Matriz Calor")
        escribir_matriz_calor(ws_calor, df_cierres)
        print("    Hoja 'Matriz Calor'       → OK")
    else:
        print("    ADVERTENCIA: sin datos de cierres — hojas omitidas.")

    # ── 11. Guardar ───────────────────────────────────────────────────────────
    os.makedirs(os.path.dirname(RUTA_SALIDA), exist_ok=True)
    print(f"\n[10] Guardando en {RUTA_SALIDA}...")
    wb.save(RUTA_SALIDA)
    print("    OK → archivo guardado.")

    print(f"\n{sep}")
    print("  Proceso completado exitosamente.")
    print(f"  Archivo: {RUTA_SALIDA}")
    print(sep)


if __name__ == "__main__":
    main()
