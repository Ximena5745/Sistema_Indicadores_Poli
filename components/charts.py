"""
components/charts.py — Componentes de gráficos reutilizables (Plotly + Streamlit).
"""
import io
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl.styles import PatternFill, Font

from core.config import COLOR_CATEGORIA, COLOR_CATEGORIA_CLARO
from services.data_loader import cargar_analisis_usuarios
from services.ai_analysis import analizar_texto_indicador

# Alias para compatibilidad interna — fuente única en core/config.py
COLOR_CAT = COLOR_CATEGORIA


def grafico_historico_indicador(df_ind: pd.DataFrame, titulo: str = "") -> go.Figure:
    """
    Gráfico de línea con zonas de color para el histórico de un indicador.
    df_ind debe tener: Cumplimiento_norm, Categoria, Periodo, Fecha.
    """
    if df_ind.empty:
        fig = go.Figure()
        fig.update_layout(title="Sin datos disponibles")
        return fig

    df_ind = df_ind.sort_values("Fecha").copy()
    cum_pct = (df_ind["Cumplimiento_norm"] * 100).round(1)
    y_max = max(130, float(cum_pct.max()) + 15) if not cum_pct.empty else 130

    colores_pts = df_ind["Categoria"].map(COLOR_CAT).fillna("#9E9E9E")

    fig = go.Figure()

    # Zonas de fondo
    fig.add_hrect(
        y0=0, y1=80, fillcolor="#FFCDD2", opacity=0.45, line_width=0,
        annotation_text="Peligro 0–80%", annotation_position="top left",
        annotation_font_size=10,
    )
    fig.add_hrect(
        y0=80, y1=100, fillcolor="#FFF9C4", opacity=0.45, line_width=0,
        annotation_text="Alerta 80–100%", annotation_position="top left",
        annotation_font_size=10,
    )
    fig.add_hrect(
        y0=100, y1=105, fillcolor="#E8F5E9", opacity=0.50, line_width=0,
        annotation_text="Cumplimiento 100–105%", annotation_position="top left",
        annotation_font_size=10,
    )
    fig.add_hrect(
        y0=105, y1=y_max, fillcolor="#D0E4FF", opacity=0.45, line_width=0,
        annotation_text="Sobrecumplimiento > 105%", annotation_position="top left",
        annotation_font_size=10,
    )

    # Línea de datos
    x_vals = df_ind["Fecha"] if "Fecha" in df_ind.columns else df_ind["Periodo"]
    mes_labels = (
        df_ind["Fecha"].dt.strftime("%b %Y") if "Fecha" in df_ind.columns
        else df_ind.get("Periodo", pd.Series([""] * len(df_ind)))
    )
    custom = list(zip(mes_labels, df_ind["Categoria"])) if "Categoria" in df_ind.columns else None

    fig.add_trace(go.Scatter(
        x=x_vals,
        y=cum_pct,
        mode="lines+markers+text",
        line=dict(color="#455A64", width=2),
        marker=dict(
            size=12,
            color=colores_pts,
            line=dict(width=2, color="white"),
        ),
        text=cum_pct.astype(str) + "%",
        textposition="top center",
        textfont=dict(size=9),
        hovertemplate=(
            "<b>Mes:</b> %{customdata[0]}<br>"
            "<b>Cumplimiento:</b> %{y:.1f}%<br>"
            "<b>Estado:</b> %{customdata[1]}<extra></extra>"
        ) if custom is not None else "%{y:.1f}%<extra></extra>",
        customdata=custom,
        showlegend=False,
    ))

    # Líneas de referencia
    fig.add_hline(y=100, line_dash="dash", line_color="#2E7D32", line_width=1.5)
    fig.add_hline(y=80,  line_dash="dot",  line_color="#C62828", line_width=1.5)

    # Leyenda de colores
    for cat, color in COLOR_CAT.items():
        if cat == "Sin dato":
            continue
        fig.add_trace(go.Scatter(
            x=[None], y=[None], mode="markers",
            marker=dict(size=10, color=color),
            name=cat, showlegend=True,
        ))

    fig.update_layout(
        title=titulo,
        yaxis=dict(title="Cumplimiento (%)", ticksuffix="%", range=[0, y_max]),
        xaxis_title="",
        plot_bgcolor="white",
        paper_bgcolor="white",
        legend=dict(orientation="h", y=-0.2),
        margin=dict(t=50, b=60),
    )
    return fig


def tabla_historica_indicador(df_ind: pd.DataFrame) -> pd.DataFrame:
    """Prepara DataFrame histórico con columnas formateadas para mostrar."""
    df_work = df_ind.copy()

    # Generar columna Mes desde Fecha (reemplaza Periodo)
    if "Fecha" in df_work.columns:
        df_work["Mes"] = pd.to_datetime(df_work["Fecha"], errors="coerce").dt.strftime("%b %Y")
    elif "Periodo" in df_work.columns:
        df_work["Mes"] = df_work["Periodo"].astype(str)

    # Renombrar Anio -> Año
    if "Anio" in df_work.columns:
        df_work.rename(columns={"Anio": "Año"}, inplace=True)

    cols = ["Mes", "Año", "Meta", "Ejecucion", "Cumplimiento_norm", "Categoria"]
    cols_disp = [c for c in cols if c in df_work.columns]
    df_t = df_work[cols_disp].copy()
    if "Año" in df_t.columns:
        df_t = df_t.sort_values("Año")

    for col in ("Meta", "Ejecucion"):
        if col in df_t.columns:
            df_t[col] = pd.to_numeric(df_t[col], errors="coerce").map(
                lambda v: "" if pd.isna(v) else (f"{v:,.0f}" if v == int(v) else f"{v:,.2f}")
            )
    if "Cumplimiento_norm" in df_t.columns:
        def _fmt_cumpl(v):
            try:
                return f"{round(float(v) * 100, 1)}%" if pd.notna(v) else "No aplica"
            except (TypeError, ValueError):
                return "No aplica"
        df_t["Cumplimiento_norm"] = df_t["Cumplimiento_norm"].apply(_fmt_cumpl)
        df_t.rename(columns={"Cumplimiento_norm": "Cumplimiento%"}, inplace=True)
    return df_t


def exportar_excel(df: pd.DataFrame, nombre_hoja: str = "Datos") -> bytes:
    """Retorna bytes xlsx con formato institucional para st.download_button."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=nombre_hoja)
        ws = writer.sheets[nombre_hoja]
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1A3A5C")
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
    return output.getvalue()


def colorear_tabla_categoria(df: pd.DataFrame, col_categoria: str = "Categoria") -> pd.DataFrame.style:
    """Aplica estilos de color a filas según categoría (colores desde core/config.py)."""
    def estilo_fila(row):
        bg = COLOR_CATEGORIA_CLARO.get(row.get(col_categoria, ""), "")
        return [f"background-color: {bg}" if bg else "" for _ in row]

    return df.style.apply(estilo_fila, axis=1)


def grafico_detalle_indicador(df_ind: pd.DataFrame) -> go.Figure:
    """
    Gráfico combinado para el modal de detalle:
    - Barras agrupadas: Meta y Ejecución (eje izquierdo)
    - Línea con marcadores: Cumplimiento% (eje derecho)
    - X-axis: etiquetas de Periodo (solo períodos con datos, formato legible)
    """
    df = df_ind.sort_values("Fecha").copy()

    # Usar mes específico desde Fecha en lugar del código de Periodo
    if "Fecha" in df.columns:
        periodos = pd.to_datetime(df["Fecha"], errors="coerce").dt.strftime("%b %Y").tolist()
    else:
        periodos = df["Periodo"].astype(str).tolist()
    cum_pct = (df["Cumplimiento_norm"] * 100).round(1)
    colores_pts = df["Categoria"].map(COLOR_CAT).fillna("#9E9E9E").tolist()

    meta_vals = pd.to_numeric(df.get("Meta", pd.Series(dtype=float)), errors="coerce")
    ejec_vals = pd.to_numeric(df.get("Ejecucion", pd.Series(dtype=float)), errors="coerce")

    fig = go.Figure()

    # Barras Meta y Ejecución
    if meta_vals.notna().any():
        fig.add_trace(go.Bar(
            x=periodos, y=meta_vals.tolist(),
            name="Meta", marker_color="#90CAF9", opacity=0.85, yaxis="y1",
            hovertemplate="<b>%{x}</b><br>Meta: %{y:,.2f}<extra></extra>",
        ))
    if ejec_vals.notna().any():
        fig.add_trace(go.Bar(
            x=periodos, y=ejec_vals.tolist(),
            name="Ejecución", marker_color="#1A3A5C", opacity=0.85, yaxis="y1",
            hovertemplate="<b>%{x}</b><br>Ejecución: %{y:,.2f}<extra></extra>",
        ))

    # Línea Cumplimiento%
    fig.add_trace(go.Scatter(
        x=periodos, y=cum_pct.tolist(),
        name="Cumplimiento %", mode="lines+markers+text",
        line=dict(color="#E65100", width=2.5),
        marker=dict(size=10, color=colores_pts, line=dict(width=2, color="white")),
        text=(cum_pct.astype(str) + "%").tolist(),
        textposition="top center",
        textfont=dict(size=9),
        yaxis="y2",
        hovertemplate="<b>%{x}</b><br>Cumplimiento: %{y:.1f}%<extra></extra>",
    ))

    y_max_cum = max(130.0, float(cum_pct.max()) + 15) if not cum_pct.dropna().empty else 130.0
    for y_ref, color_ref, dash_ref in [
        (100, "#2E7D32", "dash"),
        (80,  "#C62828", "dot"),
    ]:
        fig.add_shape(
            type="line", xref="paper", x0=0, x1=1,
            yref="y2", y0=y_ref, y1=y_ref,
            line=dict(color=color_ref, width=1.5, dash=dash_ref),
        )

    for y0_z, y1_z, fill_z in [
        (0, 80, "#FFCDD2"), (80, 100, "#FFF9C4"),
        (100, 105, "#E8F5E9"), (105, y_max_cum, "#D0E4FF"),
    ]:
        fig.add_shape(
            type="rect", xref="paper", x0=0, x1=1,
            yref="y2", y0=y0_z, y1=min(y1_z, y_max_cum),
            fillcolor=fill_z, opacity=0.25, line_width=0,
            layer="below",
        )

    fig.update_layout(
        barmode="group",
        height=440,
        yaxis=dict(title="Meta / Ejecución", side="left", showgrid=False),
        yaxis2=dict(
            title="Cumplimiento (%)", ticksuffix="%",
            overlaying="y", side="right",
            range=[0, y_max_cum],
            showgrid=True, gridcolor="#EEEEEE",
        ),
        xaxis=dict(title="Período", type="category", tickangle=-30),
        plot_bgcolor="white",
        paper_bgcolor="white",
        legend=dict(orientation="h", y=-0.3, x=0),
        margin=dict(t=30, b=90, l=70, r=70),
    )
    return fig


def panel_detalle_indicador(df_ind: pd.DataFrame, id_ind: str, df_full: pd.DataFrame):
    """
    Renderiza el panel completo de detalle de un indicador.
    Llamar desde un st.dialog o st.expander.
    """
    from core.calculos import generar_recomendaciones

    # Disable backdrop click (clicking outside dialog should not close it)
    st.markdown("""
        <style>
        [data-baseweb="modal"] > div:first-child        { pointer-events: none !important; }
        [data-baseweb="modal"] [data-baseweb="dialog"]  { pointer-events: auto !important; }
        </style>
    """, unsafe_allow_html=True)

    if df_ind.empty:
        st.warning("Sin datos para este indicador.")
        return

    df_ind_sorted = df_ind.sort_values("Fecha")
    ultimo = df_ind_sorted.iloc[-1]

    nombre        = ultimo.get("Indicador", "—")
    proceso       = ultimo.get("Proceso", "—")
    subproceso    = ultimo.get("Subproceso", "—")
    periodicidad  = ultimo.get("Periodicidad", "—")
    clasificacion = ultimo.get("Clasificacion", "—")
    cum_norm      = ultimo.get("Cumplimiento_norm", None)
    categoria     = ultimo.get("Categoria", "Sin dato")

    cum_pct_str = f"{round(float(cum_norm)*100, 1)}%" if pd.notna(cum_norm) else "—"

    BADGE_COLOR = {
        "Peligro":          "#C62828",
        "Alerta":           "#F57F17",
        "Cumplimiento":     "#2E7D32",
        "Sobrecumplimiento":"#0277BD",
        "Sin dato":         "#9E9E9E",
    }
    badge_col = BADGE_COLOR.get(categoria, "#9E9E9E")

    # Contenedor con scroll que cubre toda la ficha
    with st.container(height=620, border=False):
        st.markdown(f"### {id_ind} — {nombre}")

        c1, c2, c3, c4 = st.columns([2, 2, 2, 1])
        with c1:
            st.markdown(f"**Proceso:** {proceso}")
        with c2:
            st.markdown(f"**Subproceso:** {subproceso}")
        with c3:
            st.markdown(f"**Periodicidad:** {periodicidad} · {clasificacion}")
        with c4:
            st.markdown(
                f"<div style='font-size:1.5rem; font-weight:bold;'>{cum_pct_str}</div>"
                f"<span style='background:{badge_col};color:white;padding:2px 10px;"
                f"border-radius:12px;font-size:0.85rem'>{categoria}</span>",
                unsafe_allow_html=True,
            )

        st.divider()

        df_tabla = tabla_historica_indicador(df_ind_sorted)
        st.markdown("**Histórico**")

        def _color_cumpl(row):
            cat = row.get("Categoria", "")
            bg  = COLOR_CATEGORIA_CLARO.get(cat, "")
            fg  = COLOR_CATEGORIA.get(cat, "")
            styles = [""] * len(row)
            if "Cumplimiento%" in row.index:
                i = row.index.get_loc("Cumplimiento%")
                if bg:
                    styles[i] = f"background-color:{bg};color:{fg};font-weight:bold"
            return styles

        st.dataframe(
            df_tabla.style.apply(_color_cumpl, axis=1),
            use_container_width=True,
            hide_index=True,
        )

        st.markdown("**Evolución: Meta, Ejecución y Cumplimiento**")
        fig = grafico_detalle_indicador(df_ind_sorted)
        st.plotly_chart(fig, use_container_width=True)

        st.divider()
        cum_series = df_ind_sorted["Cumplimiento_norm"].dropna() * 100
        tendencia, recs = generar_recomendaciones(categoria, cum_series)

        tendencia_labels = {
            "Mejorando":            "📈 Tendencia creciente",
            "Empeorando":           "📉 Tendencia decreciente",
            "Estable":              "➡️ Sin variación significativa",
            "Sin datos suficientes":"ℹ️ Datos insuficientes para análisis",
        }
        st.markdown(f"**Análisis de tendencia:** {tendencia_labels.get(tendencia, tendencia)}")
        st.markdown("**Recomendaciones:**")
        for rec in recs:
            st.markdown(f"- {rec}")

        # ── Análisis del usuario (último corte registrado) ────────────────────
        df_analisis = cargar_analisis_usuarios()
        if not df_analisis.empty:
            df_ind_an = df_analisis[df_analisis["Id"] == str(id_ind)]
            if not df_ind_an.empty:
                ultimo_an = df_ind_an.sort_values("analisis_fecha").iloc[-1]
                texto     = str(ultimo_an.get("analisis_texto", "")).strip()
                autor     = str(ultimo_an.get("analisis_autor", "")).strip()
                fecha_an  = ultimo_an.get("analisis_fecha")
                fecha_str = fecha_an.strftime("%d/%m/%Y") if pd.notna(fecha_an) else ""

                if texto and texto.lower() not in ("nan", "none", ""):
                    st.divider()
                    st.markdown(
                        f"**Análisis del responsable**"
                        + (f" · {autor}" if autor and autor not in ("nan", "None") else "")
                        + (f" · {fecha_str}" if fecha_str else ""),
                    )
                    st.info(texto)

                    # ── Insights IA ───────────────────────────────────────────
                    with st.expander("✨ Insights y oportunidades de mejora (IA)", expanded=True):
                        with st.spinner("Analizando..."):
                            ia_result = analizar_texto_indicador(
                                id_ind=str(id_ind),
                                nombre=nombre,
                                proceso=proceso,
                                categoria=categoria,
                                cumplimiento=cum_pct_str,
                                texto_analisis=texto,
                            )
                        if ia_result:
                            st.markdown(ia_result)
                        else:
                            st.caption("⚙️ Configura ANTHROPIC_API_KEY en los secretos de Streamlit Cloud para activar este análisis.")


def grafico_3d_riesgo(df_cat: pd.DataFrame) -> go.Figure:
    """
    Scatter 3D: Proceso (eje X) × Cumplimiento% (eje Y) × Periodos en riesgo (eje Z).
    Coloreado por Categoria. Ideal para Tab 1 de Gestión OM.
    df_cat debe tener: Id, Indicador, Proceso, Cumplimiento_norm, Categoria
    y opcionalmente 'Periodos en riesgo'.
    """
    df = df_cat.copy()
    if df.empty:
        return go.Figure()

    # Mapeo proceso -> índice numérico para eje X
    procesos = sorted(df["Proceso"].dropna().unique().tolist())
    proc_map = {p: i for i, p in enumerate(procesos)}
    df["proc_idx"] = df["Proceso"].map(proc_map).fillna(0)
    df["cum_pct"]  = (df["Cumplimiento_norm"] * 100).round(1)
    df["per_riesgo"] = (df["Períodos en riesgo"].fillna(0).astype(int)
                        if "Períodos en riesgo" in df.columns
                        else pd.Series(0, index=df.index))

    colores_cat = df["Categoria"].map(COLOR_CAT).fillna("#9E9E9E").tolist()

    fig = go.Figure(go.Scatter3d(
        x=df["proc_idx"].tolist(),
        y=df["cum_pct"].tolist(),
        z=df["per_riesgo"].tolist(),
        mode="markers",
        marker=dict(
            size=6,
            color=colores_cat,
            opacity=0.85,
            line=dict(width=0.5, color="white"),
        ),
        customdata=list(zip(df["Id"].tolist(), df["Proceso"].tolist(),
                            df["Indicador"].str[:50].tolist())),
        hovertemplate=(
            "<b>%{customdata[2]}</b><br>"
            "Proceso: %{customdata[1]}<br>"
            "Cumplimiento: %{y:.1f}%<br>"
            "Periodos en riesgo: %{z}<extra></extra>"
        ),
    ))

    fig.update_layout(
        height=480,
        scene=dict(
            xaxis=dict(
                title="Proceso",
                tickvals=list(range(len(procesos))),
                ticktext=[p[:15] for p in procesos],
                tickfont=dict(size=9),
            ),
            yaxis=dict(title="Cumplimiento (%)", ticksuffix="%"),
            zaxis=dict(title="Periodos en riesgo"),
            bgcolor="white",
        ),
        margin=dict(l=0, r=0, t=30, b=0),
        paper_bgcolor="white",
    )
    return fig


def grafico_3d_om(df_om: pd.DataFrame, col_proc: str, col_av: str, col_dias: str,
                  col_estado: str) -> go.Figure:
    """
    Scatter 3D para seguimiento OM:
    X = Proceso (numérico), Y = Avance (%), Z = Dias vencida.
    Coloreado por Estado.
    """
    _COLORES_ESTADO = {
        "Cerrada":   "#43A047",
        "Ejecución": "#1976D2",
        "Abierta":   "#F9A825",
        "Retrasada": "#C62828",
    }
    cols_need = [col_proc, col_av, col_dias, col_estado]
    if not all(c and c in df_om.columns for c in cols_need):
        return go.Figure()

    df = df_om[cols_need].dropna(subset=[col_av, col_dias]).copy()
    if df.empty:
        return go.Figure()

    procesos  = sorted(df[col_proc].dropna().unique().tolist())
    proc_map  = {p: i for i, p in enumerate(procesos)}
    df["proc_idx"] = df[col_proc].map(proc_map).fillna(0)
    df["color"]    = df[col_estado].map(_COLORES_ESTADO).fillna("#9E9E9E")

    id_col = "Id" if "Id" in df_om.columns else None

    fig = go.Figure(go.Scatter3d(
        x=df["proc_idx"].tolist(),
        y=pd.to_numeric(df[col_av], errors="coerce").fillna(0).tolist(),
        z=pd.to_numeric(df[col_dias], errors="coerce").fillna(0).tolist(),
        mode="markers",
        marker=dict(size=6, color=df["color"].tolist(), opacity=0.85,
                    line=dict(width=0.5, color="white")),
        customdata=list(zip(df[col_proc].tolist(), df[col_estado].tolist())),
        hovertemplate=(
            "<b>Proceso:</b> %{customdata[0]}<br>"
            "<b>Estado:</b> %{customdata[1]}<br>"
            "Avance: %{y:.1f}%<br>"
            "Días vencida: %{z}<extra></extra>"
        ),
    ))

    fig.update_layout(
        height=460,
        scene=dict(
            xaxis=dict(
                title="Proceso",
                tickvals=list(range(len(procesos))),
                ticktext=[p[:12] for p in procesos],
                tickfont=dict(size=9),
            ),
            yaxis=dict(title="Avance (%)", ticksuffix="%"),
            zaxis=dict(title="Dias vencida"),
            bgcolor="white",
        ),
        margin=dict(l=0, r=0, t=30, b=0),
        paper_bgcolor="white",
    )
    return fig
