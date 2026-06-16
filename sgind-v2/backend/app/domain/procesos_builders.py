"""Constructores de datos para CMI por Procesos — paridad con Streamlit resumen_por_proceso."""

from __future__ import annotations

import math
import unicodedata
from datetime import date
from typing import Any

import pandas as pd

from app.domain.cmi_builders import (
    COLOR_CATEGORIA,
    build_alertas,
    build_distribucion_nivel,
    build_indicadores_listado,
    calcular_kpis,
    df_to_records,
)
from app.domain.resumen_builders import compute_trends, ensure_nivel_cumplimiento

MESES_OPCIONES = [
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
]

MES_MAP: dict[str, int] = {m.upper(): i + 1 for i, m in enumerate(MESES_OPCIONES)}

TIPOS_PROCESO = ["APOYO", "ESTRATÉGICO", "EVALUACIÓN Y MEJORA", "MISIONAL"]

TIPO_PROCESO_COLORS: dict[str, str] = {
    "APOYO": "#42A5F5",
    "ESTRATÉGICO": "#1A3A5C",
    "EVALUACIÓN Y MEJORA": "#FB8C00",
    "MISIONAL": "#43A047",
}

TIPO_PROCESO_COLORS_LIGHT: dict[str, str] = {
    "APOYO": "#E3F2FD",
    "ESTRATÉGICO": "#D0E4FF",
    "EVALUACIÓN Y MEJORA": "#FFF3E0",
    "MISIONAL": "#E8F5E9",
}

TIPO_PROCESO_ICONS: dict[str, str] = {
    "APOYO": "🔧",
    "ESTRATÉGICO": "🎯",
    "EVALUACIÓN Y MEJORA": "📊",
    "MISIONAL": "⭐",
}

STRATEGIC_PALETTE = ["#FBAF17", "#42F2F2", "#EC0677", "#1FB2DE", "#A6CE38", "#0F385A"]


def cumplimiento_semaforo_color(val: float | None) -> str:
    if val is None:
        return "#9E9E9E"
    if val >= 100:
        return "#2E7D32"
    if val >= 80:
        return "#F9A825"
    return "#C62828"


def cumplimiento_estado(val: float | None) -> dict[str, str]:
    if val is None:
        return {"label": "Sin dato", "icon": "—", "color": "#6B7280"}
    if val >= 100:
        return {"label": "Saludable", "icon": "🟢", "color": "#2E7D32"}
    if val >= 80:
        return {"label": "Alerta", "icon": "🟡", "color": "#F9A825"}
    return {"label": "Crítico", "icon": "🔴", "color": "#C62828"}


def _norm_text(value: object) -> str:
    txt = str(value or "").strip().upper()
    txt = unicodedata.normalize("NFD", txt)
    return "".join(c for c in txt if unicodedata.category(c) != "Mn")


def _safe_float(val: Any, default: float | None = None) -> float | None:
    if val is None:
        return default
    try:
        f = float(val)
        if math.isnan(f) or math.isinf(f):
            return default
        return f
    except (TypeError, ValueError):
        return default


def mes_to_num(value: object, periodicidad: object = None) -> float | None:
    if pd.isna(value):
        return None
    period_norm = str(periodicidad or "").strip().lower()
    if isinstance(value, (int, float)):
        try:
            v = int(value)
            if period_norm == "semestral":
                if v in (1, 2):
                    return 6.0
                if v == 7:
                    return 12.0
            return float(v) if 1 <= v <= 12 else None
        except Exception:
            return None
    txt = str(value).strip()
    if not txt:
        return None
    txt_norm = _norm_text(txt)
    if period_norm == "semestral":
        if txt_norm in ("1", "01", "ENERO", "ENE"):
            return 6.0
        if txt_norm in ("2", "02", "JULIO", "JUL", "7", "07"):
            return 12.0
    if txt.isdigit():
        v = int(txt)
        return float(v) if 1 <= v <= 12 else None
    return float(MES_MAP[txt_norm]) if txt_norm in MES_MAP else None


def mes_nombre(mes_num: int) -> str:
    if 1 <= mes_num <= 12:
        return MESES_OPCIONES[mes_num - 1]
    return "Diciembre"


def default_mes(tracking: pd.DataFrame, anio: int) -> int:
    prev = get_prev_month_for_year(tracking, anio)
    return prev if prev is not None else 12


def get_prev_month_for_year(tracking_df: pd.DataFrame, year: int) -> int | None:
    if tracking_df.empty or "Anio" not in tracking_df.columns:
        return None
    subset = tracking_df[pd.to_numeric(tracking_df["Anio"], errors="coerce") == int(year)].copy()
    if subset.empty or "Mes" not in subset.columns:
        return None
    nums = subset["Mes"].apply(lambda m: mes_to_num(m)).dropna()
    if nums.empty:
        return None
    return int(nums.max())


def _ensure_cumplimiento_pct(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "cumplimiento_pct" not in out.columns:
        if "Cumplimiento_norm" in out.columns:
            out["cumplimiento_pct"] = pd.to_numeric(out["Cumplimiento_norm"], errors="coerce") * 100
        elif "Cumplimiento" in out.columns:
            vals = pd.to_numeric(out["Cumplimiento"], errors="coerce")
            if not vals.dropna().empty and vals.max() <= 2:
                vals = vals * 100
            out["cumplimiento_pct"] = vals
    return out


def prepare_tracking(df: pd.DataFrame, map_df: pd.DataFrame) -> pd.DataFrame:
    """Enriquece tracking con Proceso_padre, Subproceso_final, Unidad y Tipo de proceso."""
    if df.empty:
        return df

    out = df.copy()
    if "Proceso" not in out.columns:
        out["Proceso"] = "Sin proceso"

    if not map_df.empty and {"Subproceso", "Proceso"}.issubset(map_df.columns):
        sub_map = (
            map_df[["Subproceso", "Proceso"] + (["Unidad"] if "Unidad" in map_df.columns else [])]
            .dropna(subset=["Subproceso"])
            .drop_duplicates(subset=["Subproceso"])
            .copy()
        )
        sub_map["sub_norm"] = sub_map["Subproceso"].astype(str).map(_norm_text)

        proc_col = "Subproceso" if "Subproceso" in out.columns else "Proceso"
        out["proc_norm"] = out[proc_col].astype(str).map(_norm_text)
        out = out.merge(
            sub_map.rename(columns={"Proceso": "Proceso_padre_map", "Unidad": "Unidad_map"}),
            left_on="proc_norm",
            right_on="sub_norm",
            how="left",
        )
        out["Proceso_padre"] = out["Proceso_padre_map"].fillna(out["Proceso"].astype(str))
        out["Subproceso_final"] = out.get("Subproceso", out["Proceso"]).fillna(out["Proceso"].astype(str))
        if "Unidad_map" in out.columns:
            out["Unidad"] = out["Unidad_map"].fillna("")
        out = out.drop(
            columns=[c for c in ["proc_norm", "sub_norm", "Proceso_padre_map", "Unidad_map"] if c in out.columns]
        )
    else:
        out["Proceso_padre"] = out["Proceso"].astype(str)
        out["Subproceso_final"] = out.get("Subproceso", out["Proceso"]).astype(str)
        if "Unidad" not in out.columns:
            out["Unidad"] = ""

    if not map_df.empty and "Tipo de proceso" in map_df.columns and "Subproceso" in map_df.columns:
        tipo_map = map_df[["Subproceso", "Tipo de proceso"]].drop_duplicates(subset=["Subproceso"])
        sub_col = "Subproceso_final" if "Subproceso_final" in out.columns else "Subproceso"
        if sub_col in out.columns:
            out = out.merge(
                tipo_map.rename(columns={"Subproceso": "sub_merge_key"}),
                left_on=sub_col,
                right_on="sub_merge_key",
                how="left",
            )
            if "Tipo de proceso" not in out.columns and "Tipo de proceso_y" in out.columns:
                out["Tipo de proceso"] = out["Tipo de proceso_y"]
            out = out.drop(columns=[c for c in ["sub_merge_key", "Tipo de proceso_y"] if c in out.columns])

    return _ensure_cumplimiento_pct(out)


def filter_by_anio_mes(df: pd.DataFrame, *, anio: int, mes: int) -> pd.DataFrame:
    if df.empty:
        return df
    out = df.copy()
    if "Anio" in out.columns:
        out = out[pd.to_numeric(out["Anio"], errors="coerce") == int(anio)]
    if "Mes" in out.columns and mes:
        out = out[out["Mes"].apply(mes_to_num) == float(mes)]
    return out


def apply_ui_filters(
    df: pd.DataFrame,
    *,
    unidad: str | None = None,
    proceso: str | None = None,
    subproceso: str | None = None,
    clasificacion: str | None = None,
    frecuencia: str | None = None,
) -> pd.DataFrame:
    if df.empty:
        return df
    out = df.copy()
    if unidad and unidad != "Todos" and "Unidad" in out.columns:
        out = out[out["Unidad"].astype(str) == unidad]
    if proceso and proceso != "Todos":
        col = "Proceso_padre" if "Proceso_padre" in out.columns else "Proceso"
        out = out[out[col].astype(str) == proceso]
    if subproceso and subproceso != "Todos":
        col = "Subproceso_final" if "Subproceso_final" in out.columns else "Subproceso"
        if col in out.columns:
            out = out[out[col].astype(str) == subproceso]
    if clasificacion and clasificacion != "Todos" and "Clasificacion" in out.columns:
        out = out[out["Clasificacion"].astype(str) == clasificacion]
    if frecuencia and frecuencia != "Todos":
        freq_col = next((c for c in ["Frecuencia", "Periodicidad"] if c in out.columns), None)
        if freq_col:
            out = out[out[freq_col].astype(str) == frecuencia]
    return out


def latest_per_indicator(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    out = df.copy()
    if "Id" in out.columns:
        return out.drop_duplicates(subset=["Id"], keep="last")
    if "Indicador" in out.columns:
        return out.drop_duplicates(subset=["Indicador"], keep="last")
    return out


def build_procesos_kpis(df: pd.DataFrame) -> dict[str, Any]:
    base = calcular_kpis(df)
    base["n_procesos"] = int(df["Proceso_padre"].nunique()) if "Proceso_padre" in df.columns else 0
    base["n_subprocesos"] = int(df["Subproceso_final"].nunique()) if "Subproceso_final" in df.columns else 0
    base["n_unidades"] = int(df["Unidad"].replace("", pd.NA).dropna().nunique()) if "Unidad" in df.columns else 0
    return base


def build_banner(
    df: pd.DataFrame,
    *,
    anio: int,
    mes: int,
    base_year: int,
    base_month: int | None,
    cumpl_global: float | None,
    cumpl_base: float | None,
) -> dict[str, Any]:
    en_riesgo = 0
    if not df.empty and "Nivel de cumplimiento" in df.columns:
        en_riesgo = int(df["Nivel de cumplimiento"].isin(["Peligro", "Alerta"]).sum())
    variacion = None
    if cumpl_global is not None and cumpl_base is not None:
        variacion = round(cumpl_global - cumpl_base, 1)
    return {
        "titulo": "CMI por Procesos",
        "anio": anio,
        "mes": mes_nombre(mes),
        "cumplimiento_global": cumpl_global,
        "cumplimiento_base_anio": cumpl_base,
        "base_anio": base_year,
        "base_mes": mes_nombre(base_month) if base_month else None,
        "variacion_pp": variacion,
        "total_indicadores": len(df),
        "en_riesgo": en_riesgo,
    }


def _normalize_tipo(tipo: str) -> str:
    t = _norm_text(tipo)
    if "EVALUACION" in t or "EVALUACIÓN" in t:
        return "EVALUACIÓN Y MEJORA"
    for known in TIPOS_PROCESO:
        if _norm_text(known) == t:
            return known
    return str(tipo or "Sin tipo").strip()


def build_tipo_proceso_cards(df: pd.DataFrame, df_prev: pd.DataFrame | None = None) -> list[dict[str, Any]]:
    if df.empty or "Tipo de proceso" not in df.columns:
        return []
    cards = []
    for tipo_raw, group in df.groupby("Tipo de proceso", dropna=False):
        tipo = _normalize_tipo(str(tipo_raw))
        cumpl = _safe_float(group["cumplimiento_pct"].mean()) if "cumplimiento_pct" in group.columns else None
        prev_cumpl = None
        variacion = None
        if df_prev is not None and not df_prev.empty and "Tipo de proceso" in df_prev.columns:
            prev_group = df_prev[df_prev["Tipo de proceso"].astype(str) == str(tipo_raw)]
            if not prev_group.empty:
                prev_cumpl = _safe_float(prev_group["cumplimiento_pct"].mean())
                if cumpl is not None and prev_cumpl is not None:
                    variacion = round(cumpl - prev_cumpl, 1)
        riesgo = 0
        if "Nivel de cumplimiento" in group.columns:
            riesgo = int(group["Nivel de cumplimiento"].isin(["Peligro", "Alerta"]).sum())
        cards.append(
            {
                "tipo": tipo,
                "color": TIPO_PROCESO_COLORS.get(tipo, "#BDBDBD"),
                "color_light": TIPO_PROCESO_COLORS_LIGHT.get(tipo, "#F5F5F5"),
                "icon": TIPO_PROCESO_ICONS.get(tipo, "📋"),
                "cumplimiento": round(cumpl, 1) if cumpl is not None else None,
                "cumplimiento_anterior": round(prev_cumpl, 1) if prev_cumpl is not None else None,
                "variacion_pp": variacion,
                "n_indicadores": len(group),
                "n_riesgo": riesgo,
            }
        )
    cards.sort(key=lambda x: x["cumplimiento"] or 0, reverse=True)
    return cards


def build_proceso_bars(df: pd.DataFrame, df_prev: pd.DataFrame | None = None) -> list[dict[str, Any]]:
    col = "Proceso_padre" if "Proceso_padre" in df.columns else "Proceso"
    if df.empty or col not in df.columns:
        return []
    bars = []
    for proceso, group in df.groupby(col, dropna=True):
        if not proceso or str(proceso).strip() in ("", "nan"):
            continue
        actual = _safe_float(group["cumplimiento_pct"].mean()) if "cumplimiento_pct" in group.columns else None
        anterior = None
        if df_prev is not None and not df_prev.empty and col in df_prev.columns:
            prev_g = df_prev[df_prev[col].astype(str) == str(proceso)]
            if not prev_g.empty:
                anterior = _safe_float(prev_g["cumplimiento_pct"].mean())
        bars.append(
            {
                "proceso": str(proceso),
                "cumplimiento": round(actual, 1) if actual is not None else None,
                "cumplimiento_anterior": round(anterior, 1) if anterior is not None else None,
                "n_indicadores": len(group),
                "color": cumplimiento_semaforo_color(actual),
            }
        )
    bars.sort(key=lambda x: x["cumplimiento"] or 0)
    return bars


def build_procesos_detalle(df: pd.DataFrame) -> list[dict[str, Any]]:
    col = "Proceso_padre" if "Proceso_padre" in df.columns else "Proceso"
    if df.empty or col not in df.columns:
        return []
    items = []
    for proceso, group in df.groupby(col, dropna=True):
        if not proceso or str(proceso).strip() in ("", "nan"):
            continue
        cumpl = _safe_float(group["cumplimiento_pct"].mean())
        conteo: dict[str, int] = {}
        if "Nivel de cumplimiento" in group.columns:
            for cat in ("Peligro", "Alerta", "Cumplimiento", "Sobrecumplimiento"):
                conteo[cat] = int((group["Nivel de cumplimiento"] == cat).sum())
        subprocesos = []
        sub_col = "Subproceso_final" if "Subproceso_final" in group.columns else "Subproceso"
        if sub_col in group.columns:
            for sub, sg in group.groupby(sub_col, dropna=True):
                if not sub or str(sub).strip() in ("", "nan"):
                    continue
                sub_cumpl = _safe_float(sg["cumplimiento_pct"].mean())
                sub_riesgo = int(sg["Nivel de cumplimiento"].isin(["Peligro", "Alerta"]).sum()) if "Nivel de cumplimiento" in sg.columns else 0
                subprocesos.append(
                    {
                        "subproceso": str(sub),
                        "cumplimiento": sub_cumpl,
                        "n_indicadores": len(sg),
                        "en_riesgo": sub_riesgo,
                    }
                )
            subprocesos.sort(key=lambda x: x["cumplimiento"] or 0)
        items.append(
            {
                "proceso": str(proceso),
                "unidad": str(group["Unidad"].iloc[0]) if "Unidad" in group.columns and len(group) else "",
                "tipo_proceso": str(group["Tipo de proceso"].iloc[0]) if "Tipo de proceso" in group.columns and len(group) else "",
                "total_indicadores": len(group),
                "cumplimiento_promedio": cumpl,
                "categorias": conteo,
                "en_riesgo": conteo.get("Peligro", 0) + conteo.get("Alerta", 0),
                "subprocesos": subprocesos,
            }
        )
    items.sort(key=lambda x: x["cumplimiento_promedio"] or 0)
    return items


def build_unidades_detalle(df: pd.DataFrame) -> list[dict[str, Any]]:
    if df.empty or "Unidad" not in df.columns:
        return []
    items = []
    for unidad, group in df.groupby("Unidad", dropna=True):
        u = str(unidad).strip()
        if not u:
            continue
        cumpl = _safe_float(group["cumplimiento_pct"].mean())
        riesgo = int(group["Nivel de cumplimiento"].isin(["Peligro", "Alerta"]).sum()) if "Nivel de cumplimiento" in group.columns else 0
        criticos = int((group["Nivel de cumplimiento"] == "Peligro").sum()) if "Nivel de cumplimiento" in group.columns else 0
        procesos = int(group["Proceso_padre"].nunique()) if "Proceso_padre" in group.columns else 0
        estado = cumplimiento_estado(cumpl)
        items.append(
            {
                "unidad": u,
                "cumplimiento_promedio": cumpl,
                "n_indicadores": len(group),
                "n_procesos": procesos,
                "en_riesgo": riesgo,
                "n_criticos": criticos,
                "color": cumplimiento_semaforo_color(cumpl),
                "estado": estado["label"],
                "estado_icon": estado["icon"],
                "estado_color": estado["color"],
            }
        )
    items.sort(key=lambda x: x["cumplimiento_promedio"] or 0, reverse=True)
    return items


def build_catalog_charts(cmi_df: pd.DataFrame, active_ids: set[str]) -> dict[str, list[dict[str, Any]]]:
    periodicidad: list[dict[str, Any]] = []
    tipo_ind: list[dict[str, Any]] = []
    if cmi_df.empty or "Id" not in cmi_df.columns:
        return {"periodicidad": periodicidad, "tipo_indicador": tipo_ind}

    work = cmi_df.copy()
    work["Id_norm"] = work["Id"].astype(str).str.strip()
    work = work[work["Id_norm"].isin(active_ids)]
    if work.empty:
        return {"periodicidad": periodicidad, "tipo_indicador": tipo_ind}

    per_col = next((c for c in work.columns if "periodicidad" in c.lower() or c == "Frecuencia"), None)
    if per_col:
        vc = work[per_col].fillna("Sin dato").astype(str).value_counts()
        periodicidad = [{"label": str(k), "count": int(v)} for k, v in vc.items()]

    tipo_col = next((c for c in work.columns if "tipo" in c.lower() and "indicador" in c.lower()), None)
    if tipo_col:
        vc = work[tipo_col].fillna("Sin dato").astype(str).value_counts()
        tipo_ind = [{"label": str(k), "count": int(v)} for k, v in vc.items()]

    return {"periodicidad": periodicidad, "tipo_indicador": tipo_ind}


def build_indicadores_summary(df: pd.DataFrame) -> dict[str, int]:
    if df.empty:
        return {"total": 0, "metricas": 0, "sobrecumplimiento": 0, "cumplimiento": 0, "alerta": 0, "peligro": 0}
    unique = latest_per_indicator(df)
    total = len(unique)
    metricas = int(unique["Id"].nunique()) if "Id" in unique.columns else total
    counts = {"sobrecumplimiento": 0, "cumplimiento": 0, "alerta": 0, "peligro": 0}
    if "Nivel de cumplimiento" in unique.columns:
        for nivel, key in [
            ("Sobrecumplimiento", "sobrecumplimiento"),
            ("Cumplimiento", "cumplimiento"),
            ("Alerta", "alerta"),
            ("Peligro", "peligro"),
        ]:
            counts[key] = int((unique["Nivel de cumplimiento"] == nivel).sum())
    return {"total": total, "metricas": metricas, **counts}


def build_variacion_analisis(df: pd.DataFrame, df_prev: pd.DataFrame) -> dict[str, Any]:
    if df.empty:
        return {"mejoraron": [], "empeoraron": [], "top_riesgo_procesos": []}

    mejoraron, empeoraron = compute_trends(df, df_prev)

    col = "Proceso_padre" if "Proceso_padre" in df.columns else "Proceso"
    top_riesgo: list[dict[str, Any]] = []
    if col in df.columns and "Nivel de cumplimiento" in df.columns:
        riesgo_df = df[df["Nivel de cumplimiento"].isin(["Peligro", "Alerta"])]
        if not riesgo_df.empty:
            agg = (
                riesgo_df.groupby(col)
                .agg(n_riesgo=("Nivel de cumplimiento", "count"), cumplimiento=("cumplimiento_pct", "mean"))
                .reset_index()
                .sort_values("n_riesgo", ascending=False)
                .head(8)
            )
            for _, row in agg.iterrows():
                top_riesgo.append(
                    {
                        "proceso": str(row[col]),
                        "n_riesgo": int(row["n_riesgo"]),
                        "cumplimiento": _safe_float(row["cumplimiento"]),
                    }
                )

    return {"mejoraron": mejoraron[:10], "empeoraron": empeoraron[:10], "top_riesgo_procesos": top_riesgo}


def build_comparativa_procesos(df: pd.DataFrame, df_prev: pd.DataFrame | None = None) -> list[dict[str, Any]]:
    col = "Proceso_padre" if "Proceso_padre" in df.columns else "Proceso"
    if df.empty or col not in df.columns:
        return []
    items: list[dict[str, Any]] = []
    for proceso, group in df.groupby(col, dropna=True):
        if not proceso or str(proceso).strip() in ("", "nan"):
            continue
        actual = _safe_float(group["cumplimiento_pct"].mean()) if "cumplimiento_pct" in group.columns else None
        anterior = None
        variacion = None
        if df_prev is not None and not df_prev.empty and col in df_prev.columns:
            prev_g = df_prev[df_prev[col].astype(str) == str(proceso)]
            if not prev_g.empty:
                anterior = _safe_float(prev_g["cumplimiento_pct"].mean())
                if actual is not None and anterior is not None:
                    variacion = round(actual - anterior, 1)
        criticos = 0
        if "Nivel de cumplimiento" in group.columns:
            criticos = int((group["Nivel de cumplimiento"] == "Peligro").sum())
        estado = cumplimiento_estado(actual)
        tipo = str(group["Tipo de proceso"].iloc[0]) if "Tipo de proceso" in group.columns and len(group) else ""
        items.append(
            {
                "proceso": str(proceso),
                "tipo_proceso": _normalize_tipo(tipo),
                "tipo_color": TIPO_PROCESO_COLORS.get(_normalize_tipo(tipo), "#BDBDBD"),
                "cumplimiento": actual,
                "cumplimiento_anterior": anterior,
                "variacion": variacion,
                "n_indicadores": len(group),
                "n_criticos": criticos,
                "color": cumplimiento_semaforo_color(actual),
                "estado": estado["label"],
                "estado_icon": estado["icon"],
                "estado_color": estado["color"],
            }
        )
    items.sort(key=lambda x: x["cumplimiento"] or 0, reverse=True)
    return items


def build_alertas_criticas(df: pd.DataFrame, *, limit: int = 6) -> list[dict[str, Any]]:
    if df.empty or "Nivel de cumplimiento" not in df.columns:
        return []
    peligro = df[df["Nivel de cumplimiento"] == "Peligro"].copy()
    if peligro.empty:
        return []
    if "cumplimiento_pct" in peligro.columns:
        peligro = peligro.sort_values("cumplimiento_pct", ascending=True)
    items: list[dict[str, Any]] = []
    for _, row in peligro.head(limit).iterrows():
        cumpl = _safe_float(row.get("cumplimiento_pct"))
        brecha = round(80 - cumpl, 1) if cumpl is not None else None
        proc = str(row.get("Proceso_padre", row.get("Proceso", "Sin proceso")))
        tipo = _normalize_tipo(str(row.get("Tipo de proceso", "")))
        diag = (
            f"Cumplimiento {cumpl:.1f}% — requiere plan de acción inmediato."
            if cumpl is not None
            else "Sin dato de cumplimiento reportado."
        )
        items.append(
            {
                "id": str(row.get("Id", "")),
                "indicador": str(row.get("Indicador", "Sin nombre")),
                "proceso": proc,
                "tipo_proceso": tipo,
                "tipo_color": TIPO_PROCESO_COLORS.get(tipo, "#BDBDBD"),
                "cumplimiento": cumpl,
                "brecha_pp": brecha,
                "nivel": "Peligro",
                "diagnostico": diag,
            }
        )
    return items


def build_ejecucion_variacion(df: pd.DataFrame, df_prev: pd.DataFrame, *, limit: int = 8) -> dict[str, list[dict[str, Any]]]:
    if df.empty or df_prev.empty or "Id" not in df.columns or "Id" not in df_prev.columns:
        return {"positiva": [], "negativa": []}
    if "Ejecucion" not in df.columns or "Ejecucion" not in df_prev.columns:
        return {"positiva": [], "negativa": []}

    cur = df[["Id", "Indicador", "Ejecucion", "Mes", "Anio"]].drop_duplicates(subset=["Id"], keep="last")
    prev = df_prev[["Id", "Ejecucion"]].drop_duplicates(subset=["Id"], keep="last")
    merged = cur.merge(prev, on="Id", suffixes=("", "_prev"))
    merged["delta"] = pd.to_numeric(merged["Ejecucion"], errors="coerce") - pd.to_numeric(
        merged["Ejecucion_prev"], errors="coerce"
    )
    merged = merged.dropna(subset=["delta"])
    if merged.empty:
        return {"positiva": [], "negativa": []}

    def _row(r: pd.Series) -> dict[str, Any]:
        periodo = f"{r.get('Mes', '')} {r.get('Anio', '')}".strip()
        return {
            "indicador": str(r.get("Indicador", r["Id"])),
            "ejecucion": _safe_float(r.get("Ejecucion")),
            "delta": round(float(r["delta"]), 2),
            "periodo": periodo or "—",
        }

    pos = merged[merged["delta"] > 0].sort_values("delta", ascending=False).head(limit)
    neg = merged[merged["delta"] < 0].sort_values("delta").head(limit)
    return {
        "positiva": [_row(r) for _, r in pos.iterrows()],
        "negativa": [_row(r) for _, r in neg.iterrows()],
    }


def build_vista_global(
    df_global: pd.DataFrame,
    df_base_year: pd.DataFrame,
    cmi_catalog: pd.DataFrame,
    *,
    anio: int,
    mes_corte: int,
    base_year: int,
    base_mes: int | None,
) -> dict[str, Any]:
    latest = latest_per_indicator(df_global)
    active_ids = set(latest["Id"].astype(str).str.strip().tolist()) if "Id" in latest.columns else set()
    cumpl_global = avg_cumplimiento(latest)
    cumpl_base = avg_cumplimiento(df_base_year)
    return {
        "mes_corte": mes_corte,
        "mes_nombre": mes_nombre(mes_corte),
        "kpis": build_procesos_kpis(latest),
        "banner": build_banner(
            latest,
            anio=anio,
            mes=mes_corte,
            base_year=base_year,
            base_month=base_mes,
            cumpl_global=cumpl_global,
            cumpl_base=cumpl_base,
        ),
        "distribucion_nivel": build_distribucion_nivel(latest),
        "tipo_proceso_cards": build_tipo_proceso_cards(latest, df_base_year),
        "proceso_bars": build_proceso_bars(latest, df_base_year),
        "catalog_charts": build_catalog_charts(cmi_catalog, active_ids),
        "procesos_detalle": build_procesos_detalle(latest),
        "unidades_detalle": build_unidades_detalle(latest),
        "comparativa_procesos": build_comparativa_procesos(latest, df_base_year),
        "variacion": build_variacion_analisis(latest, df_base_year if not df_base_year.empty else df_global),
        "alertas_criticas": build_alertas_criticas(latest),
    }


def build_filtros_options(
    tracking: pd.DataFrame,
    map_df: pd.DataFrame,
    cmi_df: pd.DataFrame,
    *,
    anio: int,
) -> dict[str, Any]:
    anios = sorted(
        pd.to_numeric(tracking["Anio"], errors="coerce").dropna().astype(int).unique().tolist()
    ) if "Anio" in tracking.columns else [anio]

    prepared = prepare_tracking(tracking, map_df)
    year_slice = filter_by_anio_mes(prepared, anio=anio, mes=default_mes(tracking, anio))

    unidades = sorted(year_slice["Unidad"].replace("", pd.NA).dropna().astype(str).unique().tolist()) if "Unidad" in year_slice.columns else []
    procesos = sorted(year_slice["Proceso_padre"].dropna().astype(str).unique().tolist()) if "Proceso_padre" in year_slice.columns else []
    subprocesos = sorted(year_slice["Subproceso_final"].dropna().astype(str).unique().tolist()) if "Subproceso_final" in year_slice.columns else []

    clasificaciones: list[str] = []
    if "Clasificacion" in year_slice.columns:
        clasificaciones = sorted(year_slice["Clasificacion"].dropna().astype(str).unique().tolist())

    frecuencias: list[str] = []
    freq_col = next((c for c in ["Frecuencia", "Periodicidad"] if c in year_slice.columns), None)
    if freq_col:
        frecuencias = sorted(year_slice[freq_col].dropna().astype(str).unique().tolist())

    meses_disp: list[int] = []
    if "Anio" in tracking.columns and "Mes" in tracking.columns:
        sub = tracking[pd.to_numeric(tracking["Anio"], errors="coerce") == int(anio)]
        nums = sub["Mes"].apply(mes_to_num).dropna().astype(int).unique().tolist()
        meses_disp = sorted(nums)

    return {
        "anios": anios,
        "meses": meses_disp or list(range(1, 13)),
        "mes_default": default_mes(tracking, anio),
        "unidades": unidades,
        "procesos": procesos,
        "subprocesos": subprocesos,
        "clasificaciones": clasificaciones,
        "frecuencias": frecuencias,
    }


def build_indicadores_procesos_listado(df: pd.DataFrame) -> list[dict[str, Any]]:
    cols_extra = [c for c in ["Proceso_padre", "Subproceso_final", "Unidad", "Tipo de proceso"] if c in df.columns]
    base_cols = [
        c
        for c in [
            "Id",
            "Indicador",
            "Proceso",
            "Subproceso",
            "Meta",
            "Ejecucion",
            "cumplimiento_pct",
            "Nivel de cumplimiento",
            "Clasificacion",
            "Frecuencia",
            "Periodicidad",
            "Anio",
            "Mes",
        ]
        if c in df.columns
    ]
    return df_to_records(df, list(dict.fromkeys(base_cols + cols_extra)))


def avg_cumplimiento(df: pd.DataFrame) -> float | None:
    if df.empty or "cumplimiento_pct" not in df.columns:
        return None
    val = _safe_float(df["cumplimiento_pct"].mean())
    return round(val, 1) if val is not None else None


def default_anio_procesos(anios: list[int]) -> int:
    if 2025 in anios:
        return 2025
    if anios:
        return anios[-1]
    return date.today().year


def process_variation_for_rpp(
    df_current: pd.DataFrame,
    df_prev: pd.DataFrame,
    display_col: str = "Proceso_padre",
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    if df_current.empty or df_prev.empty or display_col not in df_current.columns:
        return [], []
    curr = (
        df_current.groupby(display_col, dropna=False)
        .agg(actual=("cumplimiento_pct", "mean"))
        .reset_index()
    )
    prev = (
        df_prev.groupby(display_col, dropna=False)
        .agg(prev=("cumplimiento_pct", "mean"))
        .reset_index()
    )
    merged = curr.merge(prev, on=display_col, how="left")
    merged["change"] = merged["actual"] - pd.to_numeric(merged["prev"], errors="coerce")
    merged = merged.dropna(subset=["change"])
    if merged.empty:
        return [], []
    best = [
        {"name": str(r[display_col]), "change": round(float(r["change"]), 1)}
        for _, r in merged.sort_values("change", ascending=False).head(5).iterrows()
    ]
    worst = [
        {"name": str(r[display_col]), "change": round(float(r["change"]), 1)}
        for _, r in merged.sort_values("change", ascending=True).head(5).iterrows()
    ]
    return best, worst


def build_propuesta_accion(df: pd.DataFrame, proceso: str = "Todos") -> dict[str, Any]:
    if df.empty:
        return {
            "proceso": proceso,
            "plan_mejoramiento": "Sin datos para priorizar acciones.",
            "pdi": "Definir metas e hitos por indicador.",
            "sga": "Validar integración de evidencias y trazabilidad.",
            "retos": "Completar reporte de indicadores faltantes.",
            "top_criticos": [],
        }
    work = df.copy()
    if "cumplimiento_pct" not in work.columns:
        return build_propuesta_accion(pd.DataFrame(), proceso)
    riesgos = work[work["cumplimiento_pct"].notna() & (work["cumplimiento_pct"] < 80)]
    top = riesgos.nsmallest(3, "cumplimiento_pct") if not riesgos.empty else pd.DataFrame()
    top_criticos = []
    for _, r in top.iterrows():
        top_criticos.append(
            {
                "indicador": str(r.get("Indicador", "Sin nombre")),
                "proceso": str(r.get("Proceso_padre", r.get("Proceso", "Sin proceso"))),
                "cumplimiento": _safe_float(r.get("cumplimiento_pct")),
            }
        )
    nombres = ", ".join(t["indicador"] for t in top_criticos) or "Ninguno crítico en el corte"
    return {
        "proceso": proceso,
        "plan_mejoramiento": f"Priorizar cierre de brechas en: {nombres}.",
        "pdi": "Alinear metas por resultados históricos y capacidad operativa.",
        "sga": "Fortalecer controles de calidad de dato y consistencia de soportes.",
        "retos": "Sostener cumplimiento mensual y reducir dispersión entre subprocesos.",
        "top_criticos": top_criticos,
    }


def generate_ficha_narrativa_heuristica(
    *,
    nombre: str,
    meta: Any,
    ejecucion: Any,
    nivel: str,
    cumplimiento: float | None,
    proceso: str | None = None,
) -> dict[str, str]:
    cump = cumplimiento or 0.0
    if cump >= 100:
        estado = "El indicador supera la meta y presenta desempeño favorable en el corte actual."
        riesgo = "Riesgo de sobreconfianza: sostener el resultado sin plan de continuidad puede generar retroceso."
        accion = "Estandarizar las prácticas que explican el resultado y documentar un plan de sostenibilidad."
    elif cump >= 95:
        estado = "El indicador está cerca de la meta y requiere ajuste fino para consolidar cumplimiento."
        riesgo = "Riesgo de cierre en alerta por variaciones menores de ejecución o retrasos operativos."
        accion = "Definir acciones de corto plazo con responsables y seguimiento semanal hasta el próximo corte."
    else:
        estado = "El indicador presenta brecha frente a la meta y requiere intervención prioritaria."
        riesgo = "Riesgo de incumplimiento del objetivo asociado y efectos en el balance del proceso."
        accion = "Implementar plan de recuperación con metas parciales y control quincenal de avance."

    proc_ctx = f" en el proceso <strong>{proceso}</strong>" if proceso else ""
    texto = (
        f"<strong>Diagnóstico:</strong> {estado}{proc_ctx}<br/>"
        f"<strong>Riesgo principal:</strong> {riesgo}<br/>"
        f"<strong>Recomendación táctica:</strong> {accion}<br/>"
        f"<strong>Meta / Ejecución:</strong> {meta} / {ejecucion} — Nivel: {nivel}"
    )
    return {
        "diagnostico": estado,
        "riesgo": riesgo,
        "recomendacion": accion,
        "texto_html": texto,
        "fuente": "heuristica",
    }


def generate_proceso_narrativa_heuristica(
    proceso: str,
    df: pd.DataFrame,
    *,
    cumplimiento_promedio: float | None = None,
) -> dict[str, Any]:
    total = len(df)
    riesgo = 0
    if "Nivel de cumplimiento" in df.columns:
        riesgo = int(df["Nivel de cumplimiento"].isin(["Peligro", "Alerta"]).sum())
    cump = cumplimiento_promedio if cumplimiento_promedio is not None else _safe_float(df["cumplimiento_pct"].mean())
    cump = cump or 0.0
    riesgo_ratio = (riesgo / total) if total else 0.0

    if cump >= 100:
        estado = f"El proceso <strong>{proceso}</strong> supera la meta institucional en el corte actual."
        color = "#16A34A"
    elif cump >= 95:
        estado = f"El proceso <strong>{proceso}</strong> mantiene desempeño estable con brechas acotadas."
        color = "#2563EB"
    else:
        estado = f"El proceso <strong>{proceso}</strong> presenta desviación relevante y requiere priorización."
        color = "#DC2626"

    dir_1 = "Enfocar seguimiento semanal en indicadores en riesgo del proceso."
    dir_2 = (
        "Activar comité táctico con responsables por subproceso."
        if riesgo_ratio > 0.25
        else "Replicar buenas prácticas de subprocesos con mejor desempeño."
    )
    texto = (
        f"{estado} Hay <strong>{riesgo}</strong> indicadores en alerta/peligro sobre <strong>{total}</strong>.<br/>"
        f"<strong>Directriz 1:</strong> {dir_1}<br/>"
        f"<strong>Directriz 2:</strong> {dir_2}"
    )
    return {
        "titulo": "Análisis heurístico del proceso",
        "estado_color": color,
        "foco_urgente": dir_1,
        "directrices": [dir_1, dir_2],
        "texto_html": texto,
        "fuente": "heuristica",
    }


def build_historico_catalog(df: pd.DataFrame, tracking: pd.DataFrame, *, limit: int = 50) -> list[dict[str, Any]]:
    if df.empty or tracking.empty or "Id" not in df.columns:
        return []
    ids = set(df["Id"].astype(str).str.strip().tolist())
    hist = tracking[tracking["Id"].astype(str).str.strip().isin(ids)].copy()
    if hist.empty:
        return []

    catalog: list[dict[str, Any]] = []
    name_col = "Indicador" if "Indicador" in hist.columns else "Id"
    for ind_id, group in hist.groupby("Id"):
        ind_id = str(ind_id)
        name = str(group[name_col].iloc[0]) if name_col in group.columns else ind_id
        puntos = []
        sort_cols = [c for c in ["Anio", "Mes"] if c in group.columns]
        sorted_g = group.sort_values(sort_cols) if sort_cols else group
        for _, row in sorted_g.iterrows():
            mes_num = mes_to_num(row.get("Mes"))
            anio = row.get("Anio")
            periodo = f"{anio}-{int(mes_num):02d}" if mes_num and pd.notna(anio) else str(anio or "")
            pct = row.get("cumplimiento_pct") or row.get("Cumplimiento_norm")
            if pct is not None:
                try:
                    val = float(pct)
                    if val <= 2:
                        val *= 100
                except (TypeError, ValueError):
                    val = None
            else:
                val = None
            if val is not None:
                puntos.append({"periodo": periodo, "cumplimiento": round(val, 1)})
        if puntos:
            catalog.append({"id": ind_id, "indicador": name, "puntos": puntos})
        if len(catalog) >= limit:
            break
    return catalog


_EXPORT_COLS = {
    "Id": "Código",
    "Indicador": "Indicador",
    "Proceso_padre": "Proceso",
    "Subproceso_final": "Subproceso",
    "Unidad": "Unidad",
    "Tipo de proceso": "Tipo de proceso",
    "Meta": "Meta",
    "Ejecucion": "Ejecución",
    "cumplimiento_pct": "Cumplimiento (%)",
    "Nivel de cumplimiento": "Estado",
    "Clasificacion": "Clasificación",
    "Frecuencia": "Frecuencia",
    "Periodicidad": "Periodicidad",
    "Anio": "Año",
    "Mes": "Mes",
}


def build_export_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    present = {k: v for k, v in _EXPORT_COLS.items() if k in df.columns}
    if not present:
        return df.copy()
    out = df[list(present.keys())].copy().rename(columns=present)
    if "Cumplimiento (%)" in out.columns:
        out["Cumplimiento (%)"] = pd.to_numeric(out["Cumplimiento (%)"], errors="coerce").round(1)
    return out


def build_export_excel_bytes(df: pd.DataFrame) -> bytes:
    import io

    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    df_exp = build_export_dataframe(df)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_exp.to_excel(writer, sheet_name="Indicadores CMI Procesos", index=False)
        ws = writer.sheets["Indicadores CMI Procesos"]
        thin = Side(style="thin", color="D1D5DB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_fill = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
        hdr_font = Font(color="FFFFFF", bold=True, size=11)
        for col_idx in range(1, len(df_exp.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        col_widths = {"Indicador": 48, "Proceso": 28, "Subproceso": 28}
        for col_idx, col_name in enumerate(df_exp.columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 18)
        ws.freeze_panes = "A2"
    return buf.getvalue()


def build_analisis_avanzado(
    df: pd.DataFrame,
    df_prev: pd.DataFrame,
    df_base_year: pd.DataFrame,
    tracking: pd.DataFrame,
    *,
    proceso: str = "Todos",
    base_anio: int,
) -> dict[str, Any]:
    col = "Proceso_padre" if "Proceso_padre" in df.columns else "Proceso"
    best_proc, worst_proc = process_variation_for_rpp(
        df if not df.empty else df_base_year,
        df_base_year if not df_base_year.empty else df_prev,
        display_col=col,
    )
    propuesta = build_propuesta_accion(df, proceso=proceso)
    narrativa = generate_proceso_narrativa_heuristica(
        proceso if proceso != "Todos" else "Todos los procesos",
        df,
    )
    variacion = build_variacion_analisis(df, df_prev if not df_prev.empty else df_base_year)
    historico = build_historico_catalog(df, tracking)

    mejora_txt = (
        f"{best_proc[0]['name']} registra la mayor variación positiva: {best_proc[0]['change']:+.1f}% vs {base_anio}."
        if best_proc
        else "Sin datos comparativos de mejora."
    )
    riesgo_txt = (
        f"{worst_proc[0]['name']} presenta la mayor caída: {worst_proc[0]['change']:+.1f}% vs {base_anio}."
        if worst_proc
        else "Sin datos comparativos de riesgo."
    )

    return {
        "propuesta_accion": propuesta,
        "variacion_procesos": {"mejoraron": best_proc, "empeoraron": worst_proc},
        "insights": {"mejora_proceso": mejora_txt, "riesgo_proceso": riesgo_txt},
        "narrativa_proceso": narrativa,
        "variacion_indicadores": variacion,
        "historico_indicadores": historico,
    }

