"""
scripts/_purgar_filas_variables_bug.py
Repara filas corruptas por el bug de fallback de variables (símbolo de
ejecución configurado ausente en el período -> se adivinaba con otra
variable -> AGENT5 capeaba a 1.3). Ver fix en scripts/etl/extraccion.py.

Elimina de Consolidado Historico/Semestral/Cierres las filas cuyo (Id, Fecha)
coincide con la lista de casos detectados, SOLO si el valor actual sigue
siendo exactamente Ejecucion == 1.3 (huella del bug) — así no se toca nada
que ya tenga un valor legítimo (p.ej. algunas filas en Cierres).

Después de correr este script, ejecutar de nuevo
scripts/actualizar_consolidado.py para que esas filas se regeneren desde la
fuente con la lógica corregida (quedarán en blanco si de verdad no hay
dato, o con el valor correcto si existe un registro fuente válido).

Uso: python scripts/_purgar_filas_variables_bug.py
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

import openpyxl
import pandas as pd

_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(_ROOT / "scripts"))

from etl.config import OUTPUT_FILE  # noqa: E402

# (Id, Fecha) detectados con variables_campo_map['ejec'] configurado pero
# ausente en el período — ver validación 2026-07-14.
CASOS = [
    ("85", "2026-12-31"),
    ("245", "2026-06-30"), ("245", "2026-12-31"),
    ("252", "2026-06-30"), ("252", "2026-12-31"),
    ("256", "2026-06-30"), ("256", "2026-12-31"),
    ("341", "2026-12-31"),
    ("371", "2026-12-31"),
    ("379", "2026-12-31"),
    ("381", "2026-12-31"),
    ("382", "2026-12-31"),
    ("383", "2026-12-31"),
    ("384", "2026-12-31"),
    ("398", "2026-12-31"),
    ("399", "2026-06-30"), ("399", "2026-12-31"),
    ("401", "2026-06-30"), ("401", "2026-12-31"),
    ("392", "2026-12-31"),
    ("453", "2026-03-31"), ("453", "2026-06-30"), ("453", "2026-09-30"), ("453", "2026-12-31"),
    ("452", "2026-06-30"), ("452", "2026-12-31"),
    ("467", "2026-12-31"),
    ("528", "2025-12-31"),
]
CASOS_SET = {(id_s, fecha) for id_s, fecha in CASOS}


def _id_str(v) -> str:
    if pd.isna(v):
        return ""
    s = str(v).strip()
    return s[:-2] if s.endswith(".0") else s


def purgar_hoja(ws, nombre: str) -> int:
    header = {cell.value: cell.column for cell in ws[1]}
    col_id = header.get("Id")
    col_fecha = header.get("Fecha")
    col_ejec = header.get("Ejecucion")
    if not (col_id and col_fecha and col_ejec):
        print(f"  [{nombre}] Columnas Id/Fecha/Ejecucion no encontradas — omitido")
        return 0

    filas_a_borrar = []
    for row_idx in range(2, ws.max_row + 1):
        id_val = ws.cell(row=row_idx, column=col_id).value
        fecha_val = ws.cell(row=row_idx, column=col_fecha).value
        ejec_val = ws.cell(row=row_idx, column=col_ejec).value
        if id_val is None or fecha_val is None:
            continue
        id_s = _id_str(id_val)
        try:
            fecha_s = pd.to_datetime(fecha_val).strftime("%Y-%m-%d")
        except Exception:
            continue
        if (id_s, fecha_s) in CASOS_SET:
            try:
                ejec_num = float(ejec_val) if ejec_val is not None else None
            except (TypeError, ValueError):
                ejec_num = None
            if ejec_num is not None and abs(ejec_num - 1.3) < 1e-9:
                filas_a_borrar.append(row_idx)

    for row_idx in sorted(filas_a_borrar, reverse=True):
        ws.delete_rows(row_idx, 1)

    print(f"  [{nombre}] {len(filas_a_borrar)} filas purgadas")
    return len(filas_a_borrar)


def main() -> None:
    if not OUTPUT_FILE.exists():
        print(f"[ERROR] No existe {OUTPUT_FILE}")
        return

    bak = OUTPUT_FILE.with_name(OUTPUT_FILE.stem + ".bak_pre_purga_variables" + OUTPUT_FILE.suffix)
    if not bak.exists():
        shutil.copy2(OUTPUT_FILE, bak)
        print(f"backup -> {bak.name}")

    wb = openpyxl.load_workbook(OUTPUT_FILE)
    total = 0
    for nombre in ["Consolidado Historico", "Consolidado Semestral", "Consolidado Cierres"]:
        if nombre in wb.sheetnames:
            total += purgar_hoja(wb[nombre], nombre)

    wb.save(OUTPUT_FILE)
    print(f"\n[OK] {total} filas purgadas en total. Correr ahora: python scripts/actualizar_consolidado.py")


if __name__ == "__main__":
    main()
