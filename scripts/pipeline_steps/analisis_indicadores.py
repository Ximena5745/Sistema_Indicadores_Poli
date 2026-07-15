#!/usr/bin/env python3
"""
=============================================================
ANÁLISIS DE INCONSISTENCIAS — INDICADORES
=============================================================
Evalúa cada indicador buscando grandes diferencias, variaciones
o inconsistencias históricas en ejecución, meta y cumplimiento.

SALIDA:
  data/output/Analisis_Indicadores_[timestamp].xlsx
  → Hoja "Resumen": un indicador por fila con métricas y severidad
  → Hoja "Periodos_Anomalos": cada período problemático con contexto

EJECUTAR:
  python scripts/pipeline_steps/analisis_indicadores.py
  python scripts/pipeline_steps/analisis_indicadores.py --fuente df_api
  python scripts/pipeline_steps/analisis_indicadores.py --solo-anomalias
=============================================================
"""

import argparse
import sys
import time
from datetime import datetime
from pathlib import Path

_STEPS_DIR  = Path(__file__).parent
_SCRIPTS_DIR = _STEPS_DIR.parent
_ROOT        = _SCRIPTS_DIR.parent
for _p in (str(_STEPS_DIR), str(_SCRIPTS_DIR), str(_ROOT)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from etl.config import OUTPUT_DIR


# ── Umbrales de anomalía ──────────────────────────────────────────────
UMBRAL_CUMPL_ALTO   = 300    # cumplimiento > 300% → anómalo
UMBRAL_CUMPL_BAJO   = 0      # cumplimiento = 0 exacto → sin ejecución
UMBRAL_SALTO_PCT    = 150    # salto inter-período > 150% del valor anterior
UMBRAL_CV           = 80     # CV (std/mean * 100) > 80% → alta variabilidad
UMBRAL_META_CV      = 50     # CV de la meta > 50% → meta inestable


# ── Colores institucionales ───────────────────────────────────────────
# ── Sin fondo en ninguna celda — diferenciación solo por fuente y borde ──
NO_FILL = PatternFill(fill_type=None)

FONT_HEADER  = Font(bold=True, color="1A3A6B",  name="Inter", size=10)   # azul oscuro
FONT_BOLD    = Font(bold=True,                   name="Inter", size=9)
FONT_NORMAL  = Font(                             name="Inter", size=9)
FONT_GRIS    = Font(color="999999",              name="Inter", size=9)    # Normal sin anomalías
FONT_ALTA    = Font(bold=True, color="CC0000",   name="Inter", size=9)    # Alta — rojo
FONT_MEDIA   = Font(bold=True, color="995500",   name="Inter", size=9)    # Media — naranja oscuro
FONT_BAJA    = Font(bold=True, color="665500",   name="Inter", size=9)    # Baja — marrón

BORDER_THIN  = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
BORDER_HEADER = Border(
    left=Side(style="thin", color="1A3A6B"),
    right=Side(style="thin", color="1A3A6B"),
    top=Side(style="medium", color="1A3A6B"),
    bottom=Side(style="medium", color="1A3A6B"),
)

ALIN_CENTRO = Alignment(horizontal="center", vertical="center", wrap_text=False)
ALIN_IZQ    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIN_DER    = Alignment(horizontal="right",  vertical="center")


def log(nivel, msg):
    ts = datetime.now().strftime("%H:%M:%S")
    iconos = {"INFO": ">>", "OK": "OK", "ERROR": "!!", "WARN": "??", "DATA": "--"}
    print(f"[{ts}] {iconos.get(nivel, '  ')} {nivel:<5} | {msg}")
    sys.stdout.flush()


# ── Carga de datos ────────────────────────────────────────────────────

def cargar_datos(fuente: str) -> pd.DataFrame:
    """Carga datos desde df_api.csv (estado pipeline) o desde el Excel de salida."""
    if fuente == "df_api":
        csv = _ROOT / ".pipeline_state" / "df_api.csv"
        if not csv.exists():
            raise FileNotFoundError(f"df_api.csv no encontrado. Ejecuta Paso 01 primero.\n  {csv}")
        log("INFO", f"Cargando df_api.csv — {csv.name}")
        df = pd.read_csv(csv, low_memory=False)
        df = df.rename(columns={
            "Id": "ID", "Indicador": "Nombre", "resultado": "Ejecucion",
            "meta": "Meta", "cumplimiento": "Cumplimiento", "fecha": "Fecha",
        })
    else:
        from etl.config import OUTPUT_FILE
        if not OUTPUT_FILE.exists():
            raise FileNotFoundError(f"Archivo de salida no encontrado: {OUTPUT_FILE}")
        log("INFO", f"Cargando hoja Histórico de {OUTPUT_FILE.name}")
        df = pd.read_excel(OUTPUT_FILE, sheet_name="Consolidado Historico")
        df = df.rename(columns={
            "Id": "ID", "Indicador": "Nombre", "Ejecucion": "Ejecucion",
            "Meta": "Meta", "Cumplimiento": "Cumplimiento", "Fecha": "Fecha",
        })

    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    df["ID"]    = pd.to_numeric(df["ID"], errors="coerce")
    df = df.dropna(subset=["ID", "Fecha"])
    df["ID"]    = df["ID"].astype(int)
    return df


# ── Análisis por indicador ────────────────────────────────────────────

def analizar_indicador(gdf: pd.DataFrame) -> dict:
    """
    Recibe el DataFrame de un solo indicador (ya ordenado por Fecha).
    Devuelve métricas y lista de anomalías detectadas.
    """
    meta_rec  = gdf[gdf["Meta"].notna()]["Meta"].iloc[-1]      if gdf["Meta"].notna().any() else None
    ejec_rec  = gdf[gdf["Ejecucion"].notna()]["Ejecucion"].iloc[-1] if gdf["Ejecucion"].notna().any() else None
    cumpl_rec = gdf[gdf["Cumplimiento"].notna()]["Cumplimiento"].iloc[-1] if gdf["Cumplimiento"].notna().any() else None
    fecha_rec = gdf["Fecha"].max()

    serie_ejec  = gdf["Ejecucion"].dropna()
    serie_meta  = gdf["Meta"].dropna()
    serie_cumpl = gdf["Cumplimiento"].dropna()

    anomalias = []
    severidad_score = 0

    # ── Coeficiente de variación de la ejecución ──────────────────
    cv_ejec = None
    if len(serie_ejec) >= 3 and serie_ejec.mean() != 0:
        cv_ejec = round(serie_ejec.std() / abs(serie_ejec.mean()) * 100, 1)
        if cv_ejec > UMBRAL_CV:
            anomalias.append(f"Alta variabilidad ejecución (CV={cv_ejec:.0f}%)")
            severidad_score += 2 if cv_ejec > 150 else 1

    # ── Salto máximo entre períodos consecutivos ──────────────────
    salto_max_pct = None
    salto_max_fecha = None
    if len(serie_ejec) >= 2:
        vals = gdf[gdf["Ejecucion"].notna()].sort_values("Fecha")
        ejec_vals = vals["Ejecucion"].values
        fechas    = vals["Fecha"].values
        saltos = []
        for i in range(1, len(ejec_vals)):
            prev = ejec_vals[i - 1]
            curr = ejec_vals[i]
            if prev != 0 and not np.isnan(prev):
                pct = abs(curr - prev) / abs(prev) * 100
                saltos.append((pct, fechas[i]))
        if saltos:
            salto_max_pct, salto_max_fecha = max(saltos, key=lambda x: x[0])
            salto_max_pct = round(salto_max_pct, 1)
            if salto_max_pct > UMBRAL_SALTO_PCT:
                fecha_str = pd.Timestamp(salto_max_fecha).strftime("%Y-%m")
                anomalias.append(f"Salto brusco en {fecha_str} ({salto_max_pct:.0f}%)")
                severidad_score += 2 if salto_max_pct > 300 else 1

    # ── Cumplimiento extremo ──────────────────────────────────────
    cumpl_max = round(serie_cumpl.max(), 1) if len(serie_cumpl) else None
    cumpl_min = round(serie_cumpl.min(), 1) if len(serie_cumpl) else None

    if cumpl_max is not None and cumpl_max > UMBRAL_CUMPL_ALTO:
        n = (serie_cumpl > UMBRAL_CUMPL_ALTO).sum()
        anomalias.append(f"Cumplimiento excesivo en {n} período(s) (máx {cumpl_max:.0f}%)")
        severidad_score += 3 if cumpl_max > 1000 else 1

    n_cero = (serie_cumpl == UMBRAL_CUMPL_BAJO).sum() if len(serie_cumpl) else 0
    if n_cero >= 2:
        anomalias.append(f"Ejecución cero en {n_cero} período(s)")
        severidad_score += 1

    # ── Meta inestable ───────────────────────────────────────────
    cv_meta = None
    if len(serie_meta) >= 3 and serie_meta.mean() != 0:
        cv_meta = round(serie_meta.std() / abs(serie_meta.mean()) * 100, 1)
        if cv_meta > UMBRAL_META_CV:
            anomalias.append(f"Meta inestable (CV={cv_meta:.0f}%)")
            severidad_score += 1

    # ── Meta en cero o negativa ──────────────────────────────────
    n_meta_cero = (serie_meta <= 0).sum() if len(serie_meta) else 0
    if n_meta_cero > 0:
        anomalias.append(f"Meta <= 0 en {n_meta_cero} periodos")
        severidad_score += 1

    # ── Severidad final ──────────────────────────────────────────
    if severidad_score == 0:
        severidad = "Normal"
    elif severidad_score <= 1:
        severidad = "Baja"
    elif severidad_score <= 3:
        severidad = "Media"
    else:
        severidad = "Alta"

    return {
        "Meta_reciente":    meta_rec,
        "Ejecucion_reciente": ejec_rec,
        "Cumplimiento_reciente": cumpl_rec,
        "Fecha_reciente":   fecha_rec,
        "N_periodos":       len(gdf),
        "N_con_ejecucion":  int(gdf["Ejecucion"].notna().sum()),
        "CV_ejecucion":     cv_ejec,
        "Salto_max_pct":    salto_max_pct,
        "Cumplimiento_max": cumpl_max,
        "Cumplimiento_min": cumpl_min,
        "CV_meta":          cv_meta,
        "N_periodos_cero":  n_cero,
        "N_anomalias":      len(anomalias),
        "Anomalias":        " | ".join(anomalias) if anomalias else "—",
        "Severidad":        severidad,
        "_severidad_score": severidad_score,
    }


def detectar_periodos_anomalos(df: pd.DataFrame) -> pd.DataFrame:
    """Genera una fila por cada período problemático."""
    filas = []
    for (iid, nombre, perio), gdf in df.groupby(["ID", "Nombre", "Periodicidad"], sort=False):
        gdf = gdf.sort_values("Fecha")
        serie_cumpl = gdf[gdf["Cumplimiento"].notna()].copy()

        for _, row in serie_cumpl.iterrows():
            tipo = None
            if row["Cumplimiento"] > UMBRAL_CUMPL_ALTO:
                tipo = f"Cumplimiento excesivo ({row['Cumplimiento']:.0f}%)"
            elif row["Cumplimiento"] == 0 and pd.notna(row["Meta"]) and row["Meta"] > 0:
                tipo = "Ejecución cero con meta válida"

            if tipo:
                filas.append({
                    "ID": iid,
                    "Nombre": nombre,
                    "Periodicidad": perio,
                    "Fecha": row["Fecha"].strftime("%Y-%m-%d") if pd.notna(row["Fecha"]) else "",
                    "Meta": row.get("Meta"),
                    "Ejecucion": row.get("Ejecucion"),
                    "Cumplimiento": row.get("Cumplimiento"),
                    "Tipo_anomalia": tipo,
                })

        # Saltos entre períodos
        vals = gdf[gdf["Ejecucion"].notna()].copy()
        if len(vals) >= 2:
            ejec_prev = vals["Ejecucion"].shift(1)
            salto_pct = ((vals["Ejecucion"] - ejec_prev) / ejec_prev.abs() * 100).abs()
            for idx, (_, row) in enumerate(vals.iterrows()):
                if idx == 0:
                    continue
                sp = salto_pct.iloc[idx]
                if sp > UMBRAL_SALTO_PCT:
                    filas.append({
                        "ID": iid,
                        "Nombre": nombre,
                        "Periodicidad": perio,
                        "Fecha": row["Fecha"].strftime("%Y-%m-%d") if pd.notna(row["Fecha"]) else "",
                        "Meta": row.get("Meta"),
                        "Ejecucion": row.get("Ejecucion"),
                        "Cumplimiento": row.get("Cumplimiento"),
                        "Tipo_anomalia": f"Salto brusco ({sp:.0f}% vs período anterior)",
                    })

    return pd.DataFrame(filas) if filas else pd.DataFrame()


# ── Construcción del Excel ────────────────────────────────────────────

def _set_header(ws, headers: list, widths: list):
    ws.append(headers)
    for j, (cell, w) in enumerate(zip(ws[1], widths), start=1):
        cell.fill      = NO_FILL
        cell.font      = FONT_HEADER
        cell.alignment = ALIN_CENTRO
        cell.border    = BORDER_HEADER
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"


def _fmt_num(val, decimales=1):
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return ""
    return round(float(val), decimales)


def construir_excel(df_resumen: pd.DataFrame, df_detalle: pd.DataFrame,
                    ts: str) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    # ── HOJA 1 — Resumen ──────────────────────────────────────────
    ws1 = wb.create_sheet("Resumen")

    headers_res = [
        "ID", "Nombre del Indicador", "Periodicidad",
        "Meta\n(último período)", "Ejecución\n(último período)", "Cumplimiento\n(%)",
        "Períodos\ntotales", "Períodos con\nejecución",
        "CV Ejecución\n(%)", "Salto máx.\ninter-período (%)",
        "Cumpl. máx.\n(%)", "Cumpl. mín.\n(%)",
        "N° Anomalías", "Detalle Anomalías", "Severidad",
    ]
    widths_res = [6, 38, 14, 14, 14, 14, 10, 12, 12, 16, 12, 12, 10, 50, 12]

    _set_header(ws1, headers_res, widths_res)
    ws1.row_dimensions[1].height = 32

    # Orden: Alta → Media → Baja → Normal
    orden_sev = {"Alta": 0, "Media": 1, "Baja": 2, "Normal": 3}
    df_resumen = df_resumen.copy()
    df_resumen["_ord"] = df_resumen["Severidad"].map(orden_sev)
    df_resumen = df_resumen.sort_values(["_ord", "ID"]).drop(columns="_ord")

    SEV_FONT = {
        "Alta":   FONT_ALTA,
        "Media":  FONT_MEDIA,
        "Baja":   FONT_BAJA,
        "Normal": FONT_GRIS,
    }

    for i, row in enumerate(df_resumen.itertuples(index=False), start=2):
        sev = row.Severidad
        font_fila = SEV_FONT.get(sev, FONT_NORMAL)

        cumpl_fmt = _fmt_num(row.Cumplimiento_reciente, 1)

        vals = [
            row.ID, row.Nombre, row.Periodicidad,
            _fmt_num(row.Meta_reciente, 2), _fmt_num(row.Ejecucion_reciente, 2), cumpl_fmt,
            row.N_periodos, row.N_con_ejecucion,
            _fmt_num(row.CV_ejecucion, 1), _fmt_num(row.Salto_max_pct, 1),
            _fmt_num(row.Cumplimiento_max, 1), _fmt_num(row.Cumplimiento_min, 1),
            row.N_anomalias, row.Anomalias, sev,
        ]
        ws1.append(vals)

        for j, cell in enumerate(ws1[i], start=1):
            cell.fill      = NO_FILL
            cell.font      = font_fila
            cell.border    = BORDER_THIN
            cell.alignment = ALIN_IZQ if j in (2, 14) else ALIN_CENTRO

    ws1.auto_filter.ref = ws1.dimensions

    # ── HOJA 2 — Períodos Anómalos ────────────────────────────────
    ws2 = wb.create_sheet("Periodos_Anomalos")

    headers_det = [
        "ID", "Nombre del Indicador", "Periodicidad",
        "Fecha", "Meta", "Ejecución", "Cumplimiento (%)", "Tipo de Anomalía",
    ]
    widths_det = [6, 38, 14, 14, 14, 14, 16, 48]
    _set_header(ws2, headers_det, widths_det)

    if not df_detalle.empty:
        df_detalle = df_detalle.sort_values(["ID", "Fecha"])
        for i, row in enumerate(df_detalle.itertuples(index=False), start=2):
            es_cumpl = "Cumplimiento excesivo" in str(row.Tipo_anomalia)
            es_salto = "Salto" in str(row.Tipo_anomalia)
            font_d = FONT_ALTA if es_cumpl else (FONT_MEDIA if es_salto else FONT_BAJA)

            vals = [
                row.ID, row.Nombre, row.Periodicidad, row.Fecha,
                _fmt_num(row.Meta, 2), _fmt_num(row.Ejecucion, 2),
                _fmt_num(row.Cumplimiento, 1), row.Tipo_anomalia,
            ]
            ws2.append(vals)
            for j, cell in enumerate(ws2[i], start=1):
                cell.fill      = NO_FILL
                cell.font      = font_d
                cell.border    = BORDER_THIN
                cell.alignment = ALIN_IZQ if j in (2, 8) else ALIN_CENTRO

    ws2.auto_filter.ref = ws2.dimensions

    # ── HOJA 3 — Leyenda ─────────────────────────────────────────
    ws3 = wb.create_sheet("Leyenda")
    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 55

    leyenda = [
        ("SEVERIDAD", "CRITERIO"),
        ("ALTA",   f"CV ejecucion > {UMBRAL_CV}% y/o salto > {UMBRAL_SALTO_PCT}%, "
                   f"o cumplimiento > {UMBRAL_CUMPL_ALTO}% en multiples periodos"),
        ("MEDIA",  "Salto brusco entre periodos O cumplimiento excesivo aislado"),
        ("BAJA",   "Meta inestable, periodos con ejecucion cero u otras alertas menores"),
        ("NORMAL", "Sin anomalias detectadas con los umbrales actuales"),
        ("", ""),
        ("METRICA", "DEFINICION"),
        ("CV Ejecucion", f"Coeficiente de variacion = std / |media| x 100. "
                         f"Umbral: > {UMBRAL_CV}%"),
        ("Salto max.",   f"Mayor variacion % entre 2 periodos consecutivos. "
                         f"Umbral: > {UMBRAL_SALTO_PCT}%"),
        ("Cumpl. max.",  f"Mayor valor de cumplimiento registrado. "
                         f"Umbral anomalo: > {UMBRAL_CUMPL_ALTO}%"),
        ("CV Meta",      f"Variabilidad de la meta a lo largo del tiempo. "
                         f"Umbral: > {UMBRAL_META_CV}%"),
        ("", ""),
        ("Generado",     datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for fila in leyenda:
        ws3.append(list(fila))

    # Headers de sección en azul, sin fill
    for ref in ("A1", "B1", "A7", "B7"):
        ws3[ref].font      = FONT_HEADER
        ws3[ref].fill      = NO_FILL
        ws3[ref].alignment = ALIN_CENTRO
        ws3[ref].border    = BORDER_HEADER

    # Filas de severidad con fuente de color
    SEV_FONTS_LEY = {"ALTA": FONT_ALTA, "MEDIA": FONT_MEDIA, "BAJA": FONT_BAJA, "NORMAL": FONT_GRIS}
    for fila in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
        for cell in fila:
            sev_font = SEV_FONTS_LEY.get(str(cell.value).strip().upper(), FONT_NORMAL)
            cell.font      = sev_font if cell.column == 1 else FONT_NORMAL
            cell.fill      = NO_FILL
            cell.alignment = ALIN_IZQ
            cell.border    = BORDER_THIN

    # ── Guardar ──────────────────────────────────────────────────
    out_path = OUTPUT_DIR / f"Analisis_Indicadores_{ts}.xlsx"
    wb.save(out_path)
    return out_path


# ── Main ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Análisis de inconsistencias por indicador")
    parser.add_argument("--fuente", choices=["df_api", "excel"], default="df_api",
                        help="Fuente de datos: df_api (estado pipeline) o excel (archivo de salida)")
    parser.add_argument("--solo-anomalias", action="store_true",
                        help="Incluir solo indicadores con al menos una anomalía")
    args = parser.parse_args()

    log("INFO", "Iniciando análisis de indicadores")
    t0 = time.time()

    try:
        df = cargar_datos(args.fuente)
        log("DATA", f"{len(df):,} filas cargadas | {df['ID'].nunique()} indicadores únicos")

        # Filtrar a filas con ejecución o meta (excluir totalmente vacíos)
        df_valido = df[df["Ejecucion"].notna() | df["Meta"].notna()].copy()
        log("DATA", f"{len(df_valido):,} filas con ejecución o meta")

        # Asegurar columnas necesarias
        for col in ["Nombre", "Periodicidad", "Ejecucion", "Meta", "Cumplimiento"]:
            if col not in df_valido.columns:
                df_valido[col] = None

        # ── Análisis por indicador ────────────────────────────────
        log("INFO", "Calculando métricas por indicador...")
        filas_resumen = []

        for (iid, nombre, perio), gdf in df_valido.groupby(
            ["ID", "Nombre", "Periodicidad"], sort=False
        ):
            gdf = gdf.sort_values("Fecha")
            metricas = analizar_indicador(gdf)
            filas_resumen.append({
                "ID": iid,
                "Nombre": nombre,
                "Periodicidad": perio,
                **metricas,
            })

        df_resumen = pd.DataFrame(filas_resumen)

        n_total    = len(df_resumen)
        n_anomalos = (df_resumen["N_anomalias"] > 0).sum()
        n_alta     = (df_resumen["Severidad"] == "Alta").sum()
        n_media    = (df_resumen["Severidad"] == "Media").sum()

        log("DATA", f"{n_total} indicadores evaluados | "
                    f"{n_anomalos} con anomalías | "
                    f"{n_alta} severidad Alta | {n_media} Media")

        if args.solo_anomalias:
            df_resumen = df_resumen[df_resumen["N_anomalias"] > 0]
            log("INFO", f"Filtro --solo-anomalias: {len(df_resumen)} indicadores")

        # ── Períodos anómalos ─────────────────────────────────────
        log("INFO", "Detectando períodos anómalos específicos...")
        df_detalle = detectar_periodos_anomalos(df_valido)
        log("DATA", f"{len(df_detalle)} períodos anómalos identificados")

        # ── Exportar Excel ────────────────────────────────────────
        ts_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        log("INFO", "Generando Excel...")
        out_path = construir_excel(df_resumen, df_detalle, ts_str)

        elapsed = round(time.time() - t0, 2)
        log("OK", f"Archivo generado en {elapsed}s")
        log("OK", f"Archivo: {out_path.relative_to(_ROOT)}")

        if n_alta > 0:
            print()
            log("WARN", f"{n_alta} indicadores con severidad ALTA:")
            top = df_resumen[df_resumen["Severidad"] == "Alta"].nlargest(5, "_severidad_score")
            for _, r in top.iterrows():
                print(f"           ID {int(r['ID']):>4}  {r['Nombre'][:50]:<50}  | {r['Anomalias']}")
        print()
        sys.exit(0)

    except FileNotFoundError as e:
        log("ERROR", str(e))
        sys.exit(1)
    except Exception as e:
        log("ERROR", f"Error inesperado: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
