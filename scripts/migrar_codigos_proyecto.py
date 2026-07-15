"""Migración de una sola pasada: recodifica IDs de proyecto (no-Kawak) a PRY-N.

Ámbito: los 44 IDs marcados Proyecto=1 en 'Indicadores por CMI.xlsx' que no
tienen fila propia en ninguna fuente Kawak cruda (verificado por búsqueda en
data/raw/Kawak/*.xlsx y data/raw/Fuentes Consolidadas/*.xlsx).

Reescribe en sitio, con backup .bak previo:
  - data/raw/Indicadores por CMI.xlsx        (hoja 'Worksheet', columna Id)
  - data/raw/Resultados_Consolidados_Fuente.xlsx
      hojas: Consolidado Historico, Consolidado Semestral, Consolidado Cierres,
             Catalogo Indicadores, Variables, Desglose Series
      columnas: Id (todas), LLAVE/Llave (recalculada con el nuevo Id)

Uso: python scripts/migrar_codigos_proyecto.py
"""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent

MAPEO: dict[str, str] = {
    "1": "PRY-1", "2": "PRY-2", "3": "PRY-3", "4": "PRY-4", "5": "PRY-5",
    "6": "PRY-6", "7": "PRY-7", "8": "PRY-8", "9": "PRY-9", "10": "PRY-10",
    "10.1": "PRY-11", "11": "PRY-12", "12": "PRY-13", "13": "PRY-14",
    "13.1": "PRY-15", "900": "PRY-16", "901": "PRY-17", "902": "PRY-18",
    "903": "PRY-19", "904": "PRY-20", "905": "PRY-21", "906": "PRY-22",
    "907": "PRY-23", "908": "PRY-24", "909": "PRY-25", "910": "PRY-26",
    "911": "PRY-27", "912": "PRY-28", "913": "PRY-29", "914": "PRY-30",
    "915": "PRY-31", "916": "PRY-32", "917": "PRY-33", "918": "PRY-34",
    "919": "PRY-35", "920": "PRY-36", "921": "PRY-37", "922": "PRY-38",
    "923": "PRY-39", "924": "PRY-40", "925": "PRY-41", "926": "PRY-42",
    "927": "PRY-43", "928": "PRY-44",
}


def _id_str(val) -> str:
    """Misma normalización que scripts/etl/normalizacion.py::_id_str."""
    s = str(val)
    return s[:-2] if s.endswith(".0") else s


def _backup(path: Path) -> None:
    bak = path.with_name(path.stem + ".bak_pre_pry" + path.suffix)
    if not bak.exists():
        shutil.copy2(path, bak)
        print(f"  backup -> {bak.name}")


def migrar_indicadores_cmi() -> int:
    path = ROOT / "data" / "raw" / "Indicadores por CMI.xlsx"
    _backup(path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Worksheet"]

    header = {cell.value: cell.column for cell in ws[1]}
    col_id = header["Id"]

    n = 0
    for row in ws.iter_rows(min_row=2):
        cell = row[col_id - 1]
        id_s = _id_str(cell.value) if cell.value is not None else None
        if id_s in MAPEO:
            cell.value = MAPEO[id_s]
            n += 1

    wb.save(path)
    print(f"  Indicadores por CMI.xlsx: {n} filas recodificadas")
    return n


def migrar_hojas(path: Path, sheet_names: list[str]) -> int:
    _backup(path)
    wb = openpyxl.load_workbook(path)

    total = 0
    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header = {cell.value: cell.column for cell in ws[1]}
        if "Id" not in header:
            continue
        col_id = header["Id"]
        col_llave = header.get("LLAVE") or header.get("Llave")

        n = 0
        for row in ws.iter_rows(min_row=2):
            cell_id = row[col_id - 1]
            id_s = _id_str(cell_id.value) if cell_id.value is not None else None
            if id_s not in MAPEO:
                continue
            nuevo_id = MAPEO[id_s]

            if col_llave:
                cell_llave = row[col_llave - 1]
                if cell_llave.value and isinstance(cell_llave.value, str) and cell_llave.value.startswith(f"{id_s}-"):
                    cell_llave.value = nuevo_id + cell_llave.value[len(id_s):]

            cell_id.value = nuevo_id
            n += 1

        print(f"  {sheet_name}: {n} filas recodificadas")
        total += n

    wb.save(path)
    return total


if __name__ == "__main__":
    print("Migrando 'Indicadores por CMI.xlsx'...")
    n1 = migrar_indicadores_cmi()

    print("Migrando 'Resultados_Consolidados_Fuente.xlsx'...")
    n2 = migrar_hojas(
        ROOT / "data" / "raw" / "Resultados_Consolidados_Fuente.xlsx",
        ["Consolidado Historico", "Consolidado Semestral",
         "Consolidado Cierres", "Catalogo Indicadores",
         "Variables", "Desglose Series"],
    )

    print("Migrando 'data/output/Resultados Consolidados.xlsx'...")
    n3 = migrar_hojas(
        ROOT / "data" / "output" / "Resultados Consolidados.xlsx",
        ["Consolidado Historico", "Consolidado Semestral",
         "Consolidado Cierres", "Catalogo Indicadores",
         "Variables", "Desglose Series"],
    )

    print(f"\nTotal filas recodificadas: {n1 + n2 + n3}")
