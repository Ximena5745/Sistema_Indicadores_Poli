"""Agrega filas de catálogo CMI para los 10 proyectos nuevos del indicador
padre 603 (cronograma 2026) que aún no tenían fila en 'Indicadores por CMI.xlsx'.

Valores genéricos temporales (Clasificación, Periodicidad, Sentido, etc.)
tomados del patrón consistente de las 44 filas de proyecto existentes.
Subproceso / Linea / Objetivo / Meta Estratégica quedan en blanco — pendientes
de que el equipo de estrategia los complete manualmente en el Excel.

Uso: python scripts/agregar_proyectos_cmi.py
"""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
PATH = ROOT / "data" / "raw" / "Indicadores por CMI.xlsx"

NUEVOS = [
    ("PRY-45", "Portal de Talento Humano"),
    ("PRY-46", "Proyecto E2E - ILUMNO"),
    ("PRY-47", "Homologaciones IA - ILUMNO"),
    ("PRY-48", "Modelo Predictivo de deserción basado en Scoring"),
    ("PRY-49", "Voice IA - Recupero"),
    ("PRY-50", "Whatsapp IA"),
    ("PRY-51", "Adecuaciones campus principal"),
    ("PRY-52", "Implementación reforma Curricular"),
    ("PRY-53", "Optimización SEO AEO Captación"),
    ("PRY-54", "Sistematización proceso de la ORII"),
]


def main() -> None:
    bak = PATH.with_name(PATH.stem + ".bak_pre_pry45_54" + PATH.suffix)
    if not bak.exists():
        shutil.copy2(PATH, bak)
        print(f"backup -> {bak.name}")

    wb = openpyxl.load_workbook(PATH)
    ws = wb["Worksheet"]
    header = {cell.value: cell.column for cell in ws[1]}

    existentes = {
        str(row[header["Id"] - 1].value)
        for row in ws.iter_rows(min_row=2)
        if row[header["Id"] - 1].value is not None
    }

    def _num(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0

    orden_cmi = max(
        _num(row[header["Orden CMI"] - 1].value)
        for row in ws.iter_rows(min_row=2)
    )

    n = 0
    for id_s, nombre in NUEVOS:
        if id_s in existentes:
            print(f"  {id_s} ya existe, se omite")
            continue
        orden_cmi += 1
        fila = {
            "Id": id_s,
            "Indicador": nombre,
            "Clasificación": "Estratégico",
            "Periodicidad": "Mensual",
            "Sentido": "Positivo",
            "Indicadores Clave": 1,
            "Indicadores Plan estrategico": 1,
            "Indicadores Vicerrectoria": 1,
            "Subprocesos": 1,
            "General": 1,
            "Orden CMI": orden_cmi,
            "Plan anual": 0,
            "Proyecto": 1,
            "Tipo de indicador": "Resultado",
            "Ind act": 0,
        }
        nueva_fila = [None] * len(header)
        for col_name, col_idx in header.items():
            if col_name in fila:
                nueva_fila[col_idx - 1] = fila[col_name]
        ws.append(nueva_fila)
        n += 1
        print(f"  {id_s} -> {nombre}")

    wb.save(PATH)
    print(f"\nTotal filas agregadas: {n}")


if __name__ == "__main__":
    main()
