"""Construye 'data/raw/Catalogo de Indicadores.xlsx' desde cero — directorio
maestro único de indicadores, fusionando:
  - data/raw/Resultados_Consolidados_Fuente.xlsx (hoja 'Catalogo Indicadores':
    universo Kawak/API + campos manuales TipoCalculo/Asociacion/Formato_Valores)
  - data/raw/Indicadores por CMI.xlsx (hoja 'Worksheet': clasificación de negocio)
  - data/raw/Ficha_Tecnica_Indicadores.xlsx (ficha técnica; solo se fusiona la
    información de indicadores con Estado_Indicador en {Stand by, Activo})

Reglas aplicadas (feedback 2026-07-14):
  - NO incluir Meta_General (redundante con Meta_Estrategica).
  - NO incluir Deficiente/Alerta/Satisfactorio/Sobresaliente en 'Catalogo
    Indicadores' (cambian por año); sí se conservan en 'Ficha Tecnica Detalle'.
  - Solo se fusiona información de Ficha Técnica para indicadores con
    Estado_Indicador en {'Stand by', 'Activo'} (no 'Inactivo' ni vacío). Un
    indicador sin ficha vigente igual aparece en el catálogo si tiene datos
    de Kawak/API o CMI, solo que sin las columnas de ficha pobladas.
  - Columnas agrupadas por color de encabezado: azul (Kawak/API/Resultados
    Consolidados), verde (Clasificación CMI), naranja (Ficha Técnica).

Archivo NUEVO y separado — Resultados_Consolidados_Fuente.xlsx conserva sus
hojas de datos históricos (Consolidado Historico/Semestral/Cierres/Logica/
Variables/Desglose Series) sin cambios.

Uso: python scripts/construir_catalogo_indicadores.py
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
CAT_FUENTE = ROOT / "data" / "raw" / "Resultados_Consolidados_Fuente.xlsx"
CMI_FILE = ROOT / "data" / "raw" / "Indicadores por CMI.xlsx"
FICHA_FILE = ROOT / "data" / "raw" / "Ficha_Tecnica_Indicadores.xlsx"
OUT_FILE = ROOT / "data" / "raw" / "Catalogo de Indicadores.xlsx"

ESTADOS_VALIDOS = {"stand by", "activo"}

COLS_BASE = [
    "Id", "Indicador", "Clasificacion", "Proceso", "Periodicidad", "Sentido",
    "Tipo_API", "Estado", "Fuente", "Extraccion", "TipoCalculo",
    "Asociacion", "Formato_Valores", "Tipo de indicador",
]
COLS_CMI = [
    "Subproceso", "Linea_Estrategica", "Objetivo_Estrategico",
    "Meta_Estrategica", "Proyecto", "Orden_CMI", "Plan_Anual",
    "Factor", "Caracteristica", "Indicadores_Clave",
    "Indicadores_Plan_Estrategico", "Indicadores_Vicerrectoria",
    "Subprocesos", "General", "Ind_Act",
]
COLS_FICHA = [
    "Formula", "Descripcion", "Unidad", "Responsable_Calculo",
    "Responsable_Analisis", "ID_Kawak", "CNA_SNIES",
]

FILL_BASE = PatternFill("solid", fgColor="BDD7EE")   # azul — Kawak/API/Resultados
FILL_CMI = PatternFill("solid", fgColor="C6E0B4")     # verde — Clasificación CMI
FILL_FICHA = PatternFill("solid", fgColor="FCE4D6")   # naranja — Ficha Técnica
HEADER_FONT = Font(bold=True)


def _id_str(v) -> str:
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def main() -> None:
    df_cat = pd.read_excel(CAT_FUENTE, sheet_name="Catalogo Indicadores")
    df_cmi = pd.read_excel(CMI_FILE, sheet_name="Worksheet")
    df_ficha = pd.read_excel(FICHA_FILE, sheet_name=0)

    df_cat["_id"] = df_cat["Id"].map(_id_str)
    df_cmi["_id"] = df_cmi["Id"].map(_id_str)
    df_ficha["_id"] = df_ficha["Id Ind"].map(_id_str)

    df_cmi = df_cmi.drop_duplicates("_id", keep="last").set_index("_id")

    estado_col = df_ficha.get("Estado_Indicador")
    ficha_vigente_mask = (
        estado_col.astype(str).str.strip().str.lower().isin(ESTADOS_VALIDOS)
        if estado_col is not None else pd.Series(False, index=df_ficha.index)
    )
    n_ficha_total = len(df_ficha)
    n_ficha_vigente = int(ficha_vigente_mask.sum())
    df_ficha_vigente = df_ficha[ficha_vigente_mask].drop_duplicates("_id", keep="last").set_index("_id")
    print(f"Ficha Técnica: {n_ficha_total} filas -> {n_ficha_vigente} vigentes (Stand by/Activo)")

    n_antes = len(df_cat)
    ids_union = set(df_cat["_id"]) | set(df_cmi.index) | set(df_ficha_vigente.index)
    ids_nuevos = sorted(ids_union - set(df_cat["_id"]))
    print(f"Catalogo base: {n_antes} filas. Ids nuevos (solo en CMI/Ficha vigente): {len(ids_nuevos)}")

    filas_nuevas = []
    for id_s in ids_nuevos:
        cmi_row = df_cmi.loc[id_s] if id_s in df_cmi.index else None
        ficha_row = df_ficha_vigente.loc[id_s] if id_s in df_ficha_vigente.index else None
        nombre = ""
        if cmi_row is not None and pd.notna(cmi_row.get("Indicador")):
            nombre = str(cmi_row["Indicador"])
        elif ficha_row is not None and pd.notna(ficha_row.get("Nombre del indicador")):
            nombre = str(ficha_row["Nombre del indicador"])
        fila = {c: None for c in COLS_BASE}
        fila["Id"] = id_s
        fila["Indicador"] = nombre
        fila["Fuente"] = "CMI/Ficha (sin Kawak)"
        filas_nuevas.append(fila)

    if filas_nuevas:
        df_cat = pd.concat([df_cat.drop(columns=["_id"]), pd.DataFrame(filas_nuevas)], ignore_index=True)
        df_cat["_id"] = df_cat["Id"].map(_id_str)

    for col in COLS_CMI + COLS_FICHA:
        df_cat[col] = None

    n_con_ficha = 0
    for i, row in df_cat.iterrows():
        id_s = row["_id"]
        cmi_row = df_cmi.loc[id_s] if id_s in df_cmi.index else None
        ficha_row = df_ficha_vigente.loc[id_s] if id_s in df_ficha_vigente.index else None

        def cmi_val(col):
            return cmi_row.get(col) if cmi_row is not None and pd.notna(cmi_row.get(col)) else None

        def ficha_val(col):
            return ficha_row.get(col) if ficha_row is not None and pd.notna(ficha_row.get(col)) else None

        df_cat.at[i, "Subproceso"] = cmi_val("Subproceso")
        df_cat.at[i, "Linea_Estrategica"] = cmi_val("Linea") or ficha_val("Linea_Estrategica")
        df_cat.at[i, "Objetivo_Estrategico"] = cmi_val("Objetivo") or ficha_val("Objetivo_Estrategico")
        df_cat.at[i, "Meta_Estrategica"] = cmi_val("Meta Estratégica")
        df_cat.at[i, "Proyecto"] = cmi_val("Proyecto")
        df_cat.at[i, "Orden_CMI"] = cmi_val("Orden CMI")
        df_cat.at[i, "Plan_Anual"] = cmi_val("Plan anual")
        df_cat.at[i, "Factor"] = cmi_val("FACTOR")
        df_cat.at[i, "Caracteristica"] = cmi_val("CARACTERÍSTICA")
        df_cat.at[i, "Indicadores_Clave"] = cmi_val("Indicadores Clave")
        df_cat.at[i, "Indicadores_Plan_Estrategico"] = cmi_val("Indicadores Plan estrategico")
        df_cat.at[i, "Indicadores_Vicerrectoria"] = cmi_val("Indicadores Vicerrectoria")
        df_cat.at[i, "Subprocesos"] = cmi_val("Subprocesos")
        df_cat.at[i, "General"] = cmi_val("General")
        df_cat.at[i, "Ind_Act"] = cmi_val("Ind act")

        if ficha_row is not None:
            n_con_ficha += 1
            df_cat.at[i, "Formula"] = ficha_val("Formula")
            df_cat.at[i, "Descripcion"] = ficha_val("Descripción del indicador")
            df_cat.at[i, "Unidad"] = ficha_val("Unidad")
            df_cat.at[i, "Responsable_Calculo"] = ficha_val("Responsable del calculo")
            df_cat.at[i, "Responsable_Analisis"] = ficha_val("Responsable del analisis")
            df_cat.at[i, "ID_Kawak"] = ficha_val("ID Kawak")
            df_cat.at[i, "CNA_SNIES"] = cmi_val("CNA") or ficha_val("CNA/SNIES")
        else:
            df_cat.at[i, "CNA_SNIES"] = cmi_val("CNA")

    df_cat_final = df_cat.drop(columns=["_id"])[COLS_BASE + COLS_CMI + COLS_FICHA]
    n_despues = len(df_cat_final)
    print(f"Catalogo Indicadores: {n_antes} -> {n_despues} filas | {len(df_cat_final.columns)} columnas")
    print(f"  Filas con datos de Ficha Técnica vigente: {n_con_ficha}")

    # ── Hoja "Ficha Tecnica Detalle" — solo filas vigentes (Stand by/Activo) ──
    df_ficha_detalle = df_ficha_vigente.reset_index().rename(columns={"_id": "Id"})
    cols_orden = ["Id"] + [c for c in df_ficha_detalle.columns if c != "Id"]
    df_ficha_detalle = df_ficha_detalle[cols_orden]

    # ── Escribir archivo nuevo con formato por grupo de color ──────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def _escribir_hoja(nombre, df, grupos=None):
        ws = wb.create_sheet(nombre)
        for j, col in enumerate(df.columns, 1):
            cell = ws.cell(1, j, col)
            cell.font = HEADER_FONT
            if grupos:
                for cols, fill in grupos:
                    if col in cols:
                        cell.fill = fill
                        break
        for i, (_, row) in enumerate(df.iterrows(), 2):
            for j, col in enumerate(df.columns, 1):
                val = row[col]
                if isinstance(val, pd.Timestamp):
                    val = val.to_pydatetime().date()
                elif isinstance(val, float) and pd.isna(val):
                    val = None
                ws.cell(i, j).value = val
        ws.freeze_panes = "A2"

    _escribir_hoja(
        "Catalogo Indicadores", df_cat_final,
        grupos=[(COLS_BASE, FILL_BASE), (COLS_CMI, FILL_CMI), (COLS_FICHA, FILL_FICHA)],
    )
    _escribir_hoja("Ficha Tecnica Detalle", df_ficha_detalle)

    wb.save(OUT_FILE)
    print(f"\nGuardado: {OUT_FILE}")
    print(f"  Catalogo Indicadores: {n_despues} filas, {len(df_cat_final.columns)} columnas")
    print(f"  Ficha Tecnica Detalle: {len(df_ficha_detalle)} filas, {len(df_ficha_detalle.columns)} columnas")


if __name__ == "__main__":
    main()
