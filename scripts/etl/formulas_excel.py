"""
scripts/etl/formulas_excel.py
Utilidades openpyxl: mapa de columnas, fórmulas Excel, materialización.

IMPORTANTE:
  openpyxl NO evalúa ni ajusta referencias de fórmulas al eliminar/insertar
  filas.  Por eso _reescribir_formulas() debe ejecutarse DESPUÉS de toda
  operación de borrado.
"""
from __future__ import annotations

import logging
from typing import Any, Dict, Optional

import numpy as np
import pandas as pd
import openpyxl

from .config import IDS_PLAN_ANUAL, IDS_TOPE_100
from .cumplimiento import _calc_cumpl, _calc_cumpl_con_tope_dinamico
from .normalizacion import COL_ALIASES, MESES_ES, make_llave, _id_str

logger = logging.getLogger(__name__)


# ── Mapa de columnas ──────────────────────────────────────────────

def _build_col_map(ws) -> Dict[str, int]:
    """Lee el encabezado de la hoja → {campo_interno: col_idx (1-based)}."""
    cm: Dict[str, int] = {}
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if cell.value is not None:
            campo = COL_ALIASES.get(str(cell.value).strip())
            if campo:
                cm[campo] = cell.column
    return cm


# ── Generadores de fórmulas ───────────────────────────────────────

def formula_G(r: int) -> str: return f"=YEAR(F{r})"
def formula_H(r: int) -> str:
    # =NOMPROPIO(TEXTO(F2;"mmmm"))
    return f'=NOMPROPIO(TEXTO(F{r};"mmmm"))'

def formula_I(r: int) -> str:
    # =SI(O(H2="Enero";H2="Febrero";H2="Marzo";H2="Abril";H2="Mayo";H2="Junio");G2&"-1";SI(O(H2="Julio";H2="Agosto";H2="Septiembre";H2="Octubre";H2="Noviembre";H2="Diciembre");G2&"-2"))
    return (
        f'=SI(O(H{r}="Enero";H{r}="Febrero";H{r}="Marzo";'
        f'H{r}="Abril";H{r}="Mayo";H{r}="Junio");G{r}&"-1";'
        f'SI(O(H{r}="Julio";H{r}="Agosto";H{r}="Septiembre";'
        f'H{r}="Octubre";H{r}="Noviembre";H{r}="Diciembre");G{r}&"-2"))'
    )

def formula_L(r: int, tope: float = 1.3) -> str:
    return (
        f'=IFERROR(IF(OR(J{r}=0,K{r}=""),"",IF(E{r}="Positivo",'
        f'MIN(MAX(K{r}/J{r},0),{tope}),'
        f'MIN(MAX(J{r}/K{r},0),{tope}))),"") '
    )

def formula_M(r: int, tope: Optional[float] = None) -> str:
    return (
        f'=IFERROR(IF(OR(J{r}=0,K{r}=""),"",IF(E{r}="Positivo",'
        f'MAX(K{r}/J{r},0),'
        f'MAX(J{r}/K{r},0))),"") '
    )
def formula_R(r: int) -> str:
    # =A2&"-"&AÑO(F2)&"-"&SI(LARGO(MES(F2))=1;"0"&MES(F2);MES(F2))&"-"&DIA(F2)
    return (
        f'=A{r}&"-"&AÑO(F{r})&"-"'
        f'&SI(LARGO(MES(F{r}))=1;"0"&MES(F{r});MES(F{r}))'
        f'&"-"&DIA(F{r})'
    )


# ── Materialización ────────────────────────────────────────────────

def _materializar_cumplimiento(ws) -> None:
    """
    Recalcula Cumplimiento y Cumplimiento Real para TODAS las filas
    usando valores reales de Meta, Ejecucion, Sentido e Id de la misma fila.

    Necesario porque las fórmulas Excel pueden apuntar a filas incorrectas
    tras inserciones/borrados.
    
    PROBLEM #4 FIX: Ahora aplica toques dinámicos según IDS_PLAN_ANUAL e IDS_TOPE_100
    en lugar de usar tope fijo 1.3.
    """
    cm = _build_col_map(ws)
    idx_id      = cm.get("Id")
    idx_meta    = cm.get("Meta")
    idx_ejec    = cm.get("Ejecucion")
    idx_sentido = cm.get("Sentido")
    idx_cumpl   = cm.get("Cumplimiento")
    idx_cumplr  = cm.get("CumplReal")
    if not (idx_meta and idx_ejec and idx_sentido and idx_cumpl):
        return
    n_ok = n_vacio = 0
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value is None:
            continue
        meta    = row[idx_meta    - 1].value
        ejec    = row[idx_ejec    - 1].value
        sentido = row[idx_sentido - 1].value or "Positivo"
        id_ind  = row[idx_id - 1].value if idx_id else None  # Obtener ID para tope dinámico
        
        if (isinstance(meta, str) and meta.startswith("=")) or \
           (isinstance(ejec, str) and ejec.startswith("=")):
            continue
        
        # Usar función con tope dinámico según ID
        cumpl_capped, cumpl_real = _calc_cumpl_con_tope_dinamico(
            meta, ejec, str(sentido),
            id_indicador=id_ind,
            ids_plan_anual=IDS_PLAN_ANUAL,
            ids_tope_100=IDS_TOPE_100
        )
        
        c_cumpl = row[idx_cumpl - 1]
        c_cumpl.value = cumpl_capped
        if cumpl_capped is not None:
            c_cumpl.number_format = "0.00%"
        if idx_cumplr:
            c_cumplr = row[idx_cumplr - 1]
            c_cumplr.value = cumpl_real
            if cumpl_real is not None:
                c_cumplr.number_format = "0.00%"
        if cumpl_capped is not None:
            n_ok += 1
        else:
            n_vacio += 1
    logger.info(f"    [{ws.title}] Cumplimiento recalculado: {n_ok} con valor, {n_vacio} vacíos.")


def _materializar_formula_año(ws) -> None:
    """
    Reemplaza fórmulas en Año, Mes, Semestre y LLAVE por valores calculados
    desde Fecha e Id (openpyxl las lee como strings si no están evaluadas).
    """
    cm = _build_col_map(ws)
    idx_fecha    = cm.get("Fecha")
    idx_anio     = cm.get("Anio")
    idx_mes      = cm.get("Mes")
    idx_semestre = cm.get("Semestre")
    idx_llave    = cm.get("LLAVE")
    idx_id       = cm.get("Id")
    if not idx_fecha:
        return
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value is None:
            continue
        celda_fecha = row[idx_fecha - 1]
        try:
            fecha = pd.to_datetime(celda_fecha.value)
        except Exception:
            continue
        if idx_anio:
            c = row[idx_anio - 1]
            if isinstance(c.value, str) and c.value.startswith("="):
                c.value = fecha.year
        if idx_mes:
            c = row[idx_mes - 1]
            if isinstance(c.value, str) and c.value.startswith("="):
                c.value = MESES_ES.get(fecha.month, "")
        if idx_semestre:
            c = row[idx_semestre - 1]
            if isinstance(c.value, str) and c.value.startswith("="):
                c.value = f"{fecha.year}-{1 if fecha.month <= 6 else 2}"
        if idx_llave and idx_id:
            c = row[idx_llave - 1]
            if isinstance(c.value, str) and c.value.startswith("="):
                id_val = row[idx_id - 1].value
                c.value = make_llave(id_val, fecha)


def _ensure_tipo_registro_header(ws) -> None:
    """Agrega header 'Tipo_Registro' al final si no existe."""
    header_row = list(next(ws.iter_rows(min_row=1, max_row=1)))
    existing   = {str(c.value).strip() for c in header_row if c.value is not None}
    if "Tipo_Registro" not in existing:
        last_col = max((c.column for c in header_row if c.value is not None), default=0)
        ws.cell(1, last_col + 1).value = "Tipo_Registro"


# ── Validación de alineación de columnas ──────────────────────────

def _validar_col_formulas(cm: Dict[str, int], nombre_hoja: str = "") -> None:
    """
    Verifica que las columnas usadas por las fórmulas Excel coincidan
    con las posiciones esperadas. Lanza ValueError si hay desalineación.
    """
    esperado = {
        "Id": 1, "Sentido": 5, "Fecha": 6, "Anio": 7, "Mes": 8,
        "Semestre": 9, "Meta": 10, "Ejecucion": 11,
        "Cumplimiento": 12, "CumplReal": 13, "LLAVE": 18,
    }
    errores = []
    for campo, col_esperada in esperado.items():
        col_real = cm.get(campo)
        if col_real is not None and col_real != col_esperada:
            errores.append(f"    {campo}: esperada={col_esperada}, real={col_real}")
    if errores:
        raise ValueError(
            f"  [ERROR] Columnas desalineadas en [{nombre_hoja}]:\n"
            + "\n".join(errores)
        )


# ── Reescritura de fórmulas (CRÍTICO) ────────────────────────────

def _reescribir_formulas(ws) -> None:
    """
    Reescribe las fórmulas de Año/Mes/Periodo/Cumplimiento/CumplReal/LLAVE
    con el número de fila ACTUAL.

    OBLIGATORIO ejecutar después de TODA eliminación/inserción de filas.
    """
    cm = _build_col_map(ws)
    idx_id      = cm.get("Id")
    idx_anio    = cm.get("Anio")
    idx_mes     = cm.get("Mes")
    idx_sem     = cm.get("Semestre")
    idx_cumpl   = cm.get("Cumplimiento")
    idx_cumplr  = cm.get("CumplReal")
    idx_llave   = cm.get("LLAVE")
    idx_tiporeg = cm.get("TipoRegistro")

    n = 0
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value is None:
            continue
        r = row[0].row

        id_val = _id_str(row[idx_id - 1].value) if idx_id else None
        tope = 1.0 if id_val in IDS_PLAN_ANUAL or id_val in IDS_TOPE_100 else 1.3

        tipo_reg_val  = row[idx_tiporeg - 1].value if idx_tiporeg else None
        es_metrica = str(tipo_reg_val or "").strip().lower() == "metrica"

        meta_val = row[cm.get("Meta")-1].value if cm.get("Meta") else None
        ejec_val = row[cm.get("Ejecucion")-1].value if cm.get("Ejecucion") else None

        if idx_anio:
            ws.cell(r, idx_anio).value = formula_G(r)
        if idx_mes:
            ws.cell(r, idx_mes).value = formula_H(r)
        if idx_sem:
            ws.cell(r, idx_sem).value = formula_I(r)
        if idx_cumpl:
            c = ws.cell(r, idx_cumpl)
            if es_metrica:
                c.value = None
            elif meta_val is not None and ejec_val is not None:
                c.value = formula_L(r, tope=tope)
                c.number_format = "0.00%"
            else:
                c.value = None
        if idx_cumplr:
            c = ws.cell(r, idx_cumplr)
            if es_metrica:
                c.value = None
            elif meta_val is not None and ejec_val is not None:
                c.value = formula_M(r)
                c.number_format = "0.00%"
            else:
                c.value = None
        if idx_llave:
            ws.cell(r, idx_llave).value = formula_R(r)
        n += 1

    logger.info(f"    [{ws.title}] Fórmulas reescritas en {n:,} filas.")
